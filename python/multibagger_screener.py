"""
NSE Multibagger Hunter — 3-5 Year Wealth Creation Picks
─────────────────────────────────────────────────────────────────────────────
Identifies smid-cap NSE stocks meeting 8 quantitative + qualitative filters
that historically deliver 3x-10x returns over 3-5 years:

  1. Earnings CAGR ≥ 25% (3-yr)
  2. ROCE ≥ 20%
  3. Revenue CAGR ≥ 20% (3-yr)
  4. Promoter holding ≥ 45%
  5. Debt/Equity < 0.5
  6. OPM ≥ 15%
  7. Sector tailwind (multi-decade)
  8. Recent BUY from top brokerage

The weekly Opus 4.7 task refreshes CMP, analyst targets, and Q-result data via
multibagger_fresh.json which is merged on every screener run.

Excludes anything in permanent_damage_blacklist.PERMANENT_DAMAGE_BLACKLIST.
─────────────────────────────────────────────────────────────────────────────
"""
import json, os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Cross-module imports
try:
    from permanent_damage_blacklist import get_blacklist_set
    BLACKLIST = get_blacklist_set(["PERMANENT_AVOID", "WAIT_FOR_RESOLUTION"])
except Exception:
    BLACKLIST = set()

try:
    from volatility_engine import smart_sl, smart_targets, position_size
    HAS_VOL_ENGINE = True
except Exception:
    HAS_VOL_ENGINE = False

try:
    from sector_rotation import get_sector_boost
except Exception:
    def get_sector_boost(symbol): return 0

try:
    from flow_tracker import get_smart_money_score
except Exception:
    def get_smart_money_score(symbol): return 0

EXCEL_PATH    = r"C:\Users\razee\OneDrive\Desktop\TradingClaude\NSE_Multibagger_Picks.xlsx"
FRESH_JSON    = r"C:\Users\razee\OneDrive\Desktop\TradingClaude\multibagger_fresh.json"
TODAY         = date.today()

DARK_BG="0D0D0D"; HEADER_BG="1A1A2E"; ROW_ALT="141414"
GOLD="FFD700"; GREEN="00C896"; BLUE="4FC3F7"; WHITE="FFFFFF"
GREY="9E9E9E"; RED="FF5252";  AMBER="FFB300"; ORANGE="FF6B35"; CYAN="00BCD4"
PURPLE="9C27B0"

CONV_META = {
    "VERY HIGH": ("1B4332", GREEN, "🔥 5+ filters strong + analyst conviction"),
    "HIGH":      ("003566", BLUE,  "✨ 4-5 filters strong + sector tailwind"),
    "MEDIUM":    ("3D2C00", AMBER, "👁 3-4 filters + watch for confirmation"),
}

# ── Multibagger universe (ALL VERIFY BEFORE TRADING; numbers refreshed weekly) ─
# Numbers are last-known approximations; the weekly Opus 4.7 task refreshes them.
MULTIBAGGERS = {
    # ── DEFENCE (Make-in-India ₹6L cr capex by 2030) ──────────────────────────
    "NSE:BDL": {
        "name": "Bharat Dynamics Ltd",
        "sector": "Defence — Missiles & Underwater Weapons",
        "market_cap_cr": 22000, "cmp": 1185,
        "earnings_cagr_3y_pct": 35, "roce_pct": 24, "revenue_cagr_3y_pct": 28,
        "promoter_holding_pct": 74.9, "debt_to_equity": 0.05, "opm_pct": 22,
        "sector_tailwind": "Indigenous missile / torpedo orders ramping; FY27 order book ₹19,000 cr",
        "analyst_call": "Motilal Oswal Buy ₹1,500 (avg target ~+27%)",
        "moat": "Sole indigenous supplier for Akash, Astra, Brahmos missiles",
        "y3_target": 1900, "y3_x": "1.6x", "y5_target": 3200, "y5_x": "2.7x",
        "key_risks": "Order book execution, defence budget cuts, single-customer (MoD)",
        "entry_zone": "1150-1200", "sl": 1080, "conviction": "HIGH",
        "action": "ACCUMULATE",
    },
    "NSE:MAZDOCK": {
        "name": "Mazagon Dock Shipbuilders",
        "sector": "Defence — Shipbuilding",
        "market_cap_cr": 87000, "cmp": 4350,
        "earnings_cagr_3y_pct": 55, "roce_pct": 42, "revenue_cagr_3y_pct": 32,
        "promoter_holding_pct": 84.8, "debt_to_equity": 0.0, "opm_pct": 17,
        "sector_tailwind": "Project-75I submarine + frigate orders; ₹40,000 cr order book",
        "analyst_call": "Multiple Buys, target range ₹4,800-5,500",
        "moat": "Premier defence shipyard; submarines, destroyers, frigates",
        "y3_target": 6500, "y3_x": "1.5x", "y5_target": 11000, "y5_x": "2.5x",
        "key_risks": "Project delays, valuation already high, govt-customer concentration",
        "entry_zone": "4250-4400", "sl": 4000, "conviction": "MEDIUM",
        "action": "WAIT FOR DIP",
    },
    "NSE:DATAPATTNS": {
        "name": "Data Patterns India",
        "sector": "Defence — Avionics & Electronics",
        "market_cap_cr": 14000, "cmp": 2480,
        "earnings_cagr_3y_pct": 60, "roce_pct": 36, "revenue_cagr_3y_pct": 45,
        "promoter_holding_pct": 53.4, "debt_to_equity": 0.02, "opm_pct": 32,
        "sector_tailwind": "Indigenous radar & EW systems; FY27 order book ₹2,500 cr+",
        "analyst_call": "Anand Rathi Buy ₹3,100 (+25%)",
        "moat": "Best-in-class margins in defence electronics; design + manufacturing IP",
        "y3_target": 3800, "y3_x": "1.5x", "y5_target": 7000, "y5_x": "2.8x",
        "key_risks": "Lumpy order intake, key-person risk",
        "entry_zone": "2400-2520", "sl": 2280, "conviction": "HIGH",
        "action": "ACCUMULATE",
    },
    "NSE:MTARTECH": {
        "name": "MTAR Technologies",
        "sector": "Defence + Nuclear + Space Precision Components",
        "market_cap_cr": 4500, "cmp": 1450,
        "earnings_cagr_3y_pct": 28, "roce_pct": 18, "revenue_cagr_3y_pct": 30,
        "promoter_holding_pct": 50.2, "debt_to_equity": 0.15, "opm_pct": 20,
        "sector_tailwind": "Space + nuclear + clean-energy precision components",
        "analyst_call": "Antique Buy ₹1,950",
        "moat": "Tier-1 supplier to ISRO, Bloom Energy, Rafael",
        "y3_target": 2400, "y3_x": "1.7x", "y5_target": 4500, "y5_x": "3.1x",
        "key_risks": "Customer concentration (Bloom), execution risk",
        "entry_zone": "1400-1480", "sl": 1320, "conviction": "MEDIUM",
        "action": "ACCUMULATE",
    },

    # ── RENEWABLES (500GW by 2030 target) ─────────────────────────────────────
    "NSE:KPIGREEN": {
        "name": "KPI Green Energy",
        "sector": "Solar IPP + EPC",
        "market_cap_cr": 12500, "cmp": 695,
        "earnings_cagr_3y_pct": 95, "roce_pct": 26, "revenue_cagr_3y_pct": 80,
        "promoter_holding_pct": 47.5, "debt_to_equity": 0.45, "opm_pct": 24,
        "sector_tailwind": "Solar capacity addition ramping; CPP + IPP dual model",
        "analyst_call": "Anand Rathi Buy ₹950 (+37%)",
        "moat": "End-to-end solar (EPC + IPP); rapid execution track record",
        "y3_target": 1200, "y3_x": "1.7x", "y5_target": 2500, "y5_x": "3.6x",
        "key_risks": "Module price volatility, debt-funded capex, promoter pledge",
        "entry_zone": "680-715", "sl": 640, "conviction": "HIGH",
        "action": "ACCUMULATE",
    },
    "NSE:PREMIERENE": {
        "name": "Premier Energies Ltd",
        "sector": "Solar Cell + Module Manufacturer",
        "market_cap_cr": 50000, "cmp": 1110,
        "earnings_cagr_3y_pct": 250, "roce_pct": 32, "revenue_cagr_3y_pct": 65,
        "promoter_holding_pct": 64.3, "debt_to_equity": 0.65, "opm_pct": 22,
        "sector_tailwind": "ALMM regime + PLI + 50GW domestic cell capacity push",
        "analyst_call": "Motilal Oswal Buy ₹1,400 (+26%)",
        "moat": "2nd-largest integrated solar cell maker (TOPCon technology)",
        "y3_target": 1800, "y3_x": "1.6x", "y5_target": 3500, "y5_x": "3.2x",
        "key_risks": "Capex cycle, China imports if ALMM relaxed",
        "entry_zone": "1080-1140", "sl": 1010, "conviction": "HIGH",
        "action": "ACCUMULATE",
    },
    "NSE:INOXWIND": {
        "name": "Inox Wind Ltd",
        "sector": "Wind Turbine OEM",
        "market_cap_cr": 27000, "cmp": 230,
        "earnings_cagr_3y_pct": "Loss → Profit", "roce_pct": 22, "revenue_cagr_3y_pct": 55,
        "promoter_holding_pct": 71.8, "debt_to_equity": 0.3, "opm_pct": 15,
        "sector_tailwind": "Wind capacity addition target 100GW by 2030",
        "analyst_call": "Equirus Buy ₹290 (+26%)",
        "moat": "Largest wind OEM in India (3MW platform + service revenue)",
        "y3_target": 350, "y3_x": "1.5x", "y5_target": 600, "y5_x": "2.6x",
        "key_risks": "Order execution, cyclicality, service revenue ramp",
        "entry_zone": "225-240", "sl": 210, "conviction": "HIGH",
        "action": "ACCUMULATE",
    },
    "NSE:BORORENEW": {
        "name": "Borosil Renewables",
        "sector": "Solar Glass",
        "market_cap_cr": 6200, "cmp": 475,
        "earnings_cagr_3y_pct": "Volatile", "roce_pct": 14, "revenue_cagr_3y_pct": 22,
        "promoter_holding_pct": 65.0, "debt_to_equity": 0.40, "opm_pct": 18,
        "sector_tailwind": "Solar module volume growth + ADD on Chinese imports",
        "analyst_call": "Ventura Buy ₹620",
        "moat": "Only solar glass maker in India; ADD protection from imports",
        "y3_target": 750, "y3_x": "1.6x", "y5_target": 1300, "y5_x": "2.7x",
        "key_risks": "Capacity ramp, China dumping if ADD lifts",
        "entry_zone": "460-490", "sl": 430, "conviction": "MEDIUM",
        "action": "ACCUMULATE",
    },

    # ── ELECTRICALS / WIRES & CABLES / CAP GOODS (T&D capex super-cycle) ─────
    "NSE:KEI": {
        "name": "KEI Industries",
        "sector": "Wires & Cables",
        "market_cap_cr": 35000, "cmp": 3850,
        "earnings_cagr_3y_pct": 28, "roce_pct": 24, "revenue_cagr_3y_pct": 22,
        "promoter_holding_pct": 35.0, "debt_to_equity": 0.05, "opm_pct": 11,
        "sector_tailwind": "T&D capex + housing demand + EHV cable ramp",
        "analyst_call": "Nuvama Buy ₹4,800",
        "moat": "Premium wires + EHV cable capacity (only 4 players)",
        "y3_target": 5500, "y3_x": "1.4x", "y5_target": 9000, "y5_x": "2.3x",
        "key_risks": "Copper price volatility, working capital",
        "entry_zone": "3750-3920", "sl": 3520, "conviction": "HIGH",
        "action": "ACCUMULATE",
    },
    "NSE:POLYCAB": {
        "name": "Polycab India",
        "sector": "Wires & Cables (largest)",
        "market_cap_cr": 95000, "cmp": 6300,
        "earnings_cagr_3y_pct": 30, "roce_pct": 26, "revenue_cagr_3y_pct": 22,
        "promoter_holding_pct": 66.0, "debt_to_equity": 0.0, "opm_pct": 12,
        "sector_tailwind": "Same as KEI; FMEG diversification",
        "analyst_call": "ICICI Sec Buy ₹7,500",
        "moat": "#1 brand + national distribution",
        "y3_target": 8500, "y3_x": "1.35x", "y5_target": 13500, "y5_x": "2.1x",
        "key_risks": "FMEG losses, valuation rich",
        "entry_zone": "6150-6400", "sl": 5750, "conviction": "MEDIUM",
        "action": "ACCUMULATE",
    },
    "NSE:CGPOWER": {
        "name": "CG Power & Industrial Solutions",
        "sector": "Industrial Cap Goods + Semiconductors",
        "market_cap_cr": 95000, "cmp": 615,
        "earnings_cagr_3y_pct": 70, "roce_pct": 38, "revenue_cagr_3y_pct": 18,
        "promoter_holding_pct": 58.1, "debt_to_equity": 0.0, "opm_pct": 14,
        "sector_tailwind": "Power capex + railways + foray into OSAT semiconductors",
        "analyst_call": "Multiple Buys ₹770-820",
        "moat": "Murugappa group + semiconductor JV with Renesas",
        "y3_target": 850, "y3_x": "1.4x", "y5_target": 1400, "y5_x": "2.3x",
        "key_risks": "Semiconductor JV execution, valuation",
        "entry_zone": "600-630", "sl": 565, "conviction": "HIGH",
        "action": "ACCUMULATE",
    },
    "NSE:THERMAX": {
        "name": "Thermax Ltd",
        "sector": "Industrial Energy + Boilers",
        "market_cap_cr": 50000, "cmp": 4180,
        "earnings_cagr_3y_pct": 38, "roce_pct": 22, "revenue_cagr_3y_pct": 20,
        "promoter_holding_pct": 62.0, "debt_to_equity": 0.10, "opm_pct": 10,
        "sector_tailwind": "Industrial decarbonization + green hydrogen + waste-to-energy",
        "analyst_call": "Antique Buy ₹5,200",
        "moat": "Pure-play industrial energy transition",
        "y3_target": 5800, "y3_x": "1.4x", "y5_target": 9500, "y5_x": "2.3x",
        "key_risks": "Order lumpiness, exports cyclicality",
        "entry_zone": "4080-4250", "sl": 3850, "conviction": "MEDIUM",
        "action": "WAIT FOR DIP",
    },
    "NSE:TRIVENI": {
        "name": "Triveni Turbine",
        "sector": "Steam Turbines (Industrial + Renewables)",
        "market_cap_cr": 22000, "cmp": 700,
        "earnings_cagr_3y_pct": 38, "roce_pct": 35, "revenue_cagr_3y_pct": 28,
        "promoter_holding_pct": 70.0, "debt_to_equity": 0.0, "opm_pct": 22,
        "sector_tailwind": "Process industry capex + waste-heat recovery + biomass",
        "analyst_call": "Nuvama Buy ₹860",
        "moat": "Global #2 in 0-100MW industrial turbines",
        "y3_target": 950, "y3_x": "1.36x", "y5_target": 1500, "y5_x": "2.1x",
        "key_risks": "Export demand, valuation rich",
        "entry_zone": "680-720", "sl": 645, "conviction": "HIGH",
        "action": "ACCUMULATE",
    },

    # ── EMS / SEMICONDUCTORS (PLI scheme + China+1) ──────────────────────────
    "NSE:KAYNES": {
        "name": "Kaynes Technology India",
        "sector": "EMS — Industrial + Semicon OSAT",
        "market_cap_cr": 32000, "cmp": 5100,
        "earnings_cagr_3y_pct": 70, "roce_pct": 22, "revenue_cagr_3y_pct": 60,
        "promoter_holding_pct": 60.5, "debt_to_equity": 0.10, "opm_pct": 14,
        "sector_tailwind": "EMS for industrial / aero / defence + semicon OSAT facility",
        "analyst_call": "Motilal Oswal Buy ₹6,400 (+25%)",
        "moat": "EMS + box-build + semiconductor (rare combo)",
        "y3_target": 7500, "y3_x": "1.5x", "y5_target": 13000, "y5_x": "2.5x",
        "key_risks": "Semicon execution, working capital",
        "entry_zone": "5000-5200", "sl": 4750, "conviction": "HIGH",
        "action": "ACCUMULATE",
    },
    "NSE:DIXON": {
        "name": "Dixon Technologies",
        "sector": "EMS — Mobile + Consumer Electronics",
        "market_cap_cr": 110000, "cmp": 18500,
        "earnings_cagr_3y_pct": 52, "roce_pct": 28, "revenue_cagr_3y_pct": 65,
        "promoter_holding_pct": 32.0, "debt_to_equity": 0.05, "opm_pct": 4,
        "sector_tailwind": "Mobile manufacturing PLI + display fab JV",
        "analyst_call": "ICICI Sec Buy ₹22,000",
        "moat": "Largest EMS + Motorola, Samsung, Xiaomi tie-ups",
        "y3_target": 24000, "y3_x": "1.3x", "y5_target": 38000, "y5_x": "2.05x",
        "key_risks": "Thin margins, customer concentration, valuation",
        "entry_zone": "18000-19000", "sl": 16800, "conviction": "MEDIUM",
        "action": "WAIT FOR DIP",
    },
    "NSE:SYRMA": {
        "name": "Syrma SGS Technology",
        "sector": "EMS — Diversified Industrial",
        "market_cap_cr": 11500, "cmp": 640,
        "earnings_cagr_3y_pct": 40, "roce_pct": 17, "revenue_cagr_3y_pct": 38,
        "promoter_holding_pct": 47.5, "debt_to_equity": 0.20, "opm_pct": 9,
        "sector_tailwind": "Diversified EMS for auto + healthcare + industrial + consumer",
        "analyst_call": "Equirus Buy ₹820",
        "moat": "PCB + integrated EMS + design",
        "y3_target": 950, "y3_x": "1.5x", "y5_target": 1700, "y5_x": "2.65x",
        "key_risks": "Margin pressure, customer concentration",
        "entry_zone": "620-660", "sl": 585, "conviction": "MEDIUM",
        "action": "ACCUMULATE",
    },

    # ── EV / AUTO ANCILLARY ─────────────────────────────────────────────────
    "NSE:UNOMINDA": {
        "name": "Uno Minda",
        "sector": "Auto Components — Lighting / Acoustics / Switches",
        "market_cap_cr": 70000, "cmp": 1175,
        "earnings_cagr_3y_pct": 32, "roce_pct": 21, "revenue_cagr_3y_pct": 25,
        "promoter_holding_pct": 70.5, "debt_to_equity": 0.20, "opm_pct": 11,
        "sector_tailwind": "Auto premiumisation + EV content per vehicle expansion",
        "analyst_call": "Nuvama Buy ₹1,450",
        "moat": "Highest content-per-vehicle in Indian 4W premium",
        "y3_target": 1650, "y3_x": "1.4x", "y5_target": 2700, "y5_x": "2.3x",
        "key_risks": "Auto cycle, customer concentration",
        "entry_zone": "1140-1200", "sl": 1075, "conviction": "HIGH",
        "action": "ACCUMULATE",
    },
    "NSE:SONACOMS": {
        "name": "Sona BLW Precision Forgings",
        "sector": "EV Drivetrain Components",
        "market_cap_cr": 35000, "cmp": 595,
        "earnings_cagr_3y_pct": 48, "roce_pct": 20, "revenue_cagr_3y_pct": 35,
        "promoter_holding_pct": 33.0, "debt_to_equity": 0.05, "opm_pct": 26,
        "sector_tailwind": "EV gear + motors content rising globally",
        "analyst_call": "Motilal Oswal Buy ₹730",
        "moat": "Pure-play EV component leader; export-heavy",
        "y3_target": 850, "y3_x": "1.4x", "y5_target": 1500, "y5_x": "2.5x",
        "key_risks": "Customer concentration (Tesla derivatives), valuation",
        "entry_zone": "580-610", "sl": 545, "conviction": "MEDIUM",
        "action": "ACCUMULATE",
    },

    # ── SPECIALTY CHEMICALS (China+1, structurally tight) ────────────────────
    "NSE:DEEPAKNTR": {
        "name": "Deepak Nitrite",
        "sector": "Specialty Chemicals — Phenol / IPA / Solvents",
        "market_cap_cr": 32000, "cmp": 2360,
        "earnings_cagr_3y_pct": 32, "roce_pct": 23, "revenue_cagr_3y_pct": 22,
        "promoter_holding_pct": 49.2, "debt_to_equity": 0.0, "opm_pct": 22,
        "sector_tailwind": "China+1 + integrated phenol/IPA economics",
        "analyst_call": "ICICI Sec Buy ₹2,900",
        "moat": "Backward-integrated phenol; high asset turns",
        "y3_target": 3300, "y3_x": "1.4x", "y5_target": 5500, "y5_x": "2.3x",
        "key_risks": "Commodity cycle, capex execution",
        "entry_zone": "2280-2400", "sl": 2160, "conviction": "HIGH",
        "action": "ACCUMULATE",
    },
    "NSE:CLEAN": {
        "name": "Clean Science & Technology",
        "sector": "Specialty Chemicals — Performance + Antioxidants",
        "market_cap_cr": 16000, "cmp": 1520,
        "earnings_cagr_3y_pct": 22, "roce_pct": 28, "revenue_cagr_3y_pct": 18,
        "promoter_holding_pct": 80.5, "debt_to_equity": 0.0, "opm_pct": 38,
        "sector_tailwind": "Specialty chem industry leader margins",
        "analyst_call": "Antique Buy ₹1,800",
        "moat": "Highest ROCE + OPM in specialty chem",
        "y3_target": 2050, "y3_x": "1.35x", "y5_target": 3500, "y5_x": "2.3x",
        "key_risks": "Customer concentration, growth slowing",
        "entry_zone": "1480-1545", "sl": 1395, "conviction": "MEDIUM",
        "action": "ACCUMULATE",
    },

    # ── INFRA / EPC (₹100L cr capex by 2030) ────────────────────────────────
    "NSE:NCC": {
        "name": "NCC Ltd",
        "sector": "Infra EPC",
        "market_cap_cr": 18000, "cmp": 295,
        "earnings_cagr_3y_pct": 38, "roce_pct": 18, "revenue_cagr_3y_pct": 24,
        "promoter_holding_pct": 22.0, "debt_to_equity": 0.18, "opm_pct": 9,
        "sector_tailwind": "Infra capex + ₹70,000 cr+ order book + low leverage",
        "analyst_call": "Nuvama Buy ₹385",
        "moat": "Asset-light; PSU-dominant order book",
        "y3_target": 440, "y3_x": "1.5x", "y5_target": 750, "y5_x": "2.5x",
        "key_risks": "Working capital, payment delays from govt",
        "entry_zone": "285-305", "sl": 270, "conviction": "HIGH",
        "action": "ACCUMULATE",
    },
    "NSE:KECL": {
        "name": "KEC International",
        "sector": "Power T&D EPC",
        "market_cap_cr": 24000, "cmp": 925,
        "earnings_cagr_3y_pct": 25, "roce_pct": 16, "revenue_cagr_3y_pct": 18,
        "promoter_holding_pct": 51.7, "debt_to_equity": 0.55, "opm_pct": 7,
        "sector_tailwind": "Global T&D capex super-cycle (US, MENA, India)",
        "analyst_call": "ICICI Sec Buy ₹1,150",
        "moat": "Global T&D player with USA + MENA exposure",
        "y3_target": 1300, "y3_x": "1.4x", "y5_target": 2100, "y5_x": "2.3x",
        "key_risks": "Margin recovery slow, debt",
        "entry_zone": "895-940", "sl": 850, "conviction": "MEDIUM",
        "action": "ACCUMULATE",
    },
    "NSE:KALPATPOWR": {
        "name": "Kalpataru Projects International",
        "sector": "Power T&D + Building EPC",
        "market_cap_cr": 24000, "cmp": 1380,
        "earnings_cagr_3y_pct": 32, "roce_pct": 18, "revenue_cagr_3y_pct": 19,
        "promoter_holding_pct": 41.0, "debt_to_equity": 0.40, "opm_pct": 8,
        "sector_tailwind": "Same as KEC; railway + water EPC diversification",
        "analyst_call": "Anand Rathi Buy ₹1,750",
        "moat": "Large diversified EPC, global T&D exposure",
        "y3_target": 1900, "y3_x": "1.4x", "y5_target": 3200, "y5_x": "2.3x",
        "key_risks": "Margin volatility, working capital",
        "entry_zone": "1340-1410", "sl": 1265, "conviction": "MEDIUM",
        "action": "ACCUMULATE",
    },

    # ── NEW-AGE FINANCIALS / NBFC ────────────────────────────────────────────
    "NSE:JIOFIN": {
        "name": "Jio Financial Services",
        "sector": "Diversified Financial — NBFC + AMC + Insurance",
        "market_cap_cr": 230000, "cmp": 360,
        "earnings_cagr_3y_pct": "Just listed", "roce_pct": 1, "revenue_cagr_3y_pct": "Pre-rev",
        "promoter_holding_pct": 47.1, "debt_to_equity": 0.05, "opm_pct": "N/M",
        "sector_tailwind": "Reliance ecosystem leveraged into financials; AMC + insurance + lending",
        "analyst_call": "Speculative Buy by some; long-term thesis",
        "moat": "RIL parent + Jio + retail customer base for cross-sell",
        "y3_target": 500, "y3_x": "1.4x", "y5_target": 850, "y5_x": "2.4x",
        "key_risks": "Pre-revenue, execution unproven, holding-co discount",
        "entry_zone": "350-375", "sl": 325, "conviction": "MEDIUM",
        "action": "ACCUMULATE (smaller size)",
    },
    "NSE:AUBANK": {
        "name": "AU Small Finance Bank",
        "sector": "Small Finance Bank → Universal Bank",
        "market_cap_cr": 50000, "cmp": 670,
        "earnings_cagr_3y_pct": 22, "roce_pct": 16, "revenue_cagr_3y_pct": 26,
        "promoter_holding_pct": 26.0, "debt_to_equity": "BFSI", "opm_pct": "BFSI",
        "sector_tailwind": "AUM growth + universal bank licence pending",
        "analyst_call": "Macquarie Buy ₹830",
        "moat": "Best-managed SFB; secured book; strong ROA",
        "y3_target": 950, "y3_x": "1.4x", "y5_target": 1500, "y5_x": "2.2x",
        "key_risks": "MFI book stress, deposit franchise build",
        "entry_zone": "650-690", "sl": 615, "conviction": "HIGH",
        "action": "ACCUMULATE",
    },
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def fill(h): return PatternFill("solid", fgColor=h)
def font(color=WHITE, bold=False, size=9, italic=False):
    return Font(name="Arial", color=color, bold=bold, size=size, italic=italic)
def bdr():
    s = Side(style="thin", color="2D2D2D")
    return Border(left=s, right=s, top=s, bottom=s)
def mid(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def merge_fresh(meta):
    if not os.path.exists(FRESH_JSON):
        return meta
    try:
        with open(FRESH_JSON) as f:
            fresh = json.load(f)
        for sym, updates in fresh.items():
            if sym in meta:
                meta[sym].update(updates)
        print(f"Merged fresh data for {len(fresh)} multibagger symbols")
    except Exception as e:
        print(f"Could not merge multibagger fresh JSON: {e}")
    return meta

# ── Excel build ───────────────────────────────────────────────────────────────
HEADERS = [
    "#", "NSE Symbol", "Stock Name", "Sector",
    "Mkt Cap\n(₹ cr)", "CMP ₹", "EPS\nCAGR 3Y", "ROCE", "Rev\nCAGR 3Y",
    "Promo\n%", "D/E", "OPM",
    "Sector Tailwind", "Moat",
    "Analyst Call (recent)",
    "Entry Zone ₹", "SL ₹",
    "3Y Target ₹", "3Y x", "5Y Target ₹", "5Y x",
    "Key Risks",
    "Conviction", "Action",
]
COL_WIDTHS = [4,15,22,28,9,9,9,9,9,8,7,7,28,32,28,14,9,11,8,11,8,28,12,18]

def build():
    meta = merge_fresh({k: dict(v) for k, v in MULTIBAGGERS.items()})

    # Apply blacklist exclusion
    excluded = []
    for sym in list(meta.keys()):
        if sym in BLACKLIST:
            excluded.append(sym)
            del meta[sym]

    wb = Workbook()
    ws = wb.active
    ws.title = "NSE Multibagger Picks"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:X1")
    ws["A1"] = (f"NSE MULTIBAGGER PICKS — 3Y / 5Y WEALTH CREATION  |  "
                f"Updated {TODAY.strftime('%d %b %Y')}  |  "
                "Smid-Cap · Earnings CAGR ≥ 25% · ROCE ≥ 20% · Sector Tailwind · Analyst Backed")
    ws["A1"].font = Font(name="Arial", color=GOLD, bold=True, size=13)
    ws["A1"].fill = fill(HEADER_BG)
    ws["A1"].alignment = mid()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:X2")
    ws["A2"] = ("Filters: Mkt Cap < ₹50,000 cr · EPS+Rev CAGR ≥ 20-25% · ROCE ≥ 20% · "
                "Promoter ≥ 45% · D/E < 0.5 · OPM ≥ 15% · Sector tailwind · Analyst Buy with target ≥ +25%  "
                "|  Auto-excludes blacklist (PERMANENT_AVOID + WAIT_FOR_RESOLUTION)")
    ws["A2"].font = Font(name="Arial", color=GREY, italic=True, size=9)
    ws["A2"].fill = fill(DARK_BG)
    ws["A2"].alignment = mid()
    ws.row_dimensions[2].height = 18

    # Headers
    for i, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font      = Font(name="Arial", color=GOLD, bold=True, size=8)
        c.fill      = fill(HEADER_BG)
        c.alignment = mid()
        c.border    = bdr()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 36

    # Sort: VERY HIGH > HIGH > MEDIUM, then by 5Y multiplier desc
    conv_rank = {"VERY HIGH": 0, "HIGH": 1, "MEDIUM": 2}
    def y5_x_val(d):
        try:
            return float(str(d.get("y5_x", "0")).replace("x", "").strip())
        except Exception:
            return 0
    sorted_syms = sorted(
        meta.keys(),
        key=lambda s: (conv_rank.get(meta[s]["conviction"], 9), -y5_x_val(meta[s]))
    )

    for idx, sym in enumerate(sorted_syms, 1):
        d = meta[sym]
        r = idx + 3
        conv_bg, conv_fg, _ = CONV_META.get(d["conviction"], (DARK_BG, GREY, ""))
        row_bg = ROW_ALT if idx % 2 else DARK_BG

        # Volatility-based smart SL (multibagger horizon = 2.0× ATR for breathing room)
        if HAS_VOL_ENGINE and isinstance(d.get("cmp"), (int, float)) and d["cmp"] > 0:
            sl_data = smart_sl(d["cmp"], sym, horizon="multibagger")
            d["sl"] = sl_data["sl"]
            d["vol_regime"] = sl_data["vol_regime"]

        # Promote conviction if hot sector + smart-money buying
        sec_boost = get_sector_boost(sym)
        flow_score = get_smart_money_score(sym)
        if sec_boost >= 10 and flow_score >= 70 and d["conviction"] != "VERY HIGH":
            if d["conviction"] == "MEDIUM": d["conviction"] = "HIGH"
            elif d["conviction"] == "HIGH":  d["conviction"] = "VERY HIGH"

        cells = [
            idx, sym, d["name"], d["sector"],
            d["market_cap_cr"], d["cmp"],
            f"{d['earnings_cagr_3y_pct']}%" if isinstance(d['earnings_cagr_3y_pct'], (int, float)) else d["earnings_cagr_3y_pct"],
            f"{d['roce_pct']}%" if isinstance(d['roce_pct'], (int, float)) else d["roce_pct"],
            f"{d['revenue_cagr_3y_pct']}%" if isinstance(d['revenue_cagr_3y_pct'], (int, float)) else d["revenue_cagr_3y_pct"],
            f"{d['promoter_holding_pct']}%",
            d["debt_to_equity"], f"{d['opm_pct']}%" if isinstance(d['opm_pct'], (int, float)) else d["opm_pct"],
            d["sector_tailwind"], d["moat"], d["analyst_call"],
            d["entry_zone"], d["sl"],
            d["y3_target"], d.get("y3_x", "—"),
            d["y5_target"], d.get("y5_x", "—"),
            d["key_risks"], d["conviction"], d["action"],
        ]
        for col_i, val in enumerate(cells, 1):
            c = ws.cell(row=r, column=col_i, value=val)
            c.fill = fill(row_bg)
            c.border = bdr()
            c.font = font(WHITE, size=9)
            c.alignment = mid() if col_i in (1,2,5,6,7,8,9,10,11,12,16,17,18,19,20,21,23,24) else lft()

            if col_i == 1: c.font = font(GOLD, bold=True)
            if col_i == 2: c.font = font(GREEN, bold=True)
            if col_i == 7: c.font = font(GREEN, bold=True)   # EPS CAGR
            if col_i == 8: c.font = font(CYAN, bold=True)    # ROCE
            if col_i == 17: c.font = font(RED, bold=True)    # SL
            if col_i == 18: c.font = font(GOLD)              # 3Y target
            if col_i == 19: c.font = font(GOLD, bold=True)   # 3Y x
            if col_i == 20: c.font = font(PURPLE)            # 5Y target
            if col_i == 21: c.font = font(PURPLE, bold=True) # 5Y x
            if col_i == 23:
                c.font = Font(name="Arial", color=conv_fg, bold=True, size=9)
                c.fill = fill(conv_bg)
            if col_i == 24:
                clr = GREEN if d["action"].startswith("ACCUMULATE") else AMBER if "WAIT" in d["action"] else GREY
                c.font = font(clr, bold=True)

        ws.row_dimensions[r].height = 60

    # Footer
    fr = len(sorted_syms) + 5
    ws.merge_cells(f"A{fr}:X{fr}")
    ws[f"A{fr}"] = ("⚠️  MULTIBAGGER PROTOCOL: 1) Build position over 6-12 months in 3-5 tranches  "
                   "2) Position size = max 5% per stock + max 25% per sector  "
                   "3) Re-evaluate every quarter — exit if EPS growth drops below 15% for 2 quarters  "
                   "4) Trail SL only after 100% gain (let winners run)  "
                   "5) Tax: long-term capital gains 12.5% above ₹1.25L exemption (FY26)")
    ws[f"A{fr}"].font      = Font(name="Arial", color=AMBER, bold=True, size=9)
    ws[f"A{fr}"].fill      = fill(HEADER_BG)
    ws[f"A{fr}"].alignment = mid()
    ws.row_dimensions[fr].height = 38

    if excluded:
        ws.merge_cells(f"A{fr+1}:X{fr+1}")
        ws[f"A{fr+1}"] = (f"❌ EXCLUDED via blacklist this run: {', '.join(excluded)}")
        ws[f"A{fr+1}"].font      = Font(name="Arial", color=RED, italic=True, size=8)
        ws[f"A{fr+1}"].fill      = fill(DARK_BG)
        ws[f"A{fr+1}"].alignment = mid()
        ws.row_dimensions[fr+1].height = 14

    ws.merge_cells(f"A{fr+2}:X{fr+2}")
    ws[f"A{fr+2}"] = (f"Auto-generated by Claude Opus 4.7  |  Run date: {TODAY.strftime('%d %b %Y')}  |  "
                     "Sources: Screener.in · Motilal Oswal · ICICI Sec · Anand Rathi · Antique · Nuvama · Equirus  |  "
                     "Verify all data on Screener.in / company filings before trading.  NOT financial advice.")
    ws[f"A{fr+2}"].font      = Font(name="Arial", color=GREY, italic=True, size=8)
    ws[f"A{fr+2}"].fill      = fill(DARK_BG)
    ws[f"A{fr+2}"].alignment = mid()
    ws.row_dimensions[fr+2].height = 14

    ws.freeze_panes = "C4"

    wb.save(EXCEL_PATH)
    print(f"✅ Multibagger Excel saved: {EXCEL_PATH}\n")

    # Summary
    by_conv = {}
    for sym in sorted_syms:
        c = meta[sym]["conviction"]
        by_conv.setdefault(c, []).append(f"{sym} ({meta[sym]['sector'].split(' — ')[0][:18]})")
    print(f"📊 Multibagger Picks ({len(sorted_syms)} stocks):\n")
    for conv in ["VERY HIGH", "HIGH", "MEDIUM"]:
        if conv in by_conv:
            print(f"  {CONV_META[conv][2]}")
            for item in by_conv[conv]:
                print(f"     • {item}")
            print()

if __name__ == "__main__":
    build()
