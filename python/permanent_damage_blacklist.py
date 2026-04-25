"""
NSE Permanent Damage Blacklist
─────────────────────────────────────────────────────────────────────────────
Stocks to AUTO-EXCLUDE from swing / multibagger / fallen-angel watchlists due
to structural damage that has NOT recovered. The weekly Opus 4.7 scheduled
task refreshes this list and writes blacklist_fresh.json which is merged in
on every screener run.

Severity levels:
  • PERMANENT_AVOID       — never trade, structural failure
  • WAIT_FOR_RESOLUTION   — wait for specific event, then re-evaluate
  • REDUCED_CONVICTION    — can trade with smaller size + tighter SL
─────────────────────────────────────────────────────────────────────────────
"""
import json, os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH    = r"C:\Users\razee\OneDrive\Desktop\TradingClaude\NSE_Permanent_Damage_Blacklist.xlsx"
FRESH_JSON    = r"C:\Users\razee\OneDrive\Desktop\TradingClaude\blacklist_fresh.json"
TODAY         = date.today()

# ── Palette ───────────────────────────────────────────────────────────────────
DARK_BG="0D0D0D"; HEADER_BG="1A1A2E"; ROW_ALT="141414"
GOLD="FFD700"; GREEN="00C896"; BLUE="4FC3F7"; WHITE="FFFFFF"
GREY="9E9E9E"; RED="FF5252";  AMBER="FFB300"; ORANGE="FF6B35"
SEV_PERM_BG="2D0000"; SEV_WAIT_BG="3D2C00"; SEV_RED_BG="1A1A2E"

SEVERITY_META = {
    "PERMANENT_AVOID":     (SEV_PERM_BG, RED,    "❌ NEVER trade — structural failure"),
    "WAIT_FOR_RESOLUTION": (SEV_WAIT_BG, AMBER,  "⏳ Wait for catalyst event before re-evaluating"),
    "REDUCED_CONVICTION":  (SEV_RED_BG,  ORANGE, "⚠️ Trade smaller (50% size) + tighter SL"),
}

# ── Blacklist seed data (refresh weekly via Opus 4.7 task) ───────────────────
PERMANENT_DAMAGE_BLACKLIST = {
    # ── PERMANENT_AVOID ──────────────────────────────────────────────────────
    "NSE:IDEA": {
        "name": "Vodafone Idea Ltd", "category": "Bankrupt-zone Telecom",
        "trigger_event": "AGR Supreme Court ruling Oct 2019; ₹2L cr+ debt; 5G capex unfunded",
        "trigger_date": "Oct 2019 (ongoing)",
        "current_status": "Govt equity conversion + FPO did not fix structural deficit; market share losing to Jio + Airtel",
        "severity": "PERMANENT_AVOID",
        "monitoring_signal": "Promoter capital infusion ≥ ₹50,000 cr OR debt write-off",
    },
    "NSE:WIPRO": {
        "name": "Wipro Ltd", "category": "AI Disruption — IT Services",
        "trigger_event": "Multi-year revenue decline; AI eats staff-augmentation revenue",
        "trigger_date": "FY23 onwards",
        "current_status": "Weakest among Tier-1 IT; CEO churn; growth -2 to 0% range",
        "severity": "PERMANENT_AVOID",
        "monitoring_signal": "4 consecutive quarters of organic CC growth ≥ 5%",
    },
    "NSE:NDTV": {
        "name": "New Delhi Television", "category": "Editorial / Float Quality",
        "trigger_event": "Adani Group takeover Dec 2022; editorial concerns; small float",
        "trigger_date": "Dec 2022",
        "current_status": "Liquidity drained; minority shareholder de-rating",
        "severity": "PERMANENT_AVOID",
        "monitoring_signal": "Promoter sale or buyback at premium",
    },
    "NSE:ZEEL": {
        "name": "Zee Entertainment Enterprises", "category": "Governance",
        "trigger_event": "Sony merger collapse Jan 2024; SEBI fund-diversion findings vs promoters",
        "trigger_date": "Jan 2024",
        "current_status": "No substitute deal; promoter SAT cases ongoing; lender losses",
        "severity": "PERMANENT_AVOID",
        "monitoring_signal": "Promoter exit + new strategic buyer at clean valuation",
    },

    # ── WAIT_FOR_RESOLUTION ─────────────────────────────────────────────────
    "NSE:ADANIENT": {
        "name": "Adani Enterprises", "category": "Governance / Regulatory Overhang",
        "trigger_event": "US DOJ + SEC indictment Nov 2024 (alleged bribery); Hindenburg legacy",
        "trigger_date": "Nov 2024",
        "current_status": "Core ops solid but DOJ overhang caps multiple",
        "severity": "WAIT_FOR_RESOLUTION",
        "monitoring_signal": "DOJ resolution / settlement or charge-drop",
    },
    "NSE:ADANIGREEN": {
        "name": "Adani Green Energy", "category": "Governance",
        "trigger_event": "Named in DOJ indictment Nov 2024; dollar-bond access tightened",
        "trigger_date": "Nov 2024",
        "current_status": "PPA portfolio intact but capex funding constrained",
        "severity": "WAIT_FOR_RESOLUTION",
        "monitoring_signal": "DOJ closure + credit-rating restoration",
    },
    "NSE:ADANIENSOL": {
        "name": "Adani Energy Solutions", "category": "Governance",
        "trigger_event": "DOJ matter directly involves transmission projects",
        "trigger_date": "Nov 2024",
        "current_status": "Capex-heavy, rating-sensitive; bond market shut",
        "severity": "WAIT_FOR_RESOLUTION",
        "monitoring_signal": "DOJ resolution + AAA-rating retention",
    },
    "NSE:PAYTM": {
        "name": "One97 Communications", "category": "Regulatory-Impaired Fintech",
        "trigger_event": "RBI action on Paytm Payments Bank Jan 2024",
        "trigger_date": "Jan 2024",
        "current_status": "PA licence live; loan distribution scaling — improving but binary",
        "severity": "WAIT_FOR_RESOLUTION",
        "monitoring_signal": "First sustained PAT-positive quarter",
    },
    "NSE:IIFL": {
        "name": "IIFL Finance", "category": "NBFC Regulatory",
        "trigger_event": "RBI gold-loan ban March 2024; trust damage",
        "trigger_date": "Mar 2024",
        "current_status": "Ban lifted but AUM growth slow; brand damage real",
        "severity": "WAIT_FOR_RESOLUTION",
        "monitoring_signal": "AUM growth back to 25%+ + ROA recovery",
    },
    "NSE:JMFINANCIL": {
        "name": "JM Financial", "category": "Regulatory",
        "trigger_event": "SEBI bar on IPO debt-financing March 2024",
        "trigger_date": "Mar 2024",
        "current_status": "Diversifying; high-conviction wealth + AMC arms",
        "severity": "WAIT_FOR_RESOLUTION",
        "monitoring_signal": "Re-entry to IPO financing + clean SEBI standing",
    },

    # ── REDUCED_CONVICTION ──────────────────────────────────────────────────
    "NSE:YESBANK": {
        "name": "Yes Bank", "category": "Damaged Franchise Bank",
        "trigger_event": "Mar 2020 reconstruction; ROE structurally weak vs private peers",
        "trigger_date": "Mar 2020 (legacy)",
        "current_status": "SBI exit + new owner transition is binary; asset quality OK but ROE 0-1%",
        "severity": "REDUCED_CONVICTION",
        "monitoring_signal": "ROE > 8% sustained for 2 quarters",
    },
    "NSE:RBLBANK": {
        "name": "RBL Bank", "category": "Weak Private Bank",
        "trigger_event": "MFI/cards stress 2021-2022; leadership churn",
        "trigger_date": "2021 onwards",
        "current_status": "ROA below peers; valuations cheap but execution risk",
        "severity": "REDUCED_CONVICTION",
        "monitoring_signal": "ROA > 1% + GNPA < 2.5%",
    },
    "NSE:INFY": {
        "name": "Infosys Ltd", "category": "AI Disruption — IT Services",
        "trigger_event": "Gen-AI deflation on staff-augmentation revenue",
        "trigger_date": "FY24-FY25",
        "current_status": "Better than Wipro but growth durably mid-single-digit",
        "severity": "REDUCED_CONVICTION",
        "monitoring_signal": "Organic CC growth ≥ 8% guidance for FY27",
    },
    "NSE:TCS": {
        "name": "Tata Consultancy Services", "category": "AI Disruption — IT Services",
        "trigger_event": "Same AI compression on labour-arbitrage model",
        "trigger_date": "FY25",
        "current_status": "Most defensive Tier-1 but growth structurally lower vs FY15-FY20",
        "severity": "REDUCED_CONVICTION",
        "monitoring_signal": "Organic CC growth ≥ 7% + AI-engagement TCV materialising",
    },
    "NSE:TECHM": {
        "name": "Tech Mahindra", "category": "AI Disruption — IT Services",
        "trigger_event": "Telecom-vertical concentration in Gen-AI compression",
        "trigger_date": "FY24",
        "current_status": "New CEO turnaround attempt; margins recovering",
        "severity": "REDUCED_CONVICTION",
        "monitoring_signal": "EBIT margin > 12% + telecom vertical growth",
    },
    "NSE:LTIMINDTREE": {
        "name": "LTI Mindtree", "category": "AI Disruption — IT Services",
        "trigger_event": "Merger synergies under-delivered; BFSI slowdown",
        "trigger_date": "FY25",
        "current_status": "Structural growth questioned",
        "severity": "REDUCED_CONVICTION",
        "monitoring_signal": "Revenue per employee inflection",
    },
    "NSE:MPHASIS": {
        "name": "Mphasis Ltd", "category": "AI Disruption — IT Services",
        "trigger_event": "Mortgage vertical exposure; AI on staff-aug",
        "trigger_date": "FY24",
        "current_status": "Diversification underway",
        "severity": "REDUCED_CONVICTION",
        "monitoring_signal": "Non-mortgage BFSI growth ≥ 10%",
    },
    "NSE:HCLTECH": {
        "name": "HCL Technologies", "category": "AI Disruption — IT Services (Defensive)",
        "trigger_event": "Same AI compression but balanced products portfolio",
        "trigger_date": "FY25",
        "current_status": "Best-positioned among IT services for AI age (products + ER&D)",
        "severity": "REDUCED_CONVICTION",
        "monitoring_signal": "Products growth > 15% sustained",
    },
    "NSE:COFORGE": {
        "name": "Coforge Ltd", "category": "Execution Risk",
        "trigger_event": "Cigniti acquisition integration; valuation rich",
        "trigger_date": "2024",
        "current_status": "Integration on track but execution risk high at valuations",
        "severity": "REDUCED_CONVICTION",
        "monitoring_signal": "Integrated revenue growth ≥ 18%",
    },
    "NSE:DIVISLAB": {
        "name": "Divi's Laboratories", "category": "Pharma Regulatory",
        "trigger_event": "USFDA history at Unit-2; API pricing pressure",
        "trigger_date": "Recurring",
        "current_status": "Resolved most issues; export demand recovering",
        "severity": "REDUCED_CONVICTION",
        "monitoring_signal": "API margin > 30% sustained 4 quarters",
    },
    "NSE:AUROPHARMA": {
        "name": "Aurobindo Pharma", "category": "Pharma Regulatory",
        "trigger_event": "Recurring USFDA observations historically",
        "trigger_date": "Recurring",
        "current_status": "Improved compliance but plant audit risk persists",
        "severity": "REDUCED_CONVICTION",
        "monitoring_signal": "12 months without 483/Warning Letter",
    },
    "NSE:GLENMARK": {
        "name": "Glenmark Pharmaceuticals", "category": "Pharma Regulatory + Leverage",
        "trigger_event": "Leverage + USFDA Monroe history",
        "trigger_date": "Recurring",
        "current_status": "IGI stake sale deleveraged BS",
        "severity": "REDUCED_CONVICTION",
        "monitoring_signal": "Net D/E < 0.3 + clean USFDA",
    },
    "NSE:BIOCON": {
        "name": "Biocon Ltd", "category": "Sector Decline / Margin Compression",
        "trigger_event": "Biosimilars margin compression; Viatris debt",
        "trigger_date": "2022 onwards",
        "current_status": "Restructuring; biosimilars price war ongoing",
        "severity": "REDUCED_CONVICTION",
        "monitoring_signal": "Biosimilar EBITDA margin > 25%",
    },
    "NSE:IRCON": {
        "name": "IRCON International", "category": "PSU Re-rating Reversal",
        "trigger_event": "PSU multiples compressed post-2024 froth",
        "trigger_date": "2024",
        "current_status": "Order book growth slowing; pricing thin",
        "severity": "REDUCED_CONVICTION",
        "monitoring_signal": "Order inflow growth > 20% YoY",
    },
    "NSE:RVNL": {
        "name": "Rail Vikas Nigam Ltd", "category": "PSU Re-rating Reversal",
        "trigger_event": "Same PSU re-rating risk; razor-thin EBIT margins",
        "trigger_date": "2024",
        "current_status": "Vande Bharat orders supportive but margin squeezed",
        "severity": "REDUCED_CONVICTION",
        "monitoring_signal": "EBIT margin > 7% + new business segments",
    },
    "NSE:IREDA": {
        "name": "Indian Renewable Energy Dev. Agency", "category": "PSU Post-IPO Froth",
        "trigger_event": "Post-IPO valuation excess; renewables NPA risk emerging",
        "trigger_date": "2024",
        "current_status": "AUM growth strong but asset quality under watch",
        "severity": "REDUCED_CONVICTION",
        "monitoring_signal": "GNPA < 2.5% + AUM growth > 25%",
    },
    "NSE:SUZLON": {
        "name": "Suzlon Energy", "category": "Capital Structure History",
        "trigger_event": "Repeated dilution history; order execution risk",
        "trigger_date": "Legacy",
        "current_status": "Turnaround real but execution slippage costly",
        "severity": "REDUCED_CONVICTION",
        "monitoring_signal": "Order book execution ≥ 90% on time",
    },
    "NSE:BPCL": {
        "name": "Bharat Petroleum Corp", "category": "PSU OMC",
        "trigger_event": "Under-recovery + privatization off-table",
        "trigger_date": "Recurring",
        "current_status": "Marketing margins govt-controlled; refinery cyclical",
        "severity": "REDUCED_CONVICTION",
        "monitoring_signal": "Decontrol of retail diesel/petrol pricing",
    },
    "NSE:HPCL": {
        "name": "Hindustan Petroleum Corp", "category": "PSU OMC",
        "trigger_event": "Same OMC issue; smaller balance sheet absorbs less shock",
        "trigger_date": "Recurring",
        "current_status": "Same as BPCL",
        "severity": "REDUCED_CONVICTION",
        "monitoring_signal": "Same as BPCL",
    },
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def fill(h): return PatternFill("solid", fgColor=h)
def font(color=WHITE, bold=False, size=9, italic=False):
    return Font(name="Arial", color=color, bold=bold, size=size, italic=italic)
def bdr():
    s = Side(style="thin", color="2D2D2D")
    return Border(left=s, right=s, top=s, bottom=s)
def mid(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def merge_fresh(meta):
    if not os.path.exists(FRESH_JSON):
        return meta
    try:
        with open(FRESH_JSON) as f:
            fresh = json.load(f)
        for sym, updates in fresh.items():
            if sym in meta:
                meta[sym].update(updates)
            else:
                meta[sym] = updates
        print(f"Merged fresh data for {len(fresh)} symbols from blacklist_fresh.json")
    except Exception as e:
        print(f"Could not merge fresh JSON: {e}")
    return meta

def get_blacklist_set(severity_filter=None):
    """Returns a set of NSE symbols to exclude from screeners.
       severity_filter: list of severities to include, default = all."""
    if severity_filter is None:
        severity_filter = ["PERMANENT_AVOID", "WAIT_FOR_RESOLUTION", "REDUCED_CONVICTION"]
    meta = merge_fresh({k: dict(v) for k, v in PERMANENT_DAMAGE_BLACKLIST.items()})
    return {sym for sym, d in meta.items() if d["severity"] in severity_filter}

# ── Excel build ───────────────────────────────────────────────────────────────
HEADERS = [
    "#", "NSE Symbol", "Stock Name", "Category",
    "Trigger Event", "Trigger Date",
    "Current Status", "Severity", "Monitoring Signal (un-blacklist trigger)",
]
COL_WIDTHS = [4, 16, 24, 24, 36, 16, 38, 22, 38]

def build():
    meta = merge_fresh({k: dict(v) for k, v in PERMANENT_DAMAGE_BLACKLIST.items()})

    wb = Workbook()
    ws = wb.active
    ws.title = "NSE Permanent Damage Blacklist"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = (f"NSE PERMANENT DAMAGE BLACKLIST  |  Updated {TODAY.strftime('%d %b %Y')}  |  "
                "Auto-Excluded from Swing / Multibagger / Fallen-Angel Screeners")
    ws["A1"].font = Font(name="Arial", color=GOLD, bold=True, size=13)
    ws["A1"].fill = fill(HEADER_BG)
    ws["A1"].alignment = mid()
    ws.row_dimensions[1].height = 28

    # Sub-header
    ws.merge_cells("A2:I2")
    ws["A2"] = ("Severity:  ❌ PERMANENT_AVOID — never trade  |  "
                "⏳ WAIT_FOR_RESOLUTION — wait for catalyst  |  "
                "⚠️ REDUCED_CONVICTION — trade smaller (50% size) + tighter SL")
    ws["A2"].font = Font(name="Arial", color=AMBER, italic=True, size=9)
    ws["A2"].fill = fill(DARK_BG)
    ws["A2"].alignment = mid()
    ws.row_dimensions[2].height = 18

    # Headers
    for i, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font      = Font(name="Arial", color=GOLD, bold=True, size=9)
        c.fill      = fill(HEADER_BG)
        c.alignment = mid()
        c.border    = bdr()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 30

    # Sort by severity (PERMANENT first, then WAIT, then REDUCED)
    sev_rank = {"PERMANENT_AVOID": 0, "WAIT_FOR_RESOLUTION": 1, "REDUCED_CONVICTION": 2}
    sorted_syms = sorted(meta.keys(), key=lambda s: (sev_rank.get(meta[s]["severity"], 9), s))

    for idx, sym in enumerate(sorted_syms, 1):
        d = meta[sym]
        r = idx + 3
        sev_bg, sev_fg, _ = SEVERITY_META.get(d["severity"], (DARK_BG, GREY, ""))

        cells = [
            idx, sym, d["name"], d["category"],
            d["trigger_event"], d["trigger_date"],
            d["current_status"], d["severity"], d["monitoring_signal"],
        ]
        for col_i, val in enumerate(cells, 1):
            c = ws.cell(row=r, column=col_i, value=val)
            c.fill   = fill(sev_bg)
            c.border = bdr()
            c.font   = font(WHITE, size=9)
            c.alignment = mid() if col_i in (1, 2, 6, 8) else lft()

            if col_i == 1: c.font = font(GOLD, bold=True)
            if col_i == 2: c.font = font(RED if d["severity"] == "PERMANENT_AVOID" else AMBER, bold=True)
            if col_i == 8:
                c.font = Font(name="Arial", color=sev_fg, bold=True, size=9)
            if col_i == 9: c.font = font(GREEN, italic=True, size=9)

        ws.row_dimensions[r].height = 48

    # Footer
    fr = len(sorted_syms) + 5
    ws.merge_cells(f"A{fr}:I{fr}")
    ws[f"A{fr}"] = (f"Auto-generated by Claude Opus 4.7  |  Run date: {TODAY.strftime('%d %b %Y')}  |  "
                   "Refreshed weekly via news scans for governance / regulatory / structural events.  |  NOT financial advice.")
    ws[f"A{fr}"].font      = Font(name="Arial", color=GREY, italic=True, size=8)
    ws[f"A{fr}"].fill      = fill(DARK_BG)
    ws[f"A{fr}"].alignment = mid()
    ws.row_dimensions[fr].height = 14

    ws.freeze_panes = "C4"

    wb.save(EXCEL_PATH)
    print(f"✅ Blacklist Excel saved: {EXCEL_PATH}")

    # Console summary
    counts = {}
    for sym in sorted_syms:
        s = meta[sym]["severity"]
        counts[s] = counts.get(s, 0) + 1
    print(f"\n📊 Blacklist Summary ({len(sorted_syms)} stocks):")
    for sev in ["PERMANENT_AVOID", "WAIT_FOR_RESOLUTION", "REDUCED_CONVICTION"]:
        if sev in counts:
            print(f"  {SEVERITY_META[sev][2]:48s} {counts[sev]} stocks")

if __name__ == "__main__":
    build()
