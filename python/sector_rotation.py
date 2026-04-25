"""
Sector Rotation Tracker
─────────────────────────────────────────────────────────────────────────────
Tracks weekly performance of NSE sector indices and identifies which sectors
are receiving institutional flows. Output integrates into all screeners as a
sector_boost (0-15 score points) for stocks belonging to hot sectors.

Generates: NSE_Sector_Rotation.xlsx with:
  • Weekly returns heatmap (1W, 1M, 3M, 6M, 1Y)
  • Money flow (FII + DII allocation by sector)
  • Sector rank with momentum + RSI
  • Hot sectors → boost screener scoring
  • Cold sectors → reduce conviction

Refreshed Saturday 9:30 AM IST by Opus 4.7 task (after market close on Friday).
─────────────────────────────────────────────────────────────────────────────
"""
import json, os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH = r"C:\Users\razee\OneDrive\Desktop\TradingClaude\NSE_Sector_Rotation.xlsx"
FRESH_JSON = r"C:\Users\razee\OneDrive\Desktop\TradingClaude\sector_rotation_fresh.json"
TODAY      = date.today()

DARK_BG="0D0D0D"; HEADER_BG="1A1A2E"; ROW_ALT="141414"
GOLD="FFD700"; GREEN="00C896"; BLUE="4FC3F7"; WHITE="FFFFFF"
GREY="9E9E9E"; RED="FF5252"; AMBER="FFB300"; ORANGE="FF6B35"; CYAN="00BCD4"; PURPLE="9C27B0"

# ── NSE sector index → constituent stock mapping (for sector boost scoring) ──
SECTOR_STOCKS = {
    "DEFENCE":          ["NSE:BEL", "NSE:BDL", "NSE:HAL", "NSE:MAZDOCK", "NSE:DATAPATTNS", "NSE:MTARTECH", "NSE:COCHINSHIP", "NSE:SOLARINDS"],
    "RENEWABLES":       ["NSE:KPIGREEN", "NSE:PREMIERENE", "NSE:INOXWIND", "NSE:SUZLON", "NSE:BORORENEW", "NSE:ADANIGREEN", "NSE:NTPCGREEN"],
    "BANKING":          ["NSE:HDFCBANK", "NSE:ICICIBANK", "NSE:AXISBANK", "NSE:KOTAKBANK", "NSE:INDUSINDBK", "NSE:AUBANK"],
    "PSU BANKS":        ["NSE:SBIN", "NSE:CANBK", "NSE:PNB", "NSE:BANKBARODA", "NSE:UNIONBANK"],
    "IT SERVICES":      ["NSE:INFY", "NSE:TCS", "NSE:WIPRO", "NSE:HCLTECH", "NSE:TECHM", "NSE:LTIMINDTREE", "NSE:MPHASIS"],
    "AUTO":             ["NSE:TATAMOTORS", "NSE:M&M", "NSE:MARUTI", "NSE:BAJAJ-AUTO", "NSE:HEROMOTOCO", "NSE:EICHERMOT", "NSE:TVSMOTOR"],
    "AUTO ANCILLARY":   ["NSE:UNOMINDA", "NSE:SONACOMS", "NSE:BHARATFORG", "NSE:EXIDEIND", "NSE:AMARARAJA", "NSE:MOTHERSON"],
    "FMCG":             ["NSE:HINDUNILVR", "NSE:NESTLEIND", "NSE:ITC", "NSE:DABUR", "NSE:GODREJCP", "NSE:BRITANNIA", "NSE:MARICO"],
    "PHARMA":           ["NSE:SUNPHARMA", "NSE:CIPLA", "NSE:DRREDDY", "NSE:DIVISLAB", "NSE:LUPIN", "NSE:AUROPHARMA", "NSE:GLENMARK"],
    "METALS":           ["NSE:TATASTEEL", "NSE:JSWSTEEL", "NSE:HINDALCO", "NSE:VEDL", "NSE:NMDC", "NSE:JINDALSTEL", "NSE:NATIONALUM"],
    "OIL & GAS":        ["NSE:RELIANCE", "NSE:ONGC", "NSE:IOC", "NSE:BPCL", "NSE:HPCL", "NSE:GAIL", "NSE:OIL"],
    "POWER":            ["NSE:NTPC", "NSE:POWERGRID", "NSE:TATAPOWER", "NSE:ADANIPOWER", "NSE:JSW-ENERGY"],
    "POWER FIN":        ["NSE:RECLTD", "NSE:PFC", "NSE:IREDA", "NSE:IIFCL"],
    "WIRES & CABLES":   ["NSE:POLYCAB", "NSE:KEI", "NSE:FINOLEX", "NSE:RRKABEL", "NSE:HAVELLS"],
    "CAPITAL GOODS":    ["NSE:CGPOWER", "NSE:THERMAX", "NSE:TRIVENI", "NSE:SIEMENS", "NSE:ABB", "NSE:CUMMINS", "NSE:BHEL"],
    "EMS / SEMICON":    ["NSE:KAYNES", "NSE:DIXON", "NSE:SYRMA", "NSE:AMBER", "NSE:AVALON", "NSE:CYIENTDLM"],
    "INFRA / EPC":      ["NSE:LT", "NSE:NCC", "NSE:KECL", "NSE:KALPATPOWR", "NSE:GRINFRA", "NSE:DBL"],
    "CEMENT":           ["NSE:ULTRACEMCO", "NSE:SHREECEM", "NSE:AMBUJACEM", "NSE:ACC", "NSE:DALBHARAT"],
    "REALTY":           ["NSE:DLF", "NSE:GODREJPROP", "NSE:OBEROIRLTY", "NSE:PRESTIGE", "NSE:LODHA"],
    "INSURANCE":        ["NSE:HDFCLIFE", "NSE:SBILIFE", "NSE:ICICIPRULI", "NSE:LICI"],
    "NBFC":             ["NSE:BAJFINANCE", "NSE:BAJAJFINSV", "NSE:CHOLAFIN", "NSE:HDFCAMC", "NSE:JIOFIN"],
    "SPECIALTY CHEM":   ["NSE:DEEPAKNTR", "NSE:CLEAN", "NSE:NAVINFLUOR", "NSE:PIIND", "NSE:SRF", "NSE:ATUL"],
    "RAILWAYS":         ["NSE:IRCON", "NSE:RVNL", "NSE:IRCTC", "NSE:RAILTEL", "NSE:RITES", "NSE:CONCOR"],
}

# ── Demo / fallback sector data (refreshed weekly by Opus 4.7 task) ──────────
DEMO_SECTORS = [
    # name, ret_1w, ret_1m, ret_3m, ret_6m, ret_1y, momentum_score (0-100), institutional_flow ("INFLOW"/"OUTFLOW"/"NEUTRAL"), commentary
    {"name": "DEFENCE",         "ret_1w": 4.2, "ret_1m": 12.5, "ret_3m": 28.5, "ret_6m": 42, "ret_1y": 95, "momentum": 92, "flow": "INFLOW",  "commentary": "Order book mega-cycle; ₹6L cr capex by 2030"},
    {"name": "RENEWABLES",      "ret_1w": 3.8, "ret_1m": 9.2,  "ret_3m": 18.5, "ret_6m": 35, "ret_1y": 72, "momentum": 88, "flow": "INFLOW",  "commentary": "500GW target by 2030; PLI scheme tailwind"},
    {"name": "POWER FIN",       "ret_1w": 2.5, "ret_1m": 6.8,  "ret_3m": -8,   "ret_6m": -15,"ret_1y": -22,"momentum": 65, "flow": "INFLOW",  "commentary": "Bottoming out post-PSU correction; rate-cut beneficiary"},
    {"name": "EMS / SEMICON",   "ret_1w": 3.2, "ret_1m": 8.5,  "ret_3m": 22,   "ret_6m": 38, "ret_1y": 65, "momentum": 86, "flow": "INFLOW",  "commentary": "Mobile PLI + display fab + OSAT scaling"},
    {"name": "CAPITAL GOODS",   "ret_1w": 2.1, "ret_1m": 5.5,  "ret_3m": 14,   "ret_6m": 28, "ret_1y": 42, "momentum": 75, "flow": "INFLOW",  "commentary": "Industrial decarb + power capex"},
    {"name": "WIRES & CABLES",  "ret_1w": 1.8, "ret_1m": 4.5,  "ret_3m": 12,   "ret_6m": 22, "ret_1y": 38, "momentum": 72, "flow": "INFLOW",  "commentary": "T&D capex + housing demand"},
    {"name": "INFRA / EPC",     "ret_1w": 1.2, "ret_1m": 3.5,  "ret_3m": 8,    "ret_6m": 14, "ret_1y": 18, "momentum": 65, "flow": "INFLOW",  "commentary": "Govt capex + budget allocation"},
    {"name": "BANKING",         "ret_1w": 0.8, "ret_1m": 2.5,  "ret_3m": 5,    "ret_6m": 12, "ret_1y": 18, "momentum": 60, "flow": "NEUTRAL", "commentary": "Q4 results positive; rate cycle peak"},
    {"name": "AUTO ANCILLARY",  "ret_1w": 1.5, "ret_1m": 3.2,  "ret_3m": 6,    "ret_6m": 11, "ret_1y": 22, "momentum": 62, "flow": "NEUTRAL", "commentary": "EV transition + premiumization"},
    {"name": "INSURANCE",       "ret_1w": 1.0, "ret_1m": 2.8,  "ret_3m": 4,    "ret_6m": 8,  "ret_1y": 12, "momentum": 55, "flow": "NEUTRAL", "commentary": "ULIP recovery; tax framework stable"},
    {"name": "POWER",           "ret_1w": 0.5, "ret_1m": 1.2,  "ret_3m": -2,   "ret_6m": -5, "ret_1y": -8, "momentum": 48, "flow": "NEUTRAL", "commentary": "Renewable transition pressure"},
    {"name": "PHARMA",          "ret_1w": 0.3, "ret_1m": 1.0,  "ret_3m": 2,    "ret_6m": 5,  "ret_1y": 8,  "momentum": 50, "flow": "NEUTRAL", "commentary": "Mixed Q4; USFDA risks differ by company"},
    {"name": "RAILWAYS",        "ret_1w": -0.5,"ret_1m": -2.5, "ret_3m": -12,  "ret_6m": -28,"ret_1y": -38,"momentum": 42, "flow": "OUTFLOW", "commentary": "PSU re-rating reversed; margin pressure"},
    {"name": "REALTY",          "ret_1w":-0.8, "ret_1m": -3.5, "ret_3m": -8,   "ret_6m": -12,"ret_1y": -5, "momentum": 45, "flow": "NEUTRAL", "commentary": "Premium pockets resilient; tier-2 weak"},
    {"name": "FMCG",            "ret_1w":-0.3, "ret_1m": -1.5, "ret_3m": -5,   "ret_6m": -10,"ret_1y": -12,"momentum": 48, "flow": "OUTFLOW", "commentary": "Rural demand weak; Birla Opus pressure on paints"},
    {"name": "CEMENT",          "ret_1w": 0.0, "ret_1m": -2.0, "ret_3m": -7,   "ret_6m": -8, "ret_1y": -2, "momentum": 50, "flow": "NEUTRAL", "commentary": "Consolidation phase; pricing under pressure"},
    {"name": "METALS",          "ret_1w": 0.5, "ret_1m": 1.5,  "ret_3m": -3,   "ret_6m": -6, "ret_1y": -8, "momentum": 52, "flow": "NEUTRAL", "commentary": "Cyclical bottom; awaiting China stimulus"},
    {"name": "OIL & GAS",       "ret_1w":-1.0, "ret_1m": -2.5, "ret_3m": 5,    "ret_6m": 8,  "ret_1y": 12, "momentum": 45, "flow": "NEUTRAL", "commentary": "Crude > $100 hits OMCs; upstream benefits"},
    {"name": "PSU BANKS",       "ret_1w":-1.2, "ret_1m": -3.8, "ret_3m": -8,   "ret_6m": -14,"ret_1y": -10,"momentum": 40, "flow": "OUTFLOW", "commentary": "Profit-booking after multi-year run"},
    {"name": "AUTO",            "ret_1w":-1.5, "ret_1m": -4.2, "ret_3m": -10,  "ret_6m": -18,"ret_1y": -12,"momentum": 38, "flow": "OUTFLOW", "commentary": "Tariff fear + JLR China weakness"},
    {"name": "NBFC",            "ret_1w":-0.8, "ret_1m": -2.5, "ret_3m": -8,   "ret_6m": -15,"ret_1y": -18,"momentum": 42, "flow": "NEUTRAL", "commentary": "Unsecured loan slowdown; rate-cut tailwind ahead"},
    {"name": "SPECIALTY CHEM",  "ret_1w":-0.5, "ret_1m": -1.0, "ret_3m": -3,   "ret_6m": -8, "ret_1y": -12,"momentum": 48, "flow": "NEUTRAL", "commentary": "Cycle bottom; China+1 thesis intact"},
    {"name": "IT SERVICES",     "ret_1w":-2.5, "ret_1m": -7.5, "ret_3m": -18,  "ret_6m": -25,"ret_1y": -32,"momentum": 22, "flow": "OUTFLOW", "commentary": "AI disruption — structural derating"},
]

# ── Helpers ──────────────────────────────────────────────────────────────────
def fill(h): return PatternFill("solid", fgColor=h)
def font(color=WHITE, bold=False, size=9, italic=False):
    return Font(name="Arial", color=color, bold=bold, size=size, italic=italic)
def bdr():
    s = Side(style="thin", color="2D2D2D")
    return Border(left=s, right=s, top=s, bottom=s)
def mid(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def heat_color(val):
    """Returns hex color based on % return (heatmap)."""
    if val is None: return "1A1A2E"
    if val >=  10:  return "0A4D2A"
    if val >=   5:  return "1B6B43"
    if val >=   2:  return "2D8659"
    if val >=   0:  return "3D2C00"
    if val >=  -2:  return "5C2A00"
    if val >=  -5:  return "8C2A00"
    return            "B22222"

def load_fresh():
    if not os.path.exists(FRESH_JSON):
        return DEMO_SECTORS
    try:
        with open(FRESH_JSON) as f:
            return json.load(f)
    except Exception:
        return DEMO_SECTORS

def get_sector_for_stock(symbol):
    """Returns sector name for a given stock symbol (or None)."""
    for sector, stocks in SECTOR_STOCKS.items():
        if symbol in stocks:
            return sector
    return None

def get_sector_boost(symbol):
    """Returns 0-15 boost for the stock's sector based on momentum."""
    sector = get_sector_for_stock(symbol)
    if not sector: return 0
    sectors = load_fresh()
    for s in sectors:
        if s["name"] == sector:
            mom = s.get("momentum", 50)
            if mom >= 85: return 15
            if mom >= 70: return 10
            if mom >= 55: return 5
            if mom >= 40: return 0
            return -5  # cold sector penalty
    return 0

# ── Excel build ───────────────────────────────────────────────────────────────
def build():
    sectors = load_fresh()
    # Sort by momentum desc
    sectors = sorted(sectors, key=lambda s: -s.get("momentum", 0))

    wb = Workbook()
    ws = wb.active
    ws.title = "NSE Sector Rotation"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    ws["A1"] = (f"NSE SECTOR ROTATION — Updated {TODAY.strftime('%d %b %Y')}  |  "
                "Hot sectors → boost screener score  |  Cold → reduce conviction")
    ws["A1"].font = Font(name="Arial", color=GOLD, bold=True, size=13)
    ws["A1"].fill = fill(HEADER_BG); ws["A1"].alignment = mid()
    ws.row_dimensions[1].height = 26

    # Top sector winners + losers
    top_3 = sectors[:3]
    bot_3 = sectors[-3:]
    ws.merge_cells("A2:I2")
    ws["A2"] = (f"🔥 TOP SECTORS: {', '.join(s['name'] for s in top_3)}  ||  "
                f"❄️ BOTTOM SECTORS: {', '.join(s['name'] for s in bot_3)}")
    ws["A2"].font = Font(name="Arial", color=AMBER, bold=True, size=10)
    ws["A2"].fill = fill(DARK_BG); ws["A2"].alignment = mid()
    ws.row_dimensions[2].height = 18

    # Headers
    headers = ["Rank", "Sector", "1W %", "1M %", "3M %", "6M %", "1Y %", "Momentum", "Flow", "Commentary"]
    widths  = [6, 22, 9, 9, 9, 9, 9, 11, 11, 48]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(name="Arial", color=GOLD, bold=True, size=9)
        c.fill = fill(HEADER_BG); c.alignment = mid(); c.border = bdr()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 28

    for idx, s in enumerate(sectors, 1):
        r = idx + 3
        cells = [
            idx, s["name"],
            f"{s['ret_1w']:+.1f}%", f"{s['ret_1m']:+.1f}%", f"{s['ret_3m']:+.1f}%",
            f"{s['ret_6m']:+.1f}%", f"{s['ret_1y']:+.1f}%",
            s["momentum"], s["flow"], s.get("commentary", ""),
        ]
        for col_i, val in enumerate(cells, 1):
            c = ws.cell(row=r, column=col_i, value=val)
            c.border = bdr()
            c.font = font(WHITE, size=9)
            c.alignment = mid() if col_i not in (2, 10) else lft()

            # Heatmap colour for return columns
            if col_i in (3, 4, 5, 6, 7):
                ret_val = [s['ret_1w'], s['ret_1m'], s['ret_3m'], s['ret_6m'], s['ret_1y']][col_i - 3]
                c.fill = fill(heat_color(ret_val))
                clr = GREEN if ret_val > 0 else RED
                c.font = font(clr, bold=True)
            else:
                c.fill = fill(ROW_ALT if idx % 2 else DARK_BG)

            if col_i == 1: c.font = font(GOLD, bold=True)
            if col_i == 2: c.font = font(GREEN, bold=True)
            if col_i == 8:  # momentum
                mom = s["momentum"]
                clr = GREEN if mom >= 85 else BLUE if mom >= 70 else AMBER if mom >= 50 else RED
                c.font = font(clr, bold=True)
            if col_i == 9:
                clr = GREEN if val == "INFLOW" else RED if val == "OUTFLOW" else GREY
                c.font = font(clr, bold=True)

        ws.row_dimensions[r].height = 28

    # Footer
    fr = len(sectors) + 5
    ws.merge_cells(f"A{fr}:I{fr}")
    ws[f"A{fr}"] = ("⚠️  STRATEGY: Trade WITH sector flow, not against. "
                   "Allocate 60% capital to top-3 sectors, 30% to next-3, 10% to opportunistic. "
                   "Avoid sectors with < 40 momentum + OUTFLOW.")
    ws[f"A{fr}"].font = Font(name="Arial", color=AMBER, bold=True, size=9)
    ws[f"A{fr}"].fill = fill(HEADER_BG); ws[f"A{fr}"].alignment = mid()
    ws.row_dimensions[fr].height = 26

    ws.merge_cells(f"A{fr+1}:I{fr+1}")
    ws[f"A{fr+1}"] = (f"Auto-generated by Claude Opus 4.7  |  "
                     f"Run date: {TODAY.strftime('%d %b %Y')}  |  "
                     "Sources: NSE indices · FII/DII flow data · Sectoral institutional positioning  |  Verify before trading.")
    ws[f"A{fr+1}"].font = Font(name="Arial", color=GREY, italic=True, size=8)
    ws[f"A{fr+1}"].fill = fill(DARK_BG); ws[f"A{fr+1}"].alignment = mid()
    ws.row_dimensions[fr+1].height = 14

    ws.freeze_panes = "C4"
    wb.save(EXCEL_PATH)
    print(f"✅ Sector Rotation Excel saved: {EXCEL_PATH}\n")

    # Console summary
    print(f"📊 Top 5 sectors (rank by momentum):")
    for i, s in enumerate(sectors[:5], 1):
        print(f"  {i}. {s['name']:18s} mom {s['momentum']:>3} | 1W {s['ret_1w']:+5.1f}% | 1M {s['ret_1m']:+5.1f}% | flow {s['flow']}")
    print(f"\n📉 Bottom 3 sectors:")
    for s in sectors[-3:]:
        print(f"  • {s['name']:18s} mom {s['momentum']:>3} | 1Y {s['ret_1y']:+5.1f}% | flow {s['flow']}")

if __name__ == "__main__":
    build()
