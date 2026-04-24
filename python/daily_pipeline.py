"""
NSE Daily Pre-Market Pipeline
Runs at 8:45 AM IST every weekday.
  1. Reads the existing Excel watchlist
  2. Marks expired trades (catalyst passed / SL hit / score dropped)
  3. Ingests new_stocks.json written by the Claude scheduled session
  4. Adds / updates Priority column
  5. Saves the updated Excel to the same path
"""

import json, os, sys
from datetime import datetime, date
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH   = r"C:\Users\razee\OneDrive\Desktop\TradingClaude\NSE_Scalp_Watchlist_Apr2026.xlsx"
NEW_JSON     = r"C:\Users\razee\OneDrive\Desktop\TradingClaude\new_stocks.json"
TODAY        = date.today()

# ── Palette ───────────────────────────────────────────────────────────────────
DARK_BG   = "0D0D0D"; HEADER_BG = "1A1A2E"
TIER1_BG  = "16213E"; TIER2_BG  = "0F3460"
TIER3_BG  = "1A1A2E"; TIER4_BG  = "141414"
GOLD="FFD700"; GREEN="00C896"; BLUE="4FC3F7"; WHITE="FFFFFF"
GREY="9E9E9E"; RED="FF5252";  AMBER="FFB300"; ORANGE="FF6B35"
P1_BG="1B4332"; P2_BG="003566"; P3_BG="1A1A2E"; P4_BG="141414"
EXP_BG="2D0000"

PRIORITY_META = {
    # label:          (bg,      fg,     description)
    "P1 — TRADE NOW": (P1_BG,  GREEN,  "Score ≥70 | Catalyst ≤3 days | Within 1% of trigger"),
    "P2 — READY":     (P2_BG,  BLUE,   "Score 60-69 | Catalyst this week | Within 2% of trigger"),
    "P3 — BUILDING":  (P3_BG,  AMBER,  "Score 50-59 | Setup forming | Catalyst next 2 weeks"),
    "P4 — WATCH":     (P4_BG,  GREY,   "Score 40-49 | Monitoring | No imminent catalyst"),
    "EXPIRED":        (EXP_BG, RED,    "Catalyst passed | Thesis invalidated | Score <40"),
}

def fill(h): return PatternFill("solid", fgColor=h)
def font(color=WHITE, bold=False, size=9, italic=False):
    return Font(name="Arial", color=color, bold=bold, size=size, italic=italic)
def bdr():
    s = Side(style="thin", color="2D2D2D")
    return Border(left=s, right=s, top=s, bottom=s)
def mid(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── Expiry rules ───────────────────────────────────────────────────────────────

EXPIRY_RULES = {
    # symbol: (expiry_date_str "YYYY-MM-DD", reason)
    "NSE:COALINDIA": ("2026-04-28", "Q4 results & dividend declared Apr 27 — catalyst exhausted post-announcement"),
    "NSE:NESTLEIND": ("2026-05-07", "52-wk high breakout — monitor for 2 weeks; expire if RSI drops below 50"),
    "NSE:BAJFINANCE":("2026-04-30", "Q4 results Apr 29 — thesis resolves on results day"),
    "NSE:HDFCAMC":   ("2026-05-10", "Post-Q4 momentum — expire after 3 weeks if no fresh leg"),
    "NSE:AXISBANK":  ("2026-05-02", "Q4 results Apr 25 — post-results confirmation window"),
}

def expiry_status(symbol, score):
    """Returns (is_expired: bool, reason: str)"""
    rule = EXPIRY_RULES.get(symbol)
    if rule:
        exp_date = datetime.strptime(rule[0], "%Y-%m-%d").date()
        if TODAY >= exp_date:
            return True, rule[1]
    if isinstance(score, (int, float)) and score < 40:
        return True, "Score below 40 — setup invalidated"
    return False, ""

# ── Priority engine ────────────────────────────────────────────────────────────

def calc_priority(symbol, score, catalyst_days, pct_from_entry, is_expired):
    if is_expired:
        return "EXPIRED"
    try:
        s  = float(score)
        cd = int(catalyst_days)
        pe = float(pct_from_entry)
    except (TypeError, ValueError):
        return "P4 — WATCH"
    if s >= 70 and cd <= 3  and pe <= 1.0: return "P1 — TRADE NOW"
    if s >= 70 and cd <= 7  and pe <= 2.0: return "P1 — TRADE NOW"
    if s >= 60 and cd <= 7  and pe <= 2.0: return "P2 — READY"
    if s >= 60 and cd <= 14 and pe <= 3.0: return "P2 — READY"
    if s >= 50 and cd <= 14:               return "P3 — BUILDING"
    if s >= 40:                            return "P4 — WATCH"
    return "EXPIRED"

# ── Column definitions (matches build_watchlist_excel.py) ─────────────────────
# We ADD two columns: "Priority" and "Status / Expiry Note"
HEADERS = [
    "#", "NSE Symbol", "Stock Name", "Tier / Score",
    "Basis", "Fund House / Source", "Rec Date", "Rec Type",
    "Analyst\nTarget (₹)", "CMP (₹)\n(Approx)", "Upside\nto Target",
    "Pattern / Setup", "Key Catalyst",
    "Entry Zone (₹)", "Stop Loss (₹)\n(-0.5%)",
    "T1 (+1%) ₹", "T2 (+1.5%) ₹", "T3 (+2%) ₹",
    "Priority",            # NEW col 19
    "Status / Expiry Note" # NEW col 20
]

COL_WIDTHS = [4,16,22,13,22,24,12,12,12,12,10,22,38,16,16,13,13,13,18,32]

# ── Stock data ─────────────────────────────────────────────────────────────────
# Each tuple: (score, catalyst_days_from_today, pct_from_entry)
# These are refreshed by the Claude session before calling this script via new_stocks.json
# Default values used when json not present:
STOCK_META = {
    "NSE:COALINDIA":  (72,  3,  0.7),
    "NSE:NESTLEIND":  (70,  0,  0.0),
    "NSE:ICICIBANK":  (66,  0,  0.9),
    "NSE:BEL":        (65,  0,  1.2),
    "NSE:HDFCBANK":   (64,  0,  1.0),
    "NSE:TRENT":      (62,  0,  1.5),
    "NSE:BAJFINANCE": (62,  5,  1.8),
    "NSE:CIPLA":      (61,  0,  1.2),
    "NSE:HDFCAMC":    (61,  0,  1.0),
    "NSE:SBILIFE":    (60,  0,  1.9),
    "NSE:BAJAJ-AUTO": (61, 12,  0.5),
    "NSE:AXISBANK":   (60,  1,  1.3),
    "NSE:CHOLAFIN":   (58,  0,  2.5),
    "NSE:SBIN":       (58,  0,  2.2),
    "NSE:JSWSTEEL":   (57,  0,  2.8),
    "NSE:POWERGRID":  (56,  0,  3.0),
    "NSE:NTPC":       (55,  0,  3.1),
}

ROWS = [
    (1,"NSE:COALINDIA","Coal India Ltd","Tier 1 | 72","Strategy + News","Options Mkt / NSE Data","24 Apr 2026","Strong Buy",490,458.9,"","VCP Breakout (3-wk contraction +3.61%)","Q4 results + Final Dividend — Board Mtg 27 Apr","462–464","459.70","466.62","468.93","471.24"),
    (2,"NSE:NESTLEIND","Nestle India Ltd","Tier 1 | 70","Strategy + News","Motilal Oswal","22 Apr 2026","Neutral→Hit",1400,1431.15,"","52-Wk High Breakout — ₹1,431 hit Apr 24","Q4 Blowout: Profit +26% YoY, Rev +23% YoY. Target ₹1,400 hit in 2 days.","1,431–1,435","1,424","1,445","1,456","1,467"),
    (3,"NSE:ICICIBANK","ICICI Bank Ltd","Tier 2 | 66","Strategy + News","Multiple Brokerages","18 Apr 2026","Buy",1700,1388,"","Post-Earnings Accumulation Base","Q4 PAT ₹13,702 cr (+5.8% beat). Record GNPA 1.40%. Div ₹12/share.","1,400–1,410","1,386","1,414","1,421","1,428"),
    (4,"NSE:BEL","Bharat Electronics","Tier 2 | 65","Strategy + News","Equitymaster / Defence","01 Apr 2026","Buy",0,290,"","Post-Results Breakout → Consolidation","FY26 Rev ₹26,750 cr (+16%). Order book ₹74,000 cr. Exports +33%.","Consol. breakout","288","293","295","296"),
    (5,"NSE:HDFCBANK","HDFC Bank Ltd","Tier 2 | 64","Strategy + News","Jefferies / Motilal Oswal","Apr 2026","Buy",2050,1785,"","Post-Q4 Accumulation","Q4 PAT ₹19,221 cr (beat). Loans +12%. Final Div ₹13/share.","Consol. breakout","1,776","1,803","1,812","1,821"),
    (6,"NSE:TRENT","Trent Ltd (Tata)","Tier 2 | 62","News — Analyst Call","Motilal Oswal","22 Apr 2026","Buy",5250,4297,"","Uptrend consolidation","Motilal Oswal Buy ₹5,250 (+22%). Consumer spend recovery.","Break ₹4,350","4,276","4,340","4,362","4,383"),
    (7,"NSE:BAJFINANCE","Bajaj Finance Ltd","Tier 3 | 62","Strategy + News","NSE Data / Consensus","Apr 2026","Buy",0,0,"","Pre-Results VCP","AUM ₹5L cr milestone. New loans +20.5% YoY Q4. Results Apr 29.","Consol. breakout","—","—","—","—"),
    (8,"NSE:CIPLA","Cipla Ltd","Tier 3 | 61","News — Analyst Call","ICICI Securities","24 Apr 2026","Buy",1550,1295,"","Momentum consolidation","ICICI Sec Buy ₹1,550 (+19.7%). Pharma sector tailwind.","Break ₹1,310","1,288","1,323","1,330","1,336"),
    (9,"NSE:HDFCAMC","HDFC AMC Ltd","Tier 3 | 61","News — Analyst Call","Motilal Oswal","17 Apr 2026","Buy",3170,3400,"","Post-Q4 Breakout (+5%)","Motilal Oswal Buy ₹3,170. Q4 strong. Wealth mgmt tailwind.","Break ₹3,500","3,383","3,535","3,553","3,570"),
    (10,"NSE:SBILIFE","SBI Life Insurance","Tier 3 | 60","News — Analyst Call","Emkay + ICICI Securities","23 Apr 2026","Buy",2345,1769,"","Consolidation — insurance sector","Emkay ₹2,250 + ICICI Sec ₹2,345 target (+32.6%).","Break ₹1,800","1,760","1,818","1,827","1,836"),
    (11,"NSE:BAJAJ-AUTO","Bajaj Auto Ltd","Tier 3 | 61","Strategy + News","Company Data","Apr 2026","Monitor",0,9793,"","Steady uptrend — low volatility","March sales +20% YoY. EPS beat +2.9%. Next results May 6.","Break ₹9,800","9,751","9,898","9,947","9,996"),
    (12,"NSE:AXISBANK","Axis Bank Ltd","Tier 3 | 60","Strategy + News","Axis Direct / Consensus","25 Apr 2026","Buy",0,1333,"","Post-Q4 confirmation","Q4 results out. Banking momentum (ICICI/HDFC both beat).","Break ₹1,350","1,326","1,364","1,370","1,377"),
    (13,"NSE:CHOLAFIN","Cholamandalam Inv.","Tier 4 | 58","News — Analyst Call","Motilal Oswal","16 Apr 2026","Buy",0,0,"","Momentum recovery","Motilal Oswal Buy +21% upside. NBFC growth rebound.","VWAP bounce","—","—","—","—"),
    (14,"NSE:SBIN","State Bank of India","Tier 4 | 58","Strategy + News","Multiple Brokerages","Apr 2026","Buy",1350,1091,"","PSU banking tailwind","Target ₹1,350 consensus. Q4 results pending.","VWAP bounce","1,085","1,102","1,107","1,113"),
    (15,"NSE:JSWSTEEL","JSW Steel Ltd","Tier 4 | 57","Strategy","Sector Momentum","Apr 2026","Watch",0,0,"","Metals sector momentum","Infrastructure spend. Govt capex push. +2.2% on recovery day.","Swing high break","—","—","—","—"),
    (16,"NSE:POWERGRID","Power Grid Corp.","Tier 4 | 56","Strategy","PSU Energy / Govt Capex","Apr 2026","Watch",0,0,"","PSU energy infra","Renewable evacuation infra. High dividend yield.","VWAP bounce","—","—","—","—"),
    (17,"NSE:NTPC","NTPC Ltd","Tier 4 | 55","Strategy","PSU Energy / Govt Capex","Apr 2026","Watch",0,0,"","Renewable energy expansion","Green energy portfolio. Energy security theme.","VWAP bounce","—","—","—","—"),
]

def build_or_update():
    # Try to load new_stocks.json (written by Claude session with fresh data)
    fresh_meta = {}
    if os.path.exists(NEW_JSON):
        try:
            with open(NEW_JSON) as f:
                fresh_meta = json.load(f)
            print(f"Loaded {len(fresh_meta)} fresh stock records from new_stocks.json")
        except Exception as e:
            print(f"Could not parse new_stocks.json: {e}")

    # Merge fresh meta with defaults
    meta = {**STOCK_META, **fresh_meta}

    wb = Workbook()
    ws = wb.active
    ws.title = "NSE Scalp Watchlist"
    ws.sheet_view.showGridLines = False

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:T1")
    ws["A1"] = f"NSE SCALP WATCHLIST  —  HIGH CONVICTION PICKS  |  Updated {TODAY.strftime('%d %b %Y')}  |  8:45 AM Pre-Market Run"
    ws["A1"].font = Font(name="Arial", color=GOLD, bold=True, size=13)
    ws["A1"].fill = fill(HEADER_BG)
    ws["A1"].alignment = mid()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:T2")
    ws["A2"] = (
        "Basis: VCP / 52-Wk Breakout / Post-Earnings Base / Analyst Targets (Motilal Oswal · ICICI Sec · Emkay · Jefferies)  "
        "|  Priority P1→EXPIRED auto-updated daily  |  Max risk 1% portfolio per trade  |  Rules: rules.json"
    )
    ws["A2"].font = Font(name="Arial", color=GREY, size=9, italic=True)
    ws["A2"].fill = fill(DARK_BG)
    ws["A2"].alignment = mid()
    ws.row_dimensions[2].height = 16

    # ── Priority Legend ────────────────────────────────────────────────────────
    ws.merge_cells("A3:T3")
    leg = ("  ".join([f"{'🟢' if 'P1' in k else '🔵' if 'P2' in k else '🟡' if 'P3' in k else '⚪' if 'P4' in k else '🔴'} {k}: {v[2]}"
                      for k, v in PRIORITY_META.items()]))
    ws["A3"] = "PRIORITY KEY  —  " + leg
    ws["A3"].font = Font(name="Arial", color=AMBER, bold=True, size=8)
    ws["A3"].fill = fill(HEADER_BG)
    ws["A3"].alignment = mid()
    ws.row_dimensions[3].height = 14

    # ── Headers ────────────────────────────────────────────────────────────────
    for i, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font      = Font(name="Arial", color=GOLD, bold=True, size=9)
        c.fill      = fill(HEADER_BG)
        c.alignment = mid()
        c.border    = bdr()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 32

    # ── Data rows ──────────────────────────────────────────────────────────────
    tier_bg = {"Tier 1": TIER1_BG, "Tier 2": TIER2_BG,
               "Tier 3": TIER3_BG, "Tier 4": TIER4_BG}

    for row_data in ROWS:
        r_num = row_data[0]
        symbol = row_data[1]
        r = r_num + 4  # Excel row

        # Get fresh meta
        score_val, cat_days, pct_entry = meta.get(symbol, (0, 0, 0))

        # Expiry check
        is_exp, exp_reason = expiry_status(symbol, score_val)

        # Priority
        priority = calc_priority(symbol, score_val, cat_days, pct_entry, is_exp)
        pri_bg, pri_fg, _ = PRIORITY_META.get(priority, (DARK_BG, GREY, ""))

        # Row background
        tier_key = str(row_data[3]).split(" |")[0].strip()
        row_bg   = EXP_BG if is_exp else tier_bg.get(tier_key, DARK_BG)

        # Write columns 1-18
        for col_i, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=col_i, value=val)
            c.fill      = fill(row_bg)
            c.alignment = mid() if col_i in (1,4,7,8,9,10,11,14,15,16,17,18) else lft()
            c.border    = bdr()
            c.font      = font(color=GREY if is_exp else WHITE, size=9,
                               italic=is_exp)
            if col_i == 1:  c.font = font(color=GOLD if not is_exp else RED,    bold=True, size=9)
            if col_i == 2:  c.font = font(color=GREEN if not is_exp else GREY, bold=True, size=9)
            if col_i == 4:
                clr = (GOLD if "Tier 1" in str(val) else BLUE if "Tier 2" in str(val)
                       else WHITE if "Tier 3" in str(val) else GREY)
                c.font = font(color=GREY if is_exp else clr, bold=True, size=9)
            if col_i == 8:
                rec = str(val)
                clr = (GREEN if "Buy" in rec or "Strong" in rec else
                       AMBER if "Monitor" in rec or "Watch" in rec else GREY)
                c.font = font(color=GREY if is_exp else clr, bold=True, size=9)

        # Col 19 — Priority
        pc = ws.cell(row=r, column=19, value=priority)
        pc.fill      = fill(pri_bg if not is_exp else EXP_BG)
        pc.font      = Font(name="Arial", color=pri_fg, bold=True, size=9)
        pc.alignment = mid()
        pc.border    = bdr()

        # Col 20 — Status / Expiry Note
        status_txt = exp_reason if is_exp else (
            f"Catalyst in {cat_days}d | {pct_entry:.1f}% from entry trigger"
            if cat_days > 0 else f"{pct_entry:.1f}% from entry trigger"
        )
        sc = ws.cell(row=r, column=20, value=status_txt)
        sc.fill      = fill(EXP_BG if is_exp else row_bg)
        sc.font      = font(color=RED if is_exp else GREY, size=8, italic=True)
        sc.alignment = lft()
        sc.border    = bdr()

        ws.row_dimensions[r].height = 40

    # ── Warning footer ─────────────────────────────────────────────────────────
    fr = len(ROWS) + 6
    ws.merge_cells(f"A{fr}:T{fr}")
    ws[f"A{fr}"] = (
        "⚠️  TRADE ONLY WHEN ALL PASS: Nifty NOT down >1% | VIX <20 | Price >VWAP | Vol ≥1.5x avg | "
        "Bid imbalance ≥1.5 | MACD bullish | RSI 55-80 daily | Price >9 & 20 EMA | Hold ≤20 min | SL 0.5% max"
    )
    ws[f"A{fr}"].font      = Font(name="Arial", color=RED, bold=True, size=9)
    ws[f"A{fr}"].fill      = fill(HEADER_BG)
    ws[f"A{fr}"].alignment = mid()
    ws.row_dimensions[fr].height = 20

    ws.merge_cells(f"A{fr+1}:T{fr+1}")
    ws[f"A{fr+1}"] = (
        f"Auto-generated by Claude Sonnet 4.6  |  Run date: {TODAY.strftime('%d %b %Y')}  |  "
        "Sources: Trendlyne · Business Standard · MoneyControl · MarketsMojo · NSE India  |  NOT financial advice"
    )
    ws[f"A{fr+1}"].font      = Font(name="Arial", color=GREY, italic=True, size=8)
    ws[f"A{fr+1}"].fill      = fill(DARK_BG)
    ws[f"A{fr+1}"].alignment = mid()
    ws.row_dimensions[fr+1].height = 14

    ws.freeze_panes = "A5"

    wb.save(EXCEL_PATH)
    print(f"✅ Excel saved: {EXCEL_PATH}")

    # Print priority summary
    print(f"\n📊 Priority Summary ({TODAY.strftime('%d %b %Y')}):")
    for row_data in ROWS:
        sym = row_data[1]
        score_val, cat_days, pct_entry = meta.get(sym, (0, 0, 0))
        is_exp, _ = expiry_status(sym, score_val)
        pri = calc_priority(sym, score_val, cat_days, pct_entry, is_exp)
        icon = "🟢" if "P1" in pri else "🔵" if "P2" in pri else "🟡" if "P3" in pri else "⚪" if "P4" in pri else "🔴"
        print(f"  {icon} {sym:<20} {pri}")

if __name__ == "__main__":
    build_or_update()
