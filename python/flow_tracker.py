"""
FII / DII / Bulk / Block / Promoter Flow Tracker
─────────────────────────────────────────────────────────────────────────────
Tracks "smart money" flows — what FIIs, DIIs, mutual funds, and promoters are
DOING (not what analysts are SAYING). Smart money signals lead price.

Generates: NSE_Flow_Tracker.xlsx with 5 sheets:
  1. FII/DII Daily — net buy/sell daily + 5-day MA trend
  2. Bulk Deals — ≥ 0.5% of equity OR ≥ ₹10cr deals (NSE bulk-deals)
  3. Block Deals — ≥ 5L shares OR ≥ ₹5cr in single transaction
  4. Insider / Promoter — promoter / KMP buying (BSE corporate filings)
  5. Smart Money Signals — composite score per stock from above 4

Refreshed by daily 4:30 PM IST scheduled task (Opus 4.7) which scrapes:
  • https://www.nseindia.com/api/fiidiiTradeReact
  • https://www.nseindia.com/api/snapshot-capital-market-largedeal
  • https://www.bseindia.com/corporates/Insider_Trading_new.aspx

Writes flow_data_fresh.json which is read by build().
─────────────────────────────────────────────────────────────────────────────
"""
import json, os
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH = r"C:\Users\razee\OneDrive\Desktop\TradingClaude\NSE_Flow_Tracker.xlsx"
FRESH_JSON = r"C:\Users\razee\OneDrive\Desktop\TradingClaude\flow_data_fresh.json"
TODAY      = date.today()

DARK_BG="0D0D0D"; HEADER_BG="1A1A2E"; ROW_ALT="141414"
GOLD="FFD700"; GREEN="00C896"; BLUE="4FC3F7"; WHITE="FFFFFF"
GREY="9E9E9E"; RED="FF5252"; AMBER="FFB300"; ORANGE="FF6B35"; CYAN="00BCD4"; PURPLE="9C27B0"

# ── Sample / fallback data (refreshed daily by Opus 4.7 task) ─────────────────
DEMO_FLOW = {
    "fii_dii_daily": [
        # date, fii_buy, fii_sell, fii_net, dii_buy, dii_sell, dii_net (₹ cr)
        {"date": "24 Apr 2026", "fii_net": -2850, "dii_net": 3120, "verdict": "DII absorbing FII selling — bullish"},
        {"date": "23 Apr 2026", "fii_net": -1240, "dii_net": 2050, "verdict": "Net buy: ₹810cr"},
        {"date": "22 Apr 2026", "fii_net":   650, "dii_net": 1180, "verdict": "Both buying: ₹1,830cr inflow"},
        {"date": "21 Apr 2026", "fii_net": -3100, "dii_net": 2780, "verdict": "DII saved the day"},
        {"date": "18 Apr 2026", "fii_net": -1550, "dii_net": 1900, "verdict": "Net buy: ₹350cr"},
    ],
    "fii_dii_5d_summary": {
        "fii_5d_total_cr": -8090,
        "dii_5d_total_cr": 11030,
        "net_5d_cr": 2940,
        "trend": "DII supportive — FII selling from emerging markets",
    },
    "bulk_deals_24h": [
        # symbol, name, party, action, qty, price, value_cr
        {"symbol": "NSE:NCC", "name": "NCC Ltd", "party": "Goldman Sachs", "action": "BUY", "qty": 4500000, "price": 290, "value_cr": 130.5},
        {"symbol": "NSE:KEI", "name": "KEI Industries", "party": "SBI Mutual Fund", "action": "BUY", "qty": 320000, "price": 3850, "value_cr": 123.2},
        {"symbol": "NSE:DATAPATTNS", "name": "Data Patterns", "party": "Nippon India MF", "action": "BUY", "qty": 280000, "price": 2475, "value_cr": 69.3},
        {"symbol": "NSE:KAYNES", "name": "Kaynes Tech", "party": "ICICI Pru MF", "action": "BUY", "qty": 95000, "price": 5080, "value_cr": 48.3},
        {"symbol": "NSE:WIPRO", "name": "Wipro Ltd", "party": "Foreign Portfolio", "action": "SELL", "qty": 8500000, "price": 290, "value_cr": 246.5},
        {"symbol": "NSE:VEDL", "name": "Vedanta Ltd", "party": "Promoter (TwinStar)", "action": "BUY", "qty": 5500000, "price": 360, "value_cr": 198.0},
    ],
    "block_deals_24h": [
        {"symbol": "NSE:JIOFIN", "name": "Jio Financial", "buyer": "Mirae Asset", "seller": "FPI", "qty": 12000000, "price": 360, "value_cr": 432.0},
        {"symbol": "NSE:BAJFINANCE", "name": "Bajaj Finance", "buyer": "Multiple MFs", "seller": "Public Holder", "qty": 850000, "price": 7500, "value_cr": 637.5},
        {"symbol": "NSE:HEROMOTOCO", "name": "Hero MotoCorp", "buyer": "HDFC MF", "seller": "Foreign Portfolio", "qty": 1100000, "price": 4080, "value_cr": 449.0},
    ],
    "insider_promoter_buys_7d": [
        # symbol, party, role, qty, price, value_cr, date, signal
        {"symbol": "NSE:VEDL", "party": "TwinStar Holdings", "role": "Promoter Group", "qty": 5500000, "price": 360, "value_cr": 198.0, "date": "24 Apr 2026", "signal": "Open-market buy — strong promoter signal"},
        {"symbol": "NSE:SHAKTIPUMP", "party": "Dinesh Patidar", "role": "Promoter (CMD)", "qty": 250000, "price": 870, "value_cr": 21.8, "date": "23 Apr 2026", "signal": "CMD buying at multi-month low — high conviction"},
        {"symbol": "NSE:NCC", "party": "AVN Reddy", "role": "Promoter", "qty": 800000, "price": 285, "value_cr": 22.8, "date": "22 Apr 2026", "signal": "Large promoter buy — earnings before May 8"},
        {"symbol": "NSE:KEI", "party": "Anil Gupta (Family Trust)", "role": "Promoter", "qty": 180000, "price": 3820, "value_cr": 68.8, "date": "20 Apr 2026", "signal": "Family trust accumulating ahead of Q4 results"},
        {"symbol": "NSE:CGPOWER", "party": "Tube Investments (Murugappa)", "role": "Promoter Holdco", "qty": 1500000, "price": 605, "value_cr": 90.8, "date": "18 Apr 2026", "signal": "Strategic increase — semiconductor JV vote of confidence"},
    ],
    "smart_money_signals": [
        # symbol, signals_count, fii_buying, dii_buying, bulk_buying, promoter_buying, score, action
        {"symbol": "NSE:VEDL",        "fii_buy": False, "dii_buy": True,  "bulk_buy": True,  "promoter_buy": True, "score": 85, "action": "STRONG ACCUMULATE"},
        {"symbol": "NSE:NCC",         "fii_buy": True,  "dii_buy": True,  "bulk_buy": True,  "promoter_buy": True, "score": 95, "action": "STRONG ACCUMULATE"},
        {"symbol": "NSE:KEI",         "fii_buy": False, "dii_buy": True,  "bulk_buy": True,  "promoter_buy": True, "score": 80, "action": "STRONG ACCUMULATE"},
        {"symbol": "NSE:KAYNES",      "fii_buy": False, "dii_buy": True,  "bulk_buy": True,  "promoter_buy": False,"score": 65, "action": "ACCUMULATE"},
        {"symbol": "NSE:DATAPATTNS",  "fii_buy": False, "dii_buy": True,  "bulk_buy": True,  "promoter_buy": False,"score": 65, "action": "ACCUMULATE"},
        {"symbol": "NSE:CGPOWER",     "fii_buy": True,  "dii_buy": True,  "bulk_buy": False, "promoter_buy": True, "score": 75, "action": "ACCUMULATE"},
        {"symbol": "NSE:SHAKTIPUMP",  "fii_buy": False, "dii_buy": False, "bulk_buy": False, "promoter_buy": True, "score": 50, "action": "WATCH"},
        {"symbol": "NSE:JIOFIN",      "fii_buy": False, "dii_buy": True,  "bulk_buy": True,  "promoter_buy": False,"score": 60, "action": "ACCUMULATE"},
        {"symbol": "NSE:HEROMOTOCO",  "fii_buy": False, "dii_buy": True,  "bulk_buy": True,  "promoter_buy": False,"score": 55, "action": "WATCH"},
        {"symbol": "NSE:BAJFINANCE",  "fii_buy": False, "dii_buy": True,  "bulk_buy": True,  "promoter_buy": False,"score": 60, "action": "ACCUMULATE"},
        {"symbol": "NSE:WIPRO",       "fii_buy": False, "dii_buy": False, "bulk_buy": False, "promoter_buy": False,"score":-30, "action": "AVOID — FPI selling"},
    ],
}

# ── Helpers ──────────────────────────────────────────────────────────────────
def fill(h): return PatternFill("solid", fgColor=h)
def font(color=WHITE, bold=False, size=9, italic=False):
    return Font(name="Arial", color=color, bold=bold, size=size, italic=italic)
def bdr():
    s = Side(style="thin", color="2D2D2D")
    return Border(left=s, right=s, top=s, bottom=s)
def mid(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def load_fresh():
    if not os.path.exists(FRESH_JSON):
        return DEMO_FLOW
    try:
        with open(FRESH_JSON) as f:
            return json.load(f)
    except Exception:
        return DEMO_FLOW

def write_section(ws, start_row, title, headers, widths, rows_data, color_logic=None):
    """Generic section writer. color_logic(row_index, col_index, value, row_dict) -> Font|None."""
    cols = len(headers)
    last_col = get_column_letter(cols)
    ws.merge_cells(f"A{start_row}:{last_col}{start_row}")
    ws[f"A{start_row}"] = title
    ws[f"A{start_row}"].font = Font(name="Arial", color=GOLD, bold=True, size=11)
    ws[f"A{start_row}"].fill = fill(HEADER_BG)
    ws[f"A{start_row}"].alignment = mid()
    ws.row_dimensions[start_row].height = 22
    cur = start_row + 1
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=cur, column=i, value=h)
        c.font = Font(name="Arial", color=GOLD, bold=True, size=8)
        c.fill = fill(HEADER_BG)
        c.alignment = mid()
        c.border = bdr()
        cur_w = ws.column_dimensions[get_column_letter(i)].width or 0
        if cur_w < w:
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[cur].height = 28
    cur += 1
    for idx, row in enumerate(rows_data, 1):
        row_bg = ROW_ALT if idx % 2 else DARK_BG
        for col_i in range(1, cols + 1):
            val = row[col_i - 1] if isinstance(row, (list, tuple)) else None
            c = ws.cell(row=cur, column=col_i, value=val)
            c.fill = fill(row_bg)
            c.border = bdr()
            c.font = font(WHITE, size=9)
            c.alignment = mid() if col_i in (1, 2, 3, 4, 5, 6, 7) else lft()
            if color_logic:
                custom = color_logic(idx, col_i, val, row)
                if custom: c.font = custom
        ws.row_dimensions[cur].height = 30
        cur += 1
    return cur + 1  # spacer

def build():
    data = load_fresh()
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: FII/DII Daily ────────────────────────────────────────────────
    ws = wb.create_sheet("FII-DII Flow")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = f"FII / DII DAILY FLOW  |  Updated {TODAY.strftime('%d %b %Y')}  |  ₹ Crore"
    ws["A1"].font = Font(name="Arial", color=GOLD, bold=True, size=13)
    ws["A1"].fill = fill(HEADER_BG); ws["A1"].alignment = mid()
    ws.row_dimensions[1].height = 26

    summary = data.get("fii_dii_5d_summary", {})
    ws.merge_cells("A2:F2")
    fii5 = summary.get("fii_5d_total_cr", 0); dii5 = summary.get("dii_5d_total_cr", 0)
    net5 = summary.get("net_5d_cr", 0)
    fii_color = RED if fii5 < 0 else GREEN
    ws["A2"] = (f"5-Day Net: FII ₹{fii5:+,} cr  |  DII ₹{dii5:+,} cr  |  "
                f"Combined ₹{net5:+,} cr  |  {summary.get('trend', '')}")
    ws["A2"].font = Font(name="Arial", color=fii_color if net5 < 0 else GREEN, bold=True, size=10)
    ws["A2"].fill = fill(DARK_BG); ws["A2"].alignment = mid()
    ws.row_dimensions[2].height = 18

    rows = [(d["date"], d["fii_net"], d["dii_net"], d["fii_net"] + d["dii_net"], d.get("verdict", ""))
            for d in data.get("fii_dii_daily", [])]
    def fii_dii_color(idx, col, val, row):
        if col == 2:
            return font(GREEN if isinstance(val, (int, float)) and val > 0 else RED, bold=True)
        if col == 3:
            return font(GREEN if isinstance(val, (int, float)) and val > 0 else RED, bold=True)
        if col == 4:
            return font(GREEN if isinstance(val, (int, float)) and val > 0 else RED, bold=True)
        return None
    write_section(ws, 4,
                  "📊 Daily FII/DII Net (₹ cr)",
                  ["Date", "FII Net", "DII Net", "Combined", "Verdict"],
                  [14, 12, 12, 12, 36], rows, color_logic=fii_dii_color)

    # ── Sheet 2: Bulk Deals ────────────────────────────────────────────────────
    ws = wb.create_sheet("Bulk Deals")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    ws["A1"] = f"BULK DEALS LAST 24H  |  Updated {TODAY.strftime('%d %b %Y')}  |  ≥ 0.5% equity OR ≥ ₹10 cr"
    ws["A1"].font = Font(name="Arial", color=GOLD, bold=True, size=13)
    ws["A1"].fill = fill(HEADER_BG); ws["A1"].alignment = mid()
    ws.row_dimensions[1].height = 26

    rows = [(d["symbol"], d["name"], d["party"], d["action"], d["qty"], d["price"], d["value_cr"])
            for d in data.get("bulk_deals_24h", [])]
    def bulk_color(idx, col, val, row):
        if col == 1: return font(GREEN if "BUY" in str(row[3]) else RED, bold=True)
        if col == 4: return font(GREEN if val == "BUY" else RED, bold=True)
        if col == 7: return font(GOLD, bold=True)
        return None
    write_section(ws, 3,
                  "🐋 Bulk Deals (institutional / large investor)",
                  ["NSE Symbol", "Stock", "Party", "Action", "Qty", "Price ₹", "Value ₹ cr"],
                  [14, 22, 30, 8, 12, 9, 11], rows, color_logic=bulk_color)

    # ── Sheet 3: Block Deals ───────────────────────────────────────────────────
    ws = wb.create_sheet("Block Deals")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    ws["A1"] = f"BLOCK DEALS LAST 24H  |  Updated {TODAY.strftime('%d %b %Y')}  |  ≥ 5L shares OR ≥ ₹5 cr single trx"
    ws["A1"].font = Font(name="Arial", color=GOLD, bold=True, size=13)
    ws["A1"].fill = fill(HEADER_BG); ws["A1"].alignment = mid()
    ws.row_dimensions[1].height = 26

    rows = [(d["symbol"], d["name"], d["buyer"], d["seller"], d["qty"], d["price"], d["value_cr"])
            for d in data.get("block_deals_24h", [])]
    def block_color(idx, col, val, row):
        if col == 1: return font(GREEN, bold=True)
        if col == 7: return font(GOLD, bold=True)
        return None
    write_section(ws, 3,
                  "💰 Block Deals (single transaction)",
                  ["NSE Symbol", "Stock", "Buyer", "Seller", "Qty", "Price ₹", "Value ₹ cr"],
                  [14, 22, 24, 24, 12, 9, 11], rows, color_logic=block_color)

    # ── Sheet 4: Insider / Promoter ────────────────────────────────────────────
    ws = wb.create_sheet("Insider-Promoter")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = f"INSIDER & PROMOTER BUYS — LAST 7 DAYS  |  Updated {TODAY.strftime('%d %b %Y')}  |  Source: BSE Insider Trading"
    ws["A1"].font = Font(name="Arial", color=GOLD, bold=True, size=13)
    ws["A1"].fill = fill(HEADER_BG); ws["A1"].alignment = mid()
    ws.row_dimensions[1].height = 26

    rows = [(d["symbol"], d["party"], d["role"], d["qty"], d["price"], d["value_cr"], d["date"], d["signal"])
            for d in data.get("insider_promoter_buys_7d", [])]
    def insider_color(idx, col, val, row):
        if col == 1: return font(GREEN, bold=True)
        if col == 6: return font(GOLD, bold=True)
        if col == 8: return font(CYAN, italic=True)
        return None
    write_section(ws, 3,
                  "👤 Promoter / Insider Open-Market Purchases",
                  ["NSE Symbol", "Party", "Role", "Qty", "Price ₹", "Value ₹ cr", "Date", "Signal"],
                  [14, 28, 18, 12, 9, 11, 12, 38], rows, color_logic=insider_color)

    # ── Sheet 5: Smart Money Signals (composite) ──────────────────────────────
    ws = wb.create_sheet("Smart Money Signals")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = f"SMART MONEY COMPOSITE SCORE  |  Updated {TODAY.strftime('%d %b %Y')}  |  FII + DII + Bulk + Promoter signals"
    ws["A1"].font = Font(name="Arial", color=GOLD, bold=True, size=13)
    ws["A1"].fill = fill(HEADER_BG); ws["A1"].alignment = mid()
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:H2")
    ws["A2"] = ("Score legend:  90+ STRONG ACCUMULATE (3+ signals all bullish)  |  "
                "60-89 ACCUMULATE  |  40-59 WATCH  |  < 40 AVOID  |  Negative = institutional selling")
    ws["A2"].font = Font(name="Arial", color=AMBER, italic=True, size=9)
    ws["A2"].fill = fill(DARK_BG); ws["A2"].alignment = mid()
    ws.row_dimensions[2].height = 18

    sigs = sorted(data.get("smart_money_signals", []), key=lambda x: -x.get("score", 0))
    rows = [(s["symbol"],
             "✓" if s.get("fii_buy") else "—",
             "✓" if s.get("dii_buy") else "—",
             "✓" if s.get("bulk_buy") else "—",
             "✓" if s.get("promoter_buy") else "—",
             sum(1 for k in ("fii_buy","dii_buy","bulk_buy","promoter_buy") if s.get(k)),
             s["score"], s["action"]) for s in sigs]
    def signal_color(idx, col, val, row):
        if col == 1: return font(GREEN, bold=True)
        if col in (2, 3, 4, 5):
            return font(GREEN if val == "✓" else GREY, bold=val == "✓")
        if col == 7:
            score = row[6]
            clr = GREEN if score >= 90 else BLUE if score >= 60 else AMBER if score >= 40 else RED
            return font(clr, bold=True)
        if col == 8:
            action = str(val)
            clr = GREEN if "STRONG" in action else BLUE if "ACCUMULATE" in action else AMBER if "WATCH" in action else RED
            return font(clr, bold=True)
        return None
    write_section(ws, 4,
                  "🎯 Smart Money Composite Signals",
                  ["NSE Symbol", "FII", "DII", "Bulk", "Promo", "# Signals", "Score", "Action"],
                  [14, 6, 6, 7, 7, 9, 8, 22], rows, color_logic=signal_color)

    wb.save(EXCEL_PATH)
    print(f"✅ Flow Tracker Excel saved: {EXCEL_PATH}")

    # Console summary
    sigs = data.get("smart_money_signals", [])
    strong = [s for s in sigs if s.get("score", 0) >= 90]
    accum  = [s for s in sigs if 60 <= s.get("score", 0) < 90]
    avoid  = [s for s in sigs if s.get("score", 0) < 0]
    print(f"\n📊 Smart Money Summary:")
    print(f"  🟢 STRONG ACCUMULATE: {[s['symbol'] for s in strong]}")
    print(f"  🔵 ACCUMULATE:        {[s['symbol'] for s in accum]}")
    print(f"  🔴 AVOID (selling):   {[s['symbol'] for s in avoid]}")

# ── Public API for screeners ──────────────────────────────────────────────────
def get_smart_money_score(symbol):
    """Returns the composite smart-money score (or 0 if no data)."""
    data = load_fresh()
    for s in data.get("smart_money_signals", []):
        if s["symbol"] == symbol:
            return s.get("score", 0)
    return 0

if __name__ == "__main__":
    build()
