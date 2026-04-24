from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "NSE Scalp Watchlist"

# ── Palette ──────────────────────────────────────────────────────────────────
DARK_BG    = "0D0D0D"
HEADER_BG  = "1A1A2E"
TIER1_BG   = "16213E"
TIER2_BG   = "0F3460"
TIER3_BG   = "1A1A2E"
TIER4_BG   = "141414"
GOLD       = "FFD700"
GREEN      = "00C896"
BLUE_LINK  = "4FC3F7"
ORANGE     = "FF6B35"
WHITE      = "FFFFFF"
GREY       = "9E9E9E"
RED_LIGHT  = "FF5252"
AMBER      = "FFB300"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(color=WHITE, bold=False, size=10, italic=False):
    return Font(name="Arial", color=color, bold=bold, size=size, italic=italic)

def border():
    s = Side(style="thin", color="2D2D2D")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Sheet background ──────────────────────────────────────────────────────────
ws.sheet_view.showGridLines = False

# ── Title block ───────────────────────────────────────────────────────────────
ws.merge_cells("A1:R1")
ws["A1"] = "NSE SCALP WATCHLIST  —  HIGH CONVICTION PICKS  |  April 2026"
ws["A1"].font  = Font(name="Arial", color=GOLD, bold=True, size=14)
ws["A1"].fill  = fill(HEADER_BG)
ws["A1"].alignment = center()

ws.merge_cells("A2:R2")
ws["A2"] = (
    "Basis: VCP / 52-Wk Breakout / Post-Earnings Base / Analyst Target (Motilal Oswal · ICICI Sec · Emkay · Jefferies · Deven Choksey)  "
    "|  Rule: ALL entry criteria must pass before trading  |  Max risk 1% portfolio per trade"
)
ws["A2"].font  = Font(name="Arial", color=GREY, size=9, italic=True)
ws["A2"].fill  = fill(DARK_BG)
ws["A2"].alignment = center()

ws.row_dimensions[1].height = 28
ws.row_dimensions[2].height = 18

# ── Column headers ────────────────────────────────────────────────────────────
HEADERS = [
    "#", "NSE Symbol", "Stock Name", "Tier / Score",
    "Basis",
    "Fund House / Source", "Rec Date", "Rec Type",
    "Analyst\nTarget (₹)", "CMP (₹)\n(Approx)",
    "Upside\nto Target",
    "Pattern / Setup",
    "Key Catalyst",
    "Entry Zone (₹)", "Stop Loss (₹)\n(-0.5%)",
    "T1 (+1%) ₹", "T2 (+1.5%) ₹", "T3 (+2%) ₹"
]

COL_WIDTHS = [4, 16, 22, 13, 22, 24, 12, 12, 12, 12, 10, 22, 38, 16, 16, 13, 13, 13]

for i, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
    cell = ws.cell(row=3, column=i, value=h)
    cell.font      = Font(name="Arial", color=GOLD, bold=True, size=9)
    cell.fill      = fill(HEADER_BG)
    cell.alignment = center()
    cell.border    = border()
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[3].height = 32

# ── Data ───────────────────────────────────────────────────────────────────────
# Columns: #, Symbol, Name, Tier/Score, Basis, FundHouse, Date, RecType,
#          Target, CMP, Upside%, Pattern, Catalyst,
#          Entry, SL, T1, T2, T3

DATA = [
    # ── TIER 1 ────────────────────────────────────────────────────────────────
    (1,  "NSE:COALINDIA",  "Coal India Ltd",
     "Tier 1 | 72/100",
     "Strategy + News",
     "Options Mkt / NSE Data",         "24 Apr 2026", "Strong Buy",
     490,    458.9,  "=IF(I{r}>0,(I{r}-J{r})/J{r},\"—\")",
     "VCP Breakout (3-wk contraction → explosive candle +3.61%)",
     "Q4 results + Final Dividend — Board Meeting 27 Apr 2026. Options: 9,859 calls at ₹460 strike.",
     "462–464", "=ROUND(J{r}*0.995,1)", "=ROUND(N{r}*1.01,1)", "=ROUND(N{r}*1.015,1)", "=ROUND(N{r}*1.02,1)"),

    (2,  "NSE:NESTLEIND",  "Nestle India Ltd",
     "Tier 1 | 70/100",
     "Strategy + News",
     "Motilal Oswal",                   "22 Apr 2026", "Neutral→Target Hit",
     1400,   1431.15, "=IF(I{r}>0,(I{r}-J{r})/J{r},\"—\")",
     "52-Week High Breakout — hit ₹1,431.15 on 24 Apr. +19% in 1 month, +13% in 5 sessions",
     "Q4 FY26 Blowout: Profit +26% YoY (₹1,114 cr), Revenue +23% YoY (₹6,748 cr). Target ₹1,400 hit 2 days after call.",
     "1,431–1,435", "=ROUND(J{r}*0.995,1)", "=ROUND(N{r}*1.01,1)", "=ROUND(N{r}*1.015,1)", "=ROUND(N{r}*1.02,1)"),

    # ── TIER 2 ────────────────────────────────────────────────────────────────
    (3,  "NSE:ICICIBANK",  "ICICI Bank Ltd",
     "Tier 2 | 66/100",
     "Strategy + News",
     "Multiple Brokerages",             "18 Apr 2026", "Buy",
     1700,   1388,   "=IF(I{r}>0,(I{r}-J{r})/J{r},\"—\")",
     "Post-Earnings Accumulation Base — tight consolidation after strong Q4",
     "Q4 FY26: PAT ₹13,702 cr (+5.8% beat vs est ₹12,949 cr). Record-low GNPA 1.40%. Dividend ₹12/share.",
     "1,400–1,410", "=ROUND(J{r}*0.995,1)", "=ROUND(N{r}*1.01,1)", "=ROUND(N{r}*1.015,1)", "=ROUND(N{r}*1.02,1)"),

    (4,  "NSE:BEL",         "Bharat Electronics",
     "Tier 2 | 65/100",
     "Strategy + News",
     "Equitymaster / Sector",           "01 Apr 2026", "Buy",
     0,      290,    "—",
     "Post-Results Breakout (+6% on Apr 1) → Tight Consolidation → Second Leg",
     "FY26 Revenue ₹26,750 cr (+16.2%). Order book ₹74,000 cr. New orders ₹30,000 cr in FY26. Exports +33.65%.",
     "Break above consol. high", "=ROUND(J{r}*0.995,1)", "=ROUND(J{r}*1.01,1)", "=ROUND(J{r}*1.015,1)", "=ROUND(J{r}*1.02,1)"),

    (5,  "NSE:HDFCBANK",   "HDFC Bank Ltd",
     "Tier 2 | 64/100",
     "Strategy + News",
     "Jefferies / Motilal Oswal",       "Apr 2026",    "Buy",
     2050,   1785,   "=IF(I{r}>0,(I{r}-J{r})/J{r},\"—\")",
     "Post-Q4 Accumulation — institutional buying post strong results",
     "Q4 FY26: PAT ₹19,221 cr (beat). Loans +12%, Deposits +14.4%. Final Dividend ₹13/share (record date Jun 19).",
     "Break above recent high", "=ROUND(J{r}*0.995,1)", "=ROUND(J{r}*1.01,1)", "=ROUND(J{r}*1.015,1)", "=ROUND(J{r}*1.02,1)"),

    (6,  "NSE:TRENT",      "Trent Ltd (Tata)",
     "Tier 2 | 62/100",
     "News — Analyst Call",
     "Motilal Oswal",                   "22 Apr 2026", "Buy",
     5250,   4297,   "=IF(I{r}>0,(I{r}-J{r})/J{r},\"—\")",
     "Uptrend with consolidation — consumer discretionary leader",
     "Motilal Oswal Buy with target ₹5,250 (+22% upside). Consumer spending recovery. Zudio + Westside expansion.",
     "Break above ₹4,350", "=ROUND(J{r}*0.995,1)", "=ROUND(J{r}*1.01,1)", "=ROUND(J{r}*1.015,1)", "=ROUND(J{r}*1.02,1)"),

    # ── TIER 3 ────────────────────────────────────────────────────────────────
    (7,  "NSE:BAJFINANCE", "Bajaj Finance Ltd",
     "Tier 3 | 62/100",
     "Strategy + News",
     "NSE Data / Consensus",            "Apr 2026",    "Buy",
     0,      0,      "—",
     "Pre-Results VCP — volatility contraction ahead of Apr 29 results",
     "AUM crossed ₹5 lakh crore milestone first-ever. New loans booked +20.5% YoY in Q4. Results Apr 29.",
     "Break above consol. high", "—", "—", "—", "—"),

    (8,  "NSE:CIPLA",      "Cipla Ltd",
     "Tier 3 | 61/100",
     "News — Analyst Call",
     "ICICI Securities",                "24 Apr 2026", "Buy",
     1550,   1295,   "=IF(I{r}>0,(I{r}-J{r})/J{r},\"—\")",
     "Momentum consolidation — pharma sector recovery",
     "ICICI Securities Buy target ₹1,550 (+19.7% upside). Pharma sector tailwind. Dr Reddy's +0.35% leading sector.",
     "Break above ₹1,310", "=ROUND(J{r}*0.995,1)", "=ROUND(J{r}*1.01,1)", "=ROUND(J{r}*1.015,1)", "=ROUND(J{r}*1.02,1)"),

    (9,  "NSE:HDFCAMC",   "HDFC AMC Ltd",
     "Tier 3 | 61/100",
     "News — Analyst Call",
     "Motilal Oswal",                   "17 Apr 2026", "Buy",
     3170,   3400,   "=IF(I{r}>0,(I{r}-J{r})/J{r},\"—\")",
     "Post-Q4 Breakout (+5% on results day) — forming higher base",
     "Motilal Oswal Buy ₹3,170 target. Q4 results strong. Stock jumped 5% on Apr 17. Wealth management tailwind.",
     "Break above ₹3,500", "=ROUND(J{r}*0.995,1)", "=ROUND(J{r}*1.01,1)", "=ROUND(J{r}*1.015,1)", "=ROUND(J{r}*1.02,1)"),

    (10, "NSE:SBILIFE",   "SBI Life Insurance",
     "Tier 3 | 60/100",
     "News — Analyst Call",
     "Emkay + ICICI Securities",        "23 Apr 2026", "Buy",
     2345,   1769,   "=IF(I{r}>0,(I{r}-J{r})/J{r},\"—\")",
     "Consolidation — insurance sector recovering",
     "Emkay target ₹2,250, ICICI Securities target ₹2,345 (+32.6% upside). Insurance sector momentum.",
     "Break above ₹1,800", "=ROUND(J{r}*0.995,1)", "=ROUND(J{r}*1.01,1)", "=ROUND(J{r}*1.015,1)", "=ROUND(J{r}*1.02,1)"),

    (11, "NSE:BAJAJ-AUTO","Bajaj Auto Ltd",
     "Tier 3 | 61/100",
     "Strategy + News",
     "Company Data / Consensus",        "Apr 2026",    "Monitor",
     0,      9793,   "—",
     "Steady uptrend — low volatility consolidation pre-results",
     "March 2026 sales +20% YoY (4.45 lakh units). FY26 total 51.17 lakh units. EPS beat +2.9%. Next results May 6.",
     "Break above ₹9,800", "=ROUND(J{r}*0.995,1)", "=ROUND(J{r}*1.01,1)", "=ROUND(J{r}*1.015,1)", "=ROUND(J{r}*1.02,1)"),

    (12, "NSE:AXISBANK",  "Axis Bank Ltd",
     "Tier 3 | 60/100",
     "Strategy + News",
     "Axis Direct / Consensus",         "25 Apr 2026", "Buy",
     0,      1333,   "—",
     "Post-Q4 confirmation — banking sector momentum",
     "Q4 FY26 results out. PAT est ₹6,200–6,800 cr. NIM 3.9–4.0%. Banking sector strong (ICICI/HDFC both beat).",
     "Break above ₹1,350", "=ROUND(J{r}*0.995,1)", "=ROUND(J{r}*1.01,1)", "=ROUND(J{r}*1.015,1)", "=ROUND(J{r}*1.02,1)"),

    # ── TIER 4 ────────────────────────────────────────────────────────────────
    (13, "NSE:CHOLAFIN",  "Cholamandalam Inv.",
     "Tier 4 | 58/100",
     "News — Analyst Call",
     "Motilal Oswal",                   "16 Apr 2026", "Buy",
     0,      0,      "21% upside (MOFSL)",
     "Momentum recovery — NBFC sector",
     "Motilal Oswal retains Buy for 21% upside. Growth rebound thesis. Strong auto loan book.",
     "On VWAP bounce", "—", "—", "—", "—"),

    (14, "NSE:SBIN",      "State Bank of India",
     "Tier 4 | 58/100",
     "Strategy + News",
     "Multiple Brokerages",             "Apr 2026",    "Buy",
     1350,   1091,   "=IF(I{r}>0,(I{r}-J{r})/J{r},\"—\")",
     "PSU banking tailwind",
     "Target ₹1,350 from multiple brokerages. PSU banking sector beneficiary. Q4 results pending.",
     "VWAP bounce on green day", "=ROUND(J{r}*0.995,1)", "=ROUND(J{r}*1.01,1)", "=ROUND(J{r}*1.015,1)", "=ROUND(J{r}*1.02,1)"),

    (15, "NSE:JSWSTEEL",  "JSW Steel Ltd",
     "Tier 4 | 57/100",
     "Strategy",
     "Sector Momentum",                 "Apr 2026",    "Watch",
     0,      0,      "—",
     "Metals sector momentum",
     "+2.20% on market recovery day. Infrastructure spend acceleration. Govt capex push in steel.",
     "Break above recent swing high", "—", "—", "—", "—"),

    (16, "NSE:POWERGRID", "Power Grid Corp.",
     "Tier 4 | 56/100",
     "Strategy",
     "PSU Energy / Govt Capex",         "Apr 2026",    "Watch",
     0,      0,      "—",
     "PSU energy infrastructure",
     "+1.89% on recovery day. Renewable energy evacuation infra. Regulated stable returns + high dividend.",
     "VWAP bounce with volume", "—", "—", "—", "—"),

    (17, "NSE:NTPC",      "NTPC Ltd",
     "Tier 4 | 55/100",
     "Strategy",
     "PSU Energy / Govt Capex",         "Apr 2026",    "Watch",
     0,      0,      "—",
     "Renewable energy expansion",
     "India's largest power generator. Aggressive green energy portfolio expansion. Energy security theme.",
     "VWAP bounce with volume", "—", "—", "—", "—"),
]

# ── Tier color map ─────────────────────────────────────────────────────────────
TIER_COLORS = {
    "Tier 1": (TIER1_BG, GOLD),
    "Tier 2": (TIER2_BG, BLUE_LINK),
    "Tier 3": (TIER3_BG, WHITE),
    "Tier 4": (TIER4_BG, GREY),
}

for idx, row_data in enumerate(DATA):
    r = idx + 4  # Excel row (data starts at row 4)
    tier_key = row_data[3].split(" |")[0].strip()
    bg_hex, fg_hex = TIER_COLORS.get(tier_key, (DARK_BG, WHITE))

    for col_idx, val in enumerate(row_data, 1):
        # Substitute row number into formulas
        if isinstance(val, str) and "{r}" in val:
            val = val.replace("{r}", str(r))

        cell = ws.cell(row=r, column=col_idx, value=val)
        cell.fill      = fill(bg_hex)
        cell.font      = font(color=fg_hex, size=9)
        cell.alignment = center() if col_idx in (1, 4, 7, 8, 9, 10, 11, 14, 15, 16, 17, 18) else left()
        cell.border    = border()

        # Special formatting
        if col_idx == 1:   # Row number bold
            cell.font = font(color=GOLD, bold=True, size=9)
        if col_idx == 2:   # Symbol bold + colour
            cell.font = font(color=GREEN, bold=True, size=9)
        if col_idx == 4:   # Score col — colour by tier
            cell.font = font(color=GOLD if "Tier 1" in str(val) else
                                   BLUE_LINK if "Tier 2" in str(val) else
                                   WHITE if "Tier 3" in str(val) else GREY,
                             bold=True, size=9)
        if col_idx == 8:   # Rec type
            colour = (GREEN if "Buy" in str(val) or "Strong" in str(val)
                      else AMBER if "Monitor" in str(val) or "Watch" in str(val)
                      else GREY)
            cell.font = font(color=colour, bold=True, size=9)

        # Format upside % column
        if col_idx == 11 and isinstance(val, str) and val.startswith("="):
            cell.number_format = '0.0%;(0.0%);"-"'

        # Format price columns
        if col_idx in (9, 10, 14, 15, 16, 17, 18) and isinstance(val, (int, float)) and val > 0:
            cell.number_format = '#,##0.00'

    ws.row_dimensions[r].height = 42

# ── Legend / Notes section ─────────────────────────────────────────────────────
note_row = len(DATA) + 5
ws.merge_cells(f"A{note_row}:R{note_row}")
ws[f"A{note_row}"] = (
    "⚠️  TRADE ONLY WHEN ALL CRITERIA MET: Nifty NOT down >1% | India VIX <20 | Price >VWAP | Volume ≥1.5x avg | "
    "Order book bid imbalance ≥1.5 | MACD bullish | RSI 55-80 on daily | Price >9 EMA & 20 EMA | "
    "Max hold: 20 min | SL: 0.5% max | Daily loss limit: 2 consecutive stops = STOP for day"
)
ws[f"A{note_row}"].font      = Font(name="Arial", color=RED_LIGHT, bold=True, size=9)
ws[f"A{note_row}"].fill      = fill(HEADER_BG)
ws[f"A{note_row}"].alignment = center()
ws.row_dimensions[note_row].height = 22

ws.merge_cells(f"A{note_row+1}:R{note_row+1}")
ws[f"A{note_row+1}"] = (
    "Sources: Trendlyne Research Reports (Apr 24, 2026) · Business Standard · MoneyControl · Upstox News · MarketsMojo · NSE India  |  "
    "Compiled by Claude Sonnet 4.6 for mdrazeeqh@gmail.com  |  NOT financial advice"
)
ws[f"A{note_row+1}"].font      = Font(name="Arial", color=GREY, italic=True, size=8)
ws[f"A{note_row+1}"].fill      = fill(DARK_BG)
ws[f"A{note_row+1}"].alignment = center()
ws.row_dimensions[note_row+1].height = 16

# ── Freeze panes ──────────────────────────────────────────────────────────────
ws.freeze_panes = "A4"

# ── Save ──────────────────────────────────────────────────────────────────────
out = r"C:\Users\razee\OneDrive\Desktop\TradingClaude\NSE_Scalp_Watchlist_Apr2026.xlsx"
wb.save(out)
print(f"Saved: {out}")
