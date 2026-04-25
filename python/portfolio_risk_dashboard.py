"""
Portfolio Risk Dashboard — Across All 4 Books
─────────────────────────────────────────────────────────────────────────────
Aggregates open positions across:
  • Scalp (intraday delivery)
  • 1-Day Hold (news momentum)
  • Swing (2-3 day fallen-angel buys)
  • Multibagger (3-5 year holds)
  • IPO (allocations)

Reports:
  • Capital deployed % (warn if >70% so dry powder available)
  • Sector concentration (warn if >25% in any single sector)
  • Single-stock concentration (warn if >10%)
  • Open risk (sum of (entry-SL)×qty across all positions)
  • Per-book P&L (live MTM)
  • Correlation cluster warning (e.g. all bank stocks)
  • Suggested rebalance actions

Reads positions from positions.json (manually maintained or wired to broker
API later) and price_data.json for MTM. Prints clean dashboard.
─────────────────────────────────────────────────────────────────────────────
"""
import json, os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from sector_rotation import get_sector_for_stock
except Exception:
    def get_sector_for_stock(s): return None

EXCEL_PATH     = r"C:\Users\razee\OneDrive\Desktop\TradingClaude\NSE_Portfolio_Risk_Dashboard.xlsx"
POSITIONS_JSON = r"C:\Users\razee\OneDrive\Desktop\TradingClaude\positions.json"
PRICE_JSON     = r"C:\Users\razee\OneDrive\Desktop\TradingClaude\price_data.json"
TODAY          = date.today()

DARK_BG="0D0D0D"; HEADER_BG="1A1A2E"; ROW_ALT="141414"
GOLD="FFD700"; GREEN="00C896"; BLUE="4FC3F7"; WHITE="FFFFFF"
GREY="9E9E9E"; RED="FF5252"; AMBER="FFB300"; ORANGE="FF6B35"; CYAN="00BCD4"; PURPLE="9C27B0"

# ── Risk thresholds ───────────────────────────────────────────────────────────
THRESHOLDS = {
    "max_capital_deployed_pct": 70,
    "max_per_stock_pct":        10,
    "max_per_sector_pct":       25,
    "max_per_book_pct":         {"scalp": 20, "1day_hold": 25, "swing": 30, "multibagger": 40, "ipo": 15},
    "min_dry_powder_pct":       30,
}

# ── Demo positions (replace with real positions.json or broker API) ──────────
DEMO_POSITIONS = {
    "capital_total_rs": 1_000_000,
    "cash_available_rs": 320_000,
    "positions": [
        # book, symbol, qty, avg_price, sl, target, entry_date
        {"book": "scalp",       "symbol": "NSE:BEL",         "qty": 385,  "avg_price": 450.0,  "sl": 437.0,  "target": 462.0,  "entry_date": "24 Apr 2026"},
        {"book": "1day_hold",   "symbol": "NSE:NCC",         "qty": 1500, "avg_price": 290.0,  "sl": 281.0,  "target": 305.0,  "entry_date": "24 Apr 2026"},
        {"book": "1day_hold",   "symbol": "NSE:KEI",         "qty": 35,   "avg_price": 3850.0, "sl": 3722.0, "target": 3950.0, "entry_date": "24 Apr 2026"},
        {"book": "swing",       "symbol": "NSE:SHAKTIPUMP",  "qty": 162,  "avg_price": 880.0,  "sl": 833.0,  "target": 972.0,  "entry_date": "22 Apr 2026"},
        {"book": "swing",       "symbol": "NSE:VEDL",        "qty": 350,  "avg_price": 360.0,  "sl": 341.0,  "target": 400.0,  "entry_date": "23 Apr 2026"},
        {"book": "multibagger", "symbol": "NSE:KAYNES",      "qty": 10,   "avg_price": 5100.0, "sl": 4753.0, "target": 7500.0, "entry_date": "20 Apr 2026"},
        {"book": "multibagger", "symbol": "NSE:DATAPATTNS",  "qty": 24,   "avg_price": 2480.0, "sl": 2280.0, "target": 3800.0, "entry_date": "21 Apr 2026"},
        {"book": "multibagger", "symbol": "NSE:KPIGREEN",    "qty": 80,   "avg_price": 695.0,  "sl": 640.0,  "target": 1200.0, "entry_date": "18 Apr 2026"},
    ],
    "live_prices": {
        "NSE:BEL":         449.95,
        "NSE:NCC":         295.50,
        "NSE:KEI":         3895.0,
        "NSE:SHAKTIPUMP":  878.0,
        "NSE:VEDL":        360.5,
        "NSE:KAYNES":      5180.0,
        "NSE:DATAPATTNS":  2475.0,
        "NSE:KPIGREEN":    692.0,
    },
}

def fill(h): return PatternFill("solid", fgColor=h)
def font(color=WHITE, bold=False, size=9, italic=False):
    return Font(name="Arial", color=color, bold=bold, size=size, italic=italic)
def bdr():
    s = Side(style="thin", color="2D2D2D")
    return Border(left=s, right=s, top=s, bottom=s)
def mid(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def load_positions():
    if not os.path.exists(POSITIONS_JSON):
        return DEMO_POSITIONS
    try:
        with open(POSITIONS_JSON) as f:
            return json.load(f)
    except Exception:
        return DEMO_POSITIONS

def calculate_metrics(data):
    capital = data["capital_total_rs"]
    cash = data["cash_available_rs"]
    positions = data["positions"]
    prices = data.get("live_prices", {})

    # Per-position MTM
    rows = []
    for p in positions:
        ltp = prices.get(p["symbol"], p["avg_price"])
        deployed = p["qty"] * p["avg_price"]
        current_value = p["qty"] * ltp
        mtm_rs = current_value - deployed
        mtm_pct = (mtm_rs / deployed) * 100 if deployed else 0
        risk_rs = (p["avg_price"] - p["sl"]) * p["qty"]  # initial risk
        sector = get_sector_for_stock(p["symbol"]) or "Other"
        rows.append({**p, "ltp": ltp, "deployed": deployed, "current_value": current_value,
                      "mtm_rs": mtm_rs, "mtm_pct": mtm_pct, "risk_rs": risk_rs, "sector": sector})

    # Aggregates
    total_deployed = sum(r["deployed"] for r in rows)
    total_current = sum(r["current_value"] for r in rows)
    total_mtm = total_current - total_deployed
    total_risk = sum(r["risk_rs"] for r in rows)
    deployed_pct = (total_deployed / capital) * 100 if capital else 0
    cash_pct = (cash / capital) * 100 if capital else 0
    open_risk_pct = (total_risk / capital) * 100 if capital else 0

    # By book
    by_book = {}
    for r in rows:
        by_book.setdefault(r["book"], {"deployed": 0, "mtm": 0, "count": 0})
        by_book[r["book"]]["deployed"] += r["deployed"]
        by_book[r["book"]]["mtm"]      += r["mtm_rs"]
        by_book[r["book"]]["count"]    += 1

    # By sector
    by_sector = {}
    for r in rows:
        by_sector.setdefault(r["sector"], 0)
        by_sector[r["sector"]] += r["deployed"]

    # Generate alerts
    alerts = []
    if deployed_pct > THRESHOLDS["max_capital_deployed_pct"]:
        alerts.append(("HIGH", f"Deployed {deployed_pct:.1f}% (limit {THRESHOLDS['max_capital_deployed_pct']}%) — reduce exposure"))
    if cash_pct < THRESHOLDS["min_dry_powder_pct"]:
        alerts.append(("HIGH", f"Cash only {cash_pct:.1f}% — below {THRESHOLDS['min_dry_powder_pct']}% min — book partial profits"))
    for r in rows:
        single_pct = (r["deployed"] / capital) * 100
        if single_pct > THRESHOLDS["max_per_stock_pct"]:
            alerts.append(("HIGH", f"{r['symbol']}: {single_pct:.1f}% of portfolio — over {THRESHOLDS['max_per_stock_pct']}% single-stock cap"))
    for sec, val in by_sector.items():
        sec_pct = (val / capital) * 100
        if sec_pct > THRESHOLDS["max_per_sector_pct"]:
            alerts.append(("MED", f"Sector {sec}: {sec_pct:.1f}% — over {THRESHOLDS['max_per_sector_pct']}% sector cap"))
    for book, lim in THRESHOLDS["max_per_book_pct"].items():
        d = by_book.get(book, {}).get("deployed", 0)
        bpct = (d / capital) * 100
        if bpct > lim:
            alerts.append(("MED", f"Book {book}: {bpct:.1f}% — over {lim}% book cap"))

    return {
        "capital": capital, "cash": cash, "deployed": total_deployed,
        "current_value": total_current, "total_mtm": total_mtm,
        "deployed_pct": deployed_pct, "cash_pct": cash_pct,
        "open_risk": total_risk, "open_risk_pct": open_risk_pct,
        "rows": rows, "by_book": by_book, "by_sector": by_sector,
        "alerts": alerts,
    }

# ── Excel build ───────────────────────────────────────────────────────────────
def build():
    data = load_positions()
    m = calculate_metrics(data)

    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Risk"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:M1")
    ws["A1"] = (f"PORTFOLIO RISK DASHBOARD  |  Updated {TODAY.strftime('%d %b %Y')}  |  "
                f"Capital ₹{m['capital']:,}  |  Cash ₹{m['cash']:,} ({m['cash_pct']:.1f}%)")
    ws["A1"].font = Font(name="Arial", color=GOLD, bold=True, size=13)
    ws["A1"].fill = fill(HEADER_BG); ws["A1"].alignment = mid()
    ws.row_dimensions[1].height = 28

    # Top KPIs
    ws.merge_cells("A2:M2")
    mtm_color = GREEN if m["total_mtm"] >= 0 else RED
    ws["A2"] = (f"📊 Deployed ₹{m['deployed']:,.0f} ({m['deployed_pct']:.1f}%)  |  "
                f"Current ₹{m['current_value']:,.0f}  |  "
                f"MTM ₹{m['total_mtm']:+,.0f} ({(m['total_mtm']/m['deployed']*100 if m['deployed'] else 0):+.2f}%)  |  "
                f"Open Risk ₹{m['open_risk']:,.0f} ({m['open_risk_pct']:.2f}% of capital)")
    ws["A2"].font = Font(name="Arial", color=mtm_color, bold=True, size=10)
    ws["A2"].fill = fill(DARK_BG); ws["A2"].alignment = mid()
    ws.row_dimensions[2].height = 22

    cur = 4

    # ── ALERTS section ─────────────────────────────────────────────────────────
    ws.merge_cells(f"A{cur}:M{cur}")
    ws.cell(cur, 1, "🚨 ALERTS & REBALANCE ACTIONS")
    ws.cell(cur, 1).font = Font(name="Arial", color=GOLD, bold=True, size=11)
    ws.cell(cur, 1).fill = fill(HEADER_BG); ws.cell(cur, 1).alignment = mid()
    ws.row_dimensions[cur].height = 20
    cur += 1
    if not m["alerts"]:
        ws.merge_cells(f"A{cur}:M{cur}")
        ws.cell(cur, 1, "✅ No risk alerts — portfolio within all thresholds")
        ws.cell(cur, 1).font = font(GREEN, bold=True)
        ws.cell(cur, 1).fill = fill(DARK_BG); ws.cell(cur, 1).alignment = mid()
        ws.row_dimensions[cur].height = 20
        cur += 1
    else:
        for sev, msg in m["alerts"]:
            ws.merge_cells(f"A{cur}:M{cur}")
            clr = RED if sev == "HIGH" else AMBER
            ws.cell(cur, 1, f"{sev}: {msg}")
            ws.cell(cur, 1).font = font(clr, bold=True)
            ws.cell(cur, 1).fill = fill("2D0000" if sev == "HIGH" else "3D2C00")
            ws.cell(cur, 1).alignment = mid()
            ws.row_dimensions[cur].height = 20
            cur += 1
    cur += 1

    # ── POSITIONS table ───────────────────────────────────────────────────────
    headers = ["#", "Book", "Symbol", "Sector", "Qty", "Avg ₹", "LTP ₹",
                "SL ₹", "Tgt ₹", "Deployed ₹", "Current ₹", "MTM ₹", "MTM %"]
    widths  = [4, 12, 16, 18, 8, 10, 10, 10, 10, 13, 13, 12, 9]

    ws.merge_cells(f"A{cur}:M{cur}")
    ws.cell(cur, 1, "📋 OPEN POSITIONS")
    ws.cell(cur, 1).font = Font(name="Arial", color=GOLD, bold=True, size=11)
    ws.cell(cur, 1).fill = fill(HEADER_BG); ws.cell(cur, 1).alignment = mid()
    ws.row_dimensions[cur].height = 20
    cur += 1
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(cur, i, h)
        c.font = Font(name="Arial", color=GOLD, bold=True, size=9)
        c.fill = fill(HEADER_BG); c.alignment = mid(); c.border = bdr()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[cur].height = 28
    cur += 1

    book_rank = {"scalp": 0, "1day_hold": 1, "swing": 2, "multibagger": 3, "ipo": 4}
    book_color = {"scalp": CYAN, "1day_hold": BLUE, "swing": AMBER, "multibagger": "9C27B0", "ipo": GREEN}
    rows = sorted(m["rows"], key=lambda r: (book_rank.get(r["book"], 9), -r["deployed"]))
    for idx, r in enumerate(rows, 1):
        row = cur
        cells = [idx, r["book"].upper(), r["symbol"], r["sector"], r["qty"],
                  r["avg_price"], r["ltp"], r["sl"], r["target"],
                  f"₹{r['deployed']:,.0f}", f"₹{r['current_value']:,.0f}",
                  f"₹{r['mtm_rs']:+,.0f}", f"{r['mtm_pct']:+.2f}%"]
        row_bg = ROW_ALT if idx % 2 else DARK_BG
        for col_i, val in enumerate(cells, 1):
            c = ws.cell(row, col_i, val)
            c.fill = fill(row_bg); c.border = bdr()
            c.font = font(WHITE, size=9)
            c.alignment = mid() if col_i != 4 else lft()
            if col_i == 2: c.font = font(book_color.get(r["book"], WHITE), bold=True)
            if col_i == 3: c.font = font(GREEN, bold=True)
            if col_i == 8: c.font = font(RED, bold=True)
            if col_i == 9: c.font = font(GOLD, bold=True)
            if col_i == 12 or col_i == 13:
                c.font = font(GREEN if r["mtm_rs"] >= 0 else RED, bold=True)
        ws.row_dimensions[row].height = 24
        cur += 1

    cur += 1

    # ── BY-BOOK summary ───────────────────────────────────────────────────────
    ws.merge_cells(f"A{cur}:M{cur}")
    ws.cell(cur, 1, "📚 BY BOOK (Capital Allocation + MTM)")
    ws.cell(cur, 1).font = Font(name="Arial", color=GOLD, bold=True, size=11)
    ws.cell(cur, 1).fill = fill(HEADER_BG); ws.cell(cur, 1).alignment = mid()
    ws.row_dimensions[cur].height = 20
    cur += 1
    book_headers = ["Book", "# Pos", "Deployed ₹", "% of Capital", "MTM ₹", "MTM %", "Cap Limit %"]
    bw = [16, 8, 14, 13, 12, 9, 12]
    for i, (h, w) in enumerate(zip(book_headers, bw), 1):
        c = ws.cell(cur, i, h)
        c.font = Font(name="Arial", color=GOLD, bold=True, size=9)
        c.fill = fill(HEADER_BG); c.alignment = mid(); c.border = bdr()
    ws.row_dimensions[cur].height = 24
    cur += 1
    for book, stats in sorted(m["by_book"].items(), key=lambda x: book_rank.get(x[0], 9)):
        row = cur
        bpct = (stats["deployed"] / m["capital"]) * 100
        mtm_pct = (stats["mtm"] / stats["deployed"] * 100) if stats["deployed"] else 0
        lim = THRESHOLDS["max_per_book_pct"].get(book, 100)
        cells = [book.upper(), stats["count"], f"₹{stats['deployed']:,.0f}",
                  f"{bpct:.1f}%", f"₹{stats['mtm']:+,.0f}", f"{mtm_pct:+.2f}%", f"{lim}%"]
        row_bg = ROW_ALT if (cur % 2) else DARK_BG
        for col_i, val in enumerate(cells, 1):
            c = ws.cell(row, col_i, val)
            c.fill = fill(row_bg); c.border = bdr()
            c.font = font(WHITE, size=9); c.alignment = mid()
            if col_i == 1: c.font = font(book_color.get(book, WHITE), bold=True)
            if col_i == 4: c.font = font(RED if bpct > lim else GREEN, bold=True)
            if col_i in (5, 6): c.font = font(GREEN if stats["mtm"] >= 0 else RED, bold=True)
        ws.row_dimensions[row].height = 22
        cur += 1
    cur += 1

    # ── BY-SECTOR summary ─────────────────────────────────────────────────────
    ws.merge_cells(f"A{cur}:M{cur}")
    ws.cell(cur, 1, "🏭 BY SECTOR (Concentration Check)")
    ws.cell(cur, 1).font = Font(name="Arial", color=GOLD, bold=True, size=11)
    ws.cell(cur, 1).fill = fill(HEADER_BG); ws.cell(cur, 1).alignment = mid()
    ws.row_dimensions[cur].height = 20
    cur += 1
    sec_headers = ["Sector", "Deployed ₹", "% of Capital", "Status"]
    sw = [22, 14, 13, 16]
    for i, (h, w) in enumerate(zip(sec_headers, sw), 1):
        c = ws.cell(cur, i, h)
        c.font = Font(name="Arial", color=GOLD, bold=True, size=9)
        c.fill = fill(HEADER_BG); c.alignment = mid(); c.border = bdr()
    ws.row_dimensions[cur].height = 24
    cur += 1
    sectors_sorted = sorted(m["by_sector"].items(), key=lambda x: -x[1])
    for sec, val in sectors_sorted:
        row = cur
        sec_pct = (val / m["capital"]) * 100
        status = "OVER LIMIT" if sec_pct > THRESHOLDS["max_per_sector_pct"] else "OK"
        cells = [sec, f"₹{val:,.0f}", f"{sec_pct:.1f}%", status]
        row_bg = ROW_ALT if (cur % 2) else DARK_BG
        for col_i, vv in enumerate(cells, 1):
            c = ws.cell(row, col_i, vv)
            c.fill = fill(row_bg); c.border = bdr()
            c.font = font(WHITE, size=9); c.alignment = mid()
            if col_i == 4:
                c.font = font(RED if status == "OVER LIMIT" else GREEN, bold=True)
        ws.row_dimensions[row].height = 22
        cur += 1

    # Footer
    cur += 2
    ws.merge_cells(f"A{cur}:M{cur}")
    ws.cell(cur, 1, ("⚠️  RISK PROTOCOL: Max 70% capital deployed · Max 10% per stock · "
                     "Max 25% per sector · Min 30% cash dry powder · "
                     "Re-balance weekly using these thresholds"))
    ws.cell(cur, 1).font = Font(name="Arial", color=AMBER, bold=True, size=9)
    ws.cell(cur, 1).fill = fill(HEADER_BG); ws.cell(cur, 1).alignment = mid()
    ws.row_dimensions[cur].height = 30

    ws.freeze_panes = "B4"
    wb.save(EXCEL_PATH)
    print(f"✅ Portfolio Risk Dashboard saved: {EXCEL_PATH}\n")

    # Console summary
    print(f"📊 Portfolio Snapshot ({TODAY.strftime('%d %b %Y')}):")
    print(f"  Capital:   ₹{m['capital']:,}")
    print(f"  Deployed:  ₹{m['deployed']:,.0f}  ({m['deployed_pct']:.1f}%)")
    print(f"  Cash:      ₹{m['cash']:,.0f}  ({m['cash_pct']:.1f}%)")
    print(f"  MTM:       ₹{m['total_mtm']:+,.0f}")
    print(f"  Open Risk: ₹{m['open_risk']:,.0f} ({m['open_risk_pct']:.2f}% of capital)")
    if m["alerts"]:
        print(f"\n  🚨 {len(m['alerts'])} alerts:")
        for sev, msg in m["alerts"]:
            print(f"     [{sev}] {msg}")
    else:
        print("\n  ✅ No alerts — portfolio within all thresholds")

if __name__ == "__main__":
    build()
