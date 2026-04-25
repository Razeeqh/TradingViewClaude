"""
NSE Fallen Angels Swing Screener
─────────────────────────────────────────────────────────────────────────────
Identifies fundamentally STRONG NSE stocks beaten down 25-50%+ from peak due to
TEMPORARY reasons (tariffs, war, sector rotation, one-time events) where the
fundamentals are intact and a catalyst-driven mean-reversion is likely.

EXCLUDES:
  • IT services (Infy, TCS, Wipro, TechM, HCL, LTIM, Mphasis) — AI disruption
  • Governance/fraud cases (Yes Bank legacy, Vodafone Idea structural debt)
  • Permanent decliners (lost market share in dying industries)

OUTPUT: NSE_Swing_FallenAngels.xlsx — dark-theme Excel with:
  • Drawdown %, drop reason, TEMP/PERM tag, breakout catalyst, catalyst date
  • Swing entry zone, SL (5-7%), T1 (+10%), T2 (+20%), T3 (+30-50%)
  • Status column auto-updated daily by the Claude scheduled session
─────────────────────────────────────────────────────────────────────────────
"""
import json, os
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Optional cross-module imports ────────────────────────────────────────────
try:
    from permanent_damage_blacklist import get_blacklist_set
    BLACKLIST = get_blacklist_set(["PERMANENT_AVOID", "WAIT_FOR_RESOLUTION"])
except Exception:
    BLACKLIST = set()

EXCEL_PATH    = r"C:\Users\razee\OneDrive\Desktop\TradingClaude\NSE_Swing_FallenAngels.xlsx"
FRESH_JSON    = r"C:\Users\razee\OneDrive\Desktop\TradingClaude\fallen_angels_fresh.json"
MACRO_JSON    = r"C:\Users\razee\OneDrive\Desktop\TradingClaude\macro_context.json"
TODAY         = date.today()

# ── Palette ───────────────────────────────────────────────────────────────────
DARK_BG="0D0D0D"; HEADER_BG="1A1A2E"
HIGH_BG="1B4332"; MED_BG="003566"; LOW_BG="2D2D2D"
GOLD="FFD700"; GREEN="00C896"; BLUE="4FC3F7"; WHITE="FFFFFF"
GREY="9E9E9E"; RED="FF5252";  AMBER="FFB300"; ORANGE="FF6B35"; CYAN="00BCD4"

STATUS_META = {
    "🚀 BREAKOUT IMMINENT":  (HIGH_BG, GREEN, "Catalyst within 7 days + price near entry zone"),
    "🟢 ACCUMULATE":          (HIGH_BG, GREEN, "At entry zone — start staggered buying"),
    "🔵 WATCH":               (MED_BG,  BLUE,  "Above entry — wait for retest or catalyst"),
    "🟡 BUILDING":            ("3D2C00", AMBER, "Below entry — let setup form, monitor weekly"),
    "⏳ EARNINGS THIS WEEK":  ("3D2C00", CYAN,  "Defer entry — binary event in next 7 days"),
    "🏗 CORP ACTION":         ("2D2D2D", PURPLE := "9C27B0", "Demerger / split / bonus pending — recalibrate post-event"),
    "🔴 INVALIDATED":        ("2D0000", RED,   "Stop-loss hit OR thesis broken (governance issue surfaced)"),
}

# ── Corporate-action handling ────────────────────────────────────────────────
# When a stock is mid-corp-action, freeze its status and surface action note.
CORPORATE_ACTIONS = {
    "NSE:TATAMOTORS": ("Demerged Oct 2025 → TMPV (passenger) + TMCV (commercial). Switch tracking to NSE:TATAMOTORS-PV / NSE:TATAMOTORS-CV.", "RESTRUCTURE"),
    "NSE:VEDL":       ("Demerger record date 01 May 2026 → 1:1 split into 4 entities. Hold through; recalibrate ATH/CMP/SL/T1-3 post-event.", "PENDING"),
}

# ── Macro risk overlay ────────────────────────────────────────────────────────
# macro_context.json schema (written by Saturday Opus 4.7 task):
#   { "vix": 18.5, "fii_flow_week_cr": -3320, "nifty_week_pct": -1.14,
#     "brent_usd": 102, "geopolitical_risk": "HIGH", "rbi_stance": "HOLD" }
def load_macro_context():
    if not os.path.exists(MACRO_JSON):
        return {}
    try:
        with open(MACRO_JSON) as f:
            return json.load(f)
    except Exception:
        return {}

def macro_risk_multiplier(macro):
    """Returns (multiplier 0.0-1.0, label, reason)."""
    if not macro:
        return 1.0, "NORMAL", "No macro override"
    vix = macro.get("vix", 15)
    fii = macro.get("fii_flow_week_cr", 0)
    nifty_week = macro.get("nifty_week_pct", 0)
    geopol = macro.get("geopolitical_risk", "LOW")
    if vix > 25 or geopol == "EXTREME":
        return 0.0, "RISK-OFF", f"VIX {vix} or geopolitical EXTREME — HOLD CASH"
    if vix > 22 or fii < -10000 or nifty_week < -3:
        return 0.5, "DEFENSIVE", f"VIX {vix} / FII -₹{abs(fii)}cr / Nifty {nifty_week}% — HALF SIZE"
    if vix > 20 or fii < -5000:
        return 0.75, "CAUTIOUS", f"VIX {vix} / FII -₹{abs(fii)}cr — REDUCE BY 25%"
    return 1.0, "NORMAL", "Macro green — full size"

def fill(h): return PatternFill("solid", fgColor=h)
def font(color=WHITE, bold=False, size=9, italic=False):
    return Font(name="Arial", color=color, bold=bold, size=size, italic=italic)
def bdr():
    s = Side(style="thin", color="2D2D2D")
    return Border(left=s, right=s, top=s, bottom=s)
def mid(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── Fallen Angels — TEMPORARY drops only (verify all prices before trading) ──
FALLEN_ANGELS = {
    "NSE:SHAKTIPUMP": {
        "name": "Shakti Pumps India", "sector": "Solar Pumps / Agri",
        "ath": 1455, "cmp": 880, "drawdown_pct": 39.5,
        "drop_reason": "PM-KUSUM order lumpiness + profit booking after 8x rally",
        "temp_justification": "Solar irrigation scheme intact; FY27 order book healthy; capacity expansion on track",
        "catalyst": "Q1 FY27 results + new PM-KUSUM tender wins",
        "catalyst_date": "Jul 2026",
        "entry_zone": "870–900", "sl": 825, "t1": 970, "t2": 1060, "t3": 1250,
        "thesis": "Pure-play solar pump leader (50%+ market share). Multi-year govt tailwind on solar irrigation. Recent dip is technical profit-booking, not fundamental.",
        "confidence": "HIGH",
    },
    "NSE:TATAMOTORS": {
        "name": "Tata Motors", "sector": "Auto / JLR",
        "ath": 1180, "cmp": 680, "drawdown_pct": 42.4,
        "drop_reason": "JLR China weakness + Trump tariff fears on UK exports",
        "temp_justification": "India CV/PV cycle bottoming; JLR EV pipeline; demerger value unlock pending",
        "catalyst": "Tariff clarity + JLR China recovery + CV-PV demerger",
        "catalyst_date": "Q2 CY26",
        "entry_zone": "670–695", "sl": 630, "t1": 750, "t2": 820, "t3": 950,
        "thesis": "Cyclical headwind not structural. CV+PV demerger unlocks SOTP value. JLR EV transition + India recovery = dual driver.",
        "confidence": "HIGH",
    },
    "NSE:VEDL": {
        "name": "Vedanta Ltd", "sector": "Diversified Metals",
        "ath": 525, "cmp": 360, "drawdown_pct": 31.4,
        "drop_reason": "Commodity softness + demerger NCLT delay + parent debt overhang",
        "temp_justification": "Aluminium/zinc cycle turning up; demerger close to NCLT approval; high dividend cushion",
        "catalyst": "Demerger NCLT approval + LME aluminium > $2700",
        "catalyst_date": "Q1 FY27",
        "entry_zone": "355–375", "sl": 335, "t1": 400, "t2": 440, "t3": 500,
        "thesis": "SOTP demerger unlocks value (5 listed entities). Commodity cycle tailwind + 8%+ dividend yield support.",
        "confidence": "MEDIUM",
    },
    "NSE:ADANIGREEN": {
        "name": "Adani Green Energy", "sector": "Renewables",
        "ath": 2175, "cmp": 920, "drawdown_pct": 57.7,
        "drop_reason": "Hindenburg / SEC overhang + capex / leverage worry",
        "temp_justification": "50GW by 2030 plan on track; long-term PPAs lock-in cashflow; debt restructuring done",
        "catalyst": "SEC settlement + new capacity addition prints",
        "catalyst_date": "FY27",
        "entry_zone": "900–950", "sl": 855, "t1": 1020, "t2": 1120, "t3": 1300,
        "thesis": "India's RE leader. Governance overhang excessive — operational metrics strong. Cashflow ramp visible from FY27.",
        "confidence": "MEDIUM",
    },
    "NSE:ASIANPAINT": {
        "name": "Asian Paints", "sector": "Paints / FMCG",
        "ath": 3590, "cmp": 2280, "drawdown_pct": 36.5,
        "drop_reason": "Birla Opus competitive entry + margin reset + rural softness",
        "temp_justification": "53% market share + distribution moat (1.5L+ dealers); rural recovery imminent; premiumisation intact",
        "catalyst": "Q1 FY27 volume recovery + crude oil stability",
        "catalyst_date": "Jul 2026",
        "entry_zone": "2260–2320", "sl": 2140, "t1": 2510, "t2": 2740, "t3": 3000,
        "thesis": "Incumbent moat intact. Birla competition priced in. Margin reset bottoming. Best risk-reward in FMCG.",
        "confidence": "HIGH",
    },
    "NSE:PAYTM": {
        "name": "One97 Communications", "sector": "Fintech",
        "ath": 1060, "cmp": 690, "drawdown_pct": 34.9,
        "drop_reason": "PPB curbs aftermath + profitability uncertainty",
        "temp_justification": "Payment Aggregator licence live; loan distribution scaling; first PAT print achieved",
        "catalyst": "Maiden PAT + UPI incentive policy clarity",
        "catalyst_date": "Q1–Q2 FY27",
        "entry_zone": "680–710", "sl": 640, "t1": 760, "t2": 830, "t3": 950,
        "thesis": "Regulatory cleanup done. 4 cr+ merchant base intact. Path to sustained profitability now visible.",
        "confidence": "MEDIUM",
    },
    "NSE:HINDALCO": {
        "name": "Hindalco Industries", "sector": "Aluminium",
        "ath": 775, "cmp": 545, "drawdown_pct": 29.7,
        "drop_reason": "Novelis Bay Minette cost overrun + LME aluminium softness",
        "temp_justification": "India business margins strong; Novelis issue is one-time; aluminium structurally tight by 2027",
        "catalyst": "Novelis margin recovery + LME > $2700",
        "catalyst_date": "H2 CY26",
        "entry_zone": "535–560", "sl": 505, "t1": 600, "t2": 660, "t3": 740,
        "thesis": "India franchise undervalued vs peers. Novelis transient cost shock priced in. Aluminium supply deficit looming.",
        "confidence": "HIGH",
    },
    "NSE:TATASTEEL": {
        "name": "Tata Steel", "sector": "Steel",
        "ath": 185, "cmp": 130, "drawdown_pct": 29.7,
        "drop_reason": "China dumping + UK transition costs",
        "temp_justification": "India volumes growing; UK EAF transition fully funded; safeguard duty filed",
        "catalyst": "Indian safeguard duty + China stimulus impact",
        "catalyst_date": "CY26",
        "entry_zone": "128–135", "sl": 121, "t1": 145, "t2": 158, "t3": 175,
        "thesis": "India capacity ramp absorbing UK losses. Safeguard duty + China stimulus = double tailwind.",
        "confidence": "MEDIUM",
    },
    "NSE:NMDC": {
        "name": "NMDC Ltd", "sector": "Iron Ore PSU",
        "ath": 286, "cmp": 195, "drawdown_pct": 31.8,
        "drop_reason": "Iron ore price cycle + China steel weakness",
        "temp_justification": "Lowest-cost producer globally; volume growth intact; ~5% dividend yield cushion",
        "catalyst": "China stimulus + iron ore > $110",
        "catalyst_date": "H2 CY26",
        "entry_zone": "190–200", "sl": 180, "t1": 215, "t2": 235, "t3": 270,
        "thesis": "Cash-rich PSU at cyclical bottom. Dividend yield supports downside.",
        "confidence": "MEDIUM",
    },
    "NSE:HEROMOTOCO": {
        "name": "Hero MotoCorp", "sector": "2-Wheeler Auto",
        "ath": 6240, "cmp": 4100, "drawdown_pct": 34.3,
        "drop_reason": "EV transition fears + rural demand softness",
        "temp_justification": "Vida EV ramping; rural recovery imminent; 50M+ user base intact; cash-rich BS",
        "catalyst": "Monsoon + Vida 2.0 launch + festive season",
        "catalyst_date": "Aug–Oct 2026",
        "entry_zone": "4050–4180", "sl": 3820, "t1": 4500, "t2": 4900, "t3": 5500,
        "thesis": "Rural cycle turn + EV transition fears overdone. Cash-rich BS + 4%+ dividend yield.",
        "confidence": "MEDIUM",
    },
    "NSE:NESTLEIND": {
        "name": "Nestle India", "sector": "FMCG",
        "ath": 2780, "cmp": 1980, "drawdown_pct": 28.8,
        "drop_reason": "Rural demand + premiumisation slowdown",
        "temp_justification": "Maggi/coffee dominance unshaken; rural turn imminent post good monsoon",
        "catalyst": "Monsoon + rural FMCG revival",
        "catalyst_date": "Q2 FY27",
        "entry_zone": "1960–2020", "sl": 1850, "t1": 2180, "t2": 2380, "t3": 2650,
        "thesis": "Defensive cashflow compounder at reset valuations. ROCE 100%+ intact.",
        "confidence": "HIGH",
    },
    "NSE:BAJAJFINSV": {
        "name": "Bajaj Finserv", "sector": "NBFC + Insurance",
        "ath": 2080, "cmp": 1500, "drawdown_pct": 27.9,
        "drop_reason": "Unsecured loan slowdown + RBI tightening",
        "temp_justification": "AUM growth intact 25%+; insurance arms scaling; rate cut tailwind starting",
        "catalyst": "RBI rate cut + Q1 FY27 results",
        "catalyst_date": "Jun–Jul 2026",
        "entry_zone": "1480–1530", "sl": 1395, "t1": 1650, "t2": 1800, "t3": 2000,
        "thesis": "Best-in-class NBFC. Cycle headwind transient. Insurance subsidiaries underappreciated.",
        "confidence": "HIGH",
    },
    "NSE:INDIGO": {
        "name": "InterGlobe Aviation", "sector": "Airlines",
        "ath": 5030, "cmp": 3550, "drawdown_pct": 29.4,
        "drop_reason": "Crude oil spike + P&W engine groundings",
        "temp_justification": "60%+ market share; A321XLR fleet expanding; crude mean-reverts; international ramp",
        "catalyst": "Crude < $75 + Iran-Israel de-escalation",
        "catalyst_date": "CY26",
        "entry_zone": "3500–3620", "sl": 3300, "t1": 3900, "t2": 4250, "t3": 4800,
        "thesis": "Dominant LCC with pricing power. Transient input shock. International expansion = next leg.",
        "confidence": "MEDIUM",
    },
    "NSE:UPL": {
        "name": "UPL Ltd", "sector": "Agrochemicals",
        "ath": 815, "cmp": 540, "drawdown_pct": 33.7,
        "drop_reason": "Global agchem destocking + price erosion",
        "temp_justification": "Destocking ending; LATAM recovery; cost optimisation done; deleveraging visible",
        "catalyst": "FY27 guidance + product price recovery",
        "catalyst_date": "May–Jul 2026",
        "entry_zone": "530–555", "sl": 500, "t1": 595, "t2": 650, "t3": 740,
        "thesis": "Global #5 agrochem player. Cycle bottom + balance sheet deleveraging.",
        "confidence": "MEDIUM",
    },
    "NSE:RECLTD": {
        "name": "REC Ltd", "sector": "Power Finance PSU",
        "ath": 654, "cmp": 410, "drawdown_pct": 37.3,
        "drop_reason": "PSU profit booking + rate sensitivity fear",
        "temp_justification": "Loan book 20% CAGR; renewables financing leader; rate cut beneficiary",
        "catalyst": "RBI rate cut + interim dividend",
        "catalyst_date": "Jun 2026",
        "entry_zone": "405–425", "sl": 380, "t1": 455, "t2": 495, "t3": 575,
        "thesis": "Power capex super-cycle financier. 4%+ dividend yield. Cleanest PSU balance sheet.",
        "confidence": "HIGH",
    },
    "NSE:PFC": {
        "name": "Power Finance Corp", "sector": "Power Finance PSU",
        "ath": 580, "cmp": 380, "drawdown_pct": 34.5,
        "drop_reason": "PSU correction + rate worry (twin to REC)",
        "temp_justification": "Sanctions pipeline strong; NPA controlled; renewables tailwind",
        "catalyst": "RBI rate cut + Q1 FY27 results",
        "catalyst_date": "Jun–Jul 2026",
        "entry_zone": "375–395", "sl": 355, "t1": 420, "t2": 460, "t3": 530,
        "thesis": "Twin engine to REC. Cheapest power-capex play in market.",
        "confidence": "HIGH",
    },
    "NSE:IRCTC": {
        "name": "IRCTC", "sector": "Railways / Travel PSU",
        "ath": 1280, "cmp": 740, "drawdown_pct": 42.2,
        "drop_reason": "Convenience fee fears + broader PSU correction",
        "temp_justification": "Monopoly e-ticketing; tourism + catering scaling; debt-free BS",
        "catalyst": "Vande Bharat catering ramp + tourism revival",
        "catalyst_date": "FY27",
        "entry_zone": "725–760", "sl": 685, "t1": 815, "t2": 890, "t3": 1020,
        "thesis": "Asset-light monopoly. Growth optionality (catering, tourism, packaging) intact.",
        "confidence": "MEDIUM",
    },
    "NSE:GODREJCP": {
        "name": "Godrej Consumer", "sector": "FMCG",
        "ath": 1545, "cmp": 1080, "drawdown_pct": 30.1,
        "drop_reason": "Indonesia weakness + Household Insecticide category reset",
        "temp_justification": "India HPC strong; new HI formulation in pipeline; rural turn imminent",
        "catalyst": "Monsoon + Q2 FY27 volume recovery",
        "catalyst_date": "Aug 2026",
        "entry_zone": "1065–1110", "sl": 1010, "t1": 1190, "t2": 1295, "t3": 1450,
        "thesis": "FMCG cycle bottom + new mgmt execution. Indonesia recovery = optionality.",
        "confidence": "MEDIUM",
    },
    "NSE:DABUR": {
        "name": "Dabur India", "sector": "FMCG",
        "ath": 672, "cmp": 470, "drawdown_pct": 30.1,
        "drop_reason": "Rural softness + beverage segment reset",
        "temp_justification": "Ayurveda moat intact; rural recovery; healthcare push scaling",
        "catalyst": "Monsoon + rural FMCG revival",
        "catalyst_date": "Q2 FY27",
        "entry_zone": "465–485", "sl": 440, "t1": 520, "t2": 565, "t3": 630,
        "thesis": "Rural-heavy FMCG at cycle bottom. Defensive compounder.",
        "confidence": "MEDIUM",
    },
    "NSE:RELIANCE": {
        "name": "Reliance Industries", "sector": "Conglomerate",
        "ath": 1610, "cmp": 1200, "drawdown_pct": 25.5,
        "drop_reason": "Retail margin reset + Jio tariff hike delay",
        "temp_justification": "Jio IPO trigger pending; O2C stable; retail SSSG turning",
        "catalyst": "Jio IPO announcement + Jio tariff hike",
        "catalyst_date": "FY27",
        "entry_zone": "1190–1230", "sl": 1130, "t1": 1330, "t2": 1450, "t3": 1600,
        "thesis": "Multiple value-unlock triggers (Jio IPO, retail demerger). Conglomerate discount excessive.",
        "confidence": "HIGH",
    },
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def days_until(catalyst_date_str):
    """Best-effort parse of catalyst date string → days from today (or None)."""
    if not catalyst_date_str:
        return None
    parsers = ["%d %b %Y", "%b %Y", "%d-%b-%Y", "%Y-%m-%d", "%B %Y"]
    for fmt in parsers:
        try:
            d = datetime.strptime(catalyst_date_str.strip(), fmt).date()
            return (d - TODAY).days
        except ValueError:
            continue
    return None

def calc_status(symbol, entry_zone, sl, current_price, catalyst_date_str, earnings_dates=None):
    """Return one of STATUS_META keys based on price, catalyst proximity,
       earnings-week proximity, and corporate-action state.
    earnings_dates: list of "DD Mmm YYYY" date strings of upcoming earnings."""
    # Corporate-action freeze
    if symbol in CORPORATE_ACTIONS:
        return "🏗 CORP ACTION"
    if not isinstance(current_price, (int, float)) or current_price <= 0:
        return "🟡 BUILDING"
    try:
        lo, hi = [float(x.strip()) for x in entry_zone.replace("–", "-").split("-")]
    except Exception:
        return "🟡 BUILDING"
    if current_price <= sl:
        return "🔴 INVALIDATED"

    # Earnings-week defer: if any earnings within 7 days, defer entry
    if earnings_dates:
        for edate_str in earnings_dates:
            d = days_until(edate_str)
            if d is not None and -3 <= d <= 7:
                return "⏳ EARNINGS THIS WEEK"

    days = days_until(catalyst_date_str) or 9999
    if lo <= current_price <= hi and 0 < days <= 7:
        return "🚀 BREAKOUT IMMINENT"
    if lo <= current_price <= hi:
        return "🟢 ACCUMULATE"
    if current_price > hi:
        return "🔵 WATCH"
    return "🟡 BUILDING"

# ── Excel build ───────────────────────────────────────────────────────────────

HEADERS = [
    "#", "NSE Symbol", "Stock Name", "Sector",
    "ATH ₹", "CMP ₹", "Drawdown %",
    "Drop Reason (TEMPORARY)", "Why Mean-Reverts",
    "Breakout Catalyst", "Catalyst Window",
    "Entry Zone ₹", "SL ₹\n(-5–7%)",
    "T1 ₹\n(+10%)", "T2 ₹\n(+20%)", "T3 ₹\n(+30–50%)",
    "R:R\n(T2)", "Confidence", "Status", "Investment Thesis",
]
COL_WIDTHS = [4,16,22,18,9,9,11,32,30,28,14,13,12,10,10,12,8,11,22,42]

def merge_fresh(meta):
    """Merge fresh CMP / catalyst-date data from JSON if present."""
    if not os.path.exists(FRESH_JSON):
        return meta
    try:
        with open(FRESH_JSON) as f:
            fresh = json.load(f)
        for sym, updates in fresh.items():
            if sym in meta:
                meta[sym].update(updates)
        print(f"Merged fresh data for {len(fresh)} symbols from fallen_angels_fresh.json")
    except Exception as e:
        print(f"Could not merge fresh JSON: {e}")
    return meta

def build():
    meta = merge_fresh({k: dict(v) for k, v in FALLEN_ANGELS.items()})

    # ── Blacklist exclusion ──
    excluded = []
    for sym in list(meta.keys()):
        if sym in BLACKLIST:
            excluded.append(sym)
            del meta[sym]

    # ── Macro overlay ──
    macro = load_macro_context()
    multiplier, regime, reason = macro_risk_multiplier(macro)

    wb = Workbook()
    ws = wb.active
    ws.title = "NSE Fallen Angels — Swing"
    ws.sheet_view.showGridLines = False

    # ── Title ──
    ws.merge_cells("A1:T1")
    ws["A1"] = (f"NSE FALLEN ANGELS — SWING TRADE WATCHLIST  |  "
                f"Updated {TODAY.strftime('%d %b %Y')}  |  "
                f"Fundamentally STRONG · TEMPORARILY BEATEN DOWN")
    ws["A1"].font = Font(name="Arial", color=GOLD, bold=True, size=13)
    ws["A1"].fill = fill(HEADER_BG)
    ws["A1"].alignment = mid()
    ws.row_dimensions[1].height = 28

    # ── Macro regime banner ──
    regime_color = GREEN if regime == "NORMAL" else AMBER if regime == "CAUTIOUS" else ORANGE if regime == "DEFENSIVE" else RED
    ws.merge_cells("A2:T2")
    ws["A2"] = f"📡 MACRO REGIME: {regime}  |  Position-size multiplier: {multiplier:.0%}  |  {reason}"
    ws["A2"].font = Font(name="Arial", color=regime_color, bold=True, size=10)
    ws["A2"].fill = fill(HEADER_BG)
    ws["A2"].alignment = mid()
    ws.row_dimensions[2].height = 18

    # ── Sub-header ──
    # (Original sub-header replaced by macro regime banner above.)

    # ── Status legend ──
    ws.merge_cells("A3:T3")
    ws["A3"] = ("STATUS  →  " + "   ".join(f"{k}: {v[2]}" for k, v in STATUS_META.items()))
    ws["A3"].font = Font(name="Arial", color=AMBER, bold=True, size=8)
    ws["A3"].fill = fill(HEADER_BG)
    ws["A3"].alignment = mid()
    ws.row_dimensions[3].height = 16

    # ── Headers ──
    for i, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font      = Font(name="Arial", color=GOLD, bold=True, size=9)
        c.fill      = fill(HEADER_BG)
        c.alignment = mid()
        c.border    = bdr()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 34

    # ── Sort by drawdown desc (best discounts first), then confidence ──
    conf_rank = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    sorted_syms = sorted(
        meta.keys(),
        key=lambda s: (conf_rank.get(meta[s]["confidence"], 3), -meta[s]["drawdown_pct"])
    )

    # ── Data rows ──
    for idx, sym in enumerate(sorted_syms, 1):
        d = meta[sym]
        r = idx + 4

        status   = calc_status(sym, d["entry_zone"], d["sl"], d["cmp"],
                                d.get("catalyst_date"), d.get("earnings_dates"))
        s_bg, s_fg, _ = STATUS_META[status]
        conf_bg = HIGH_BG if d["confidence"] == "HIGH" else MED_BG if d["confidence"] == "MEDIUM" else LOW_BG
        row_bg  = "141414" if idx % 2 else "0D0D0D"
        if status == "🔴 INVALIDATED": row_bg = "2D0000"
        if status == "🏗 CORP ACTION": row_bg = "2D2D2D"
        if status == "⏳ EARNINGS THIS WEEK": row_bg = "3D2C00"

        # Risk-Reward to T2
        try:
            lo, hi = [float(x.strip()) for x in d["entry_zone"].replace("–", "-").split("-")]
            entry_mid = (lo + hi) / 2
            risk = entry_mid - d["sl"]
            reward = d["t2"] - entry_mid
            rr = f"{reward/risk:.1f}:1" if risk > 0 else "—"
        except Exception:
            rr = "—"

        cells = [
            idx, sym, d["name"], d["sector"],
            d["ath"], d["cmp"], f"-{d['drawdown_pct']:.1f}%",
            d["drop_reason"], d["temp_justification"],
            d["catalyst"], d.get("catalyst_date", "—"),
            d["entry_zone"], d["sl"],
            d["t1"], d["t2"], d["t3"],
            rr, d["confidence"], status, d["thesis"],
        ]

        for col_i, val in enumerate(cells, 1):
            c = ws.cell(row=r, column=col_i, value=val)
            c.fill   = fill(row_bg)
            c.border = bdr()
            c.font   = font(WHITE, size=9)
            c.alignment = mid() if col_i in (1,2,4,5,6,7,11,12,13,14,15,16,17,18,19) else lft()

            # Targeted color coding
            if col_i == 1:  c.font = font(GOLD, bold=True)
            if col_i == 2:  c.font = font(GREEN, bold=True)
            if col_i == 7:  # Drawdown %
                pct = d["drawdown_pct"]
                clr = RED if pct > 40 else ORANGE if pct > 30 else AMBER
                c.font = font(clr, bold=True)
            if col_i == 13: c.font = font(RED, bold=True)         # SL
            if col_i == 14: c.font = font(GREEN)                  # T1
            if col_i == 15: c.font = font(GREEN, bold=True)       # T2
            if col_i == 16: c.font = font(GOLD, bold=True)        # T3
            if col_i == 17: c.font = font(CYAN, bold=True)        # R:R
            if col_i == 18:                                        # Confidence
                conf_clr = GREEN if d["confidence"] == "HIGH" else BLUE if d["confidence"] == "MEDIUM" else GREY
                c.font = font(conf_clr, bold=True)
                c.fill = fill(conf_bg)
            if col_i == 19:                                        # Status
                c.font = Font(name="Arial", color=s_fg, bold=True, size=9)
                c.fill = fill(s_bg)

        ws.row_dimensions[r].height = 56

    # ── Footer ──
    fr = len(sorted_syms) + 6
    ws.merge_cells(f"A{fr}:T{fr}")
    ws[f"A{fr}"] = ("⚠️  SWING TRADE PROTOCOL: 1) Only enter when status = 🟢 ACCUMULATE or 🚀 BREAKOUT IMMINENT  "
                   "2) Stagger entries (40-30-30%) within entry zone  "
                   "3) Hard SL at -5 to -7% below entry  "
                   "4) Trail SL to entry after T1 hit  "
                   "5) Book 30% at T1, 40% at T2, hold 30% for T3  "
                   "6) Re-evaluate weekly — exit if THESIS BREAKS (not just price)")
    ws[f"A{fr}"].font      = Font(name="Arial", color=AMBER, bold=True, size=9)
    ws[f"A{fr}"].fill      = fill(HEADER_BG)
    ws[f"A{fr}"].alignment = mid()
    ws.row_dimensions[fr].height = 38

    ws.merge_cells(f"A{fr+1}:T{fr+1}")
    ws[f"A{fr+1}"] = ("EXCLUDED FROM THIS LIST  ❌  IT services (Infy, TCS, Wipro, TechM, HCL, LTIM, Mphasis) — AI structural risk  |  "
                     "Vodafone Idea — debt structure broken  |  Yes Bank legacy holders — governance reset  |  "
                     "Any company with active SEBI / SFIO investigation")
    ws[f"A{fr+1}"].font      = Font(name="Arial", color=RED, italic=True, size=8)
    ws[f"A{fr+1}"].fill      = fill(DARK_BG)
    ws[f"A{fr+1}"].alignment = mid()
    ws.row_dimensions[fr+1].height = 16

    ws.merge_cells(f"A{fr+2}:T{fr+2}")
    ws[f"A{fr+2}"] = (f"Auto-generated by Claude Opus 4.7  |  Run date: {TODAY.strftime('%d %b %Y')}  |  "
                     "Sources: Screener.in · MoneyControl · Trendlyne · Business Standard · Company filings  |  "
                     "Verify all prices before trading.  NOT financial advice.")
    ws[f"A{fr+2}"].font      = Font(name="Arial", color=GREY, italic=True, size=8)
    ws[f"A{fr+2}"].fill      = fill(DARK_BG)
    ws[f"A{fr+2}"].alignment = mid()
    ws.row_dimensions[fr+2].height = 14

    ws.freeze_panes = "C5"

    wb.save(EXCEL_PATH)
    print(f"✅ Excel saved: {EXCEL_PATH}\n")

    # ── Console summary ──
    print(f"📊 Fallen Angels Status Summary  ({TODAY.strftime('%d %b %Y')}):\n")
    by_status = {}
    for sym in sorted_syms:
        d = meta[sym]
        s = calc_status(sym, d["entry_zone"], d["sl"], d["cmp"], d.get("catalyst_date"), d.get("earnings_dates"))
        by_status.setdefault(s, []).append(f"{sym} (-{d['drawdown_pct']:.0f}%, {d['confidence']})")
    for status_key in STATUS_META:
        if status_key in by_status:
            print(f"  {status_key}")
            for item in by_status[status_key]:
                print(f"     • {item}")
            print()

if __name__ == "__main__":
    build()
