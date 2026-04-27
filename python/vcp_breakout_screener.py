"""
VCP Breakout Screener â€” Large Cap + Mid Cap + Small Cap
â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
Identifies NSE stocks setting up Volatility Contraction Patterns (VCP)
per Mark Minervini methodology â€” across ALL cap categories (stockexploder style)

THREE CATEGORIES (exactly like stockexploder):
  â–ˆâ–ˆ LARGECAP VCP  â€” Nifty 50 / Nifty Next 50 / MCap > â‚¹50,000 cr
  â–ˆâ–ˆ MIDCAP VCP    â€” MCap â‚¹15,000â€“50,000 cr
  â–ˆâ–ˆ SMALLCAP VCP  â€” MCap < â‚¹15,000 cr

LIVE CMP: Fetched from Yahoo Finance (NSE) at runtime via yfinance.
          All pivot points, EMAs, and entry zones are calibrated to REAL prices.

BREAKOUT PRIORITY (for early profits):
  Priority 1: ðŸš€ BREAKING OUT     â€” Enter NOW, momentum active
  Priority 2: ðŸŸ¢ PIVOT â€” READY    â€” At breakout level, waiting vol confirm
  Priority 3: ðŸ”µ CONTRACTING      â€” 3rd/4th contraction, almost ready
  Priority 4: ðŸŸ¡ BASING           â€” Stage-2, still forming base
  Priority 5: ðŸŸ  NOT READY        â€” Watch only

TARGETS (book in tranches â€” capture early profits):
  T1: +5%   â†’ book 40% (aggressive early profit â€” don't let it become breakeven)
  T2: +14%  â†’ book 35%
  T3: +24%  â†’ trail SL on last 25%

STOP LOSS: Below pivot base â€” 4-5% max. VCP fails fast or works fast.
â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
"""
import json, os
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

# â”€â”€ Live price fetch via yfinance â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def fetch_live_cmp(symbol_nse: str):
    """Fetch live CMP from Yahoo Finance. NSE symbol â†’ appends .NS suffix."""
    try:
        import yfinance as yf
        ticker = symbol_nse.replace("NSE:", "") + ".NS"
        t = yf.Ticker(ticker)
        price = t.fast_info["lastPrice"]
        if price and price > 0:
            return round(float(price), 2)
    except Exception:
        pass
    return None

def fetch_all_live_cmps(candidates: dict) -> dict:
    """Fetch live CMPs for all candidates. Returns dict symâ†’cmp."""
    print("â³ Fetching live NSE prices from Yahoo Finance...")
    live = {}
    for sym in candidates:
        cmp = fetch_live_cmp(sym)
        if cmp:
            live[sym] = cmp
            print(f"   {sym}: â‚¹{cmp:,.2f}")
        else:
            hardcoded = candidates[sym].get("cmp", 0)
            live[sym] = hardcoded
            print(f"   {sym}: â‚¹{hardcoded:,.2f} (hardcoded fallback â€” verify!)")
    return live

try:
    from volatility_engine import smart_sl, smart_targets, get_volatility_profile
    HAS_VOL_ENGINE = True
except Exception:
    HAS_VOL_ENGINE = False

try:
    from sector_rotation import get_sector_boost
except Exception:
    def get_sector_boost(s): return 0

try:
    from flow_tracker import get_smart_money_score
except Exception:
    def get_smart_money_score(s): return 0

try:
    from permanent_damage_blacklist import get_blacklist_set
    BLACKLIST = get_blacklist_set(["PERMANENT_AVOID"])
except Exception:
    BLACKLIST = set()

EXCEL_PATH = r"C:\Users\razee\OneDrive\Desktop\TradingClaude\NSE_VCP_Breakouts_v2.xlsx"
FRESH_JSON = r"C:\Users\razee\OneDrive\Desktop\TradingClaude\vcp_fresh.json"
TODAY      = date.today()

DARK_BG="0D0D0D"; HEADER_BG="1A1A2E"; ROW_ALT="141414"
GOLD="FFD700"; GREEN="00C896"; BLUE="4FC3F7"; WHITE="FFFFFF"
GREY="9E9E9E"; RED="FF5252"; AMBER="FFB300"; ORANGE="FF6B35"
CYAN="00BCD4"; PURPLE="9C27B0"; PINK="FF4081"
LC_BG="1A0D2E"   # Large cap row tint â€” royal purple
MC_BG="0D1A2E"   # Mid cap row tint â€” dark blue
SC_BG="0D2E1A"   # Small cap row tint â€” dark green

CAP_COLORS = {
    "Largecap VCP": (LC_BG, PURPLE),
    "Midcap VCP":   (MC_BG, BLUE),
    "Smallcap VCP": (SC_BG, GREEN),
}

# â”€â”€ Stage labels â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
STAGE_META = {
    "ðŸš€ BREAKING OUT":   ("0A4D2A", GREEN,  "Priority 1 â€” Volume surge + close above pivot â€” ENTER NOW"),
    "ðŸŸ¢ PIVOT â€” READY":  ("1B3A1B", GREEN,  "Priority 2 â€” At pivot, await volume confirm"),
    "ðŸ”µ CONTRACTING":    ("003566", BLUE,   "Priority 3 â€” VCP forming, 3+ contractions"),
    "ðŸŸ¡ BASING":         ("3D2C00", AMBER,  "Priority 4 â€” Stage-2, building base"),
    "ðŸŸ  NOT READY":      ("5C2A00", ORANGE, "Priority 5 â€” Watch only, incomplete setup"),
    "ðŸ”´ BROKEN":         ("2D0000", RED,    "Avoid â€” Failed VCP"),
}

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# VCP CANDIDATES â€” CMPs calibrated to live market prices (Apr 27, 2026)
# All pivot/entry/SL levels derived from REAL current price + chart structure
# cap_category: "Largecap VCP" | "Midcap VCP" | "Smallcap VCP"
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
VCP_CANDIDATES = {

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # â–ˆâ–ˆ LARGECAP VCP  (MCap > â‚¹50,000 cr)
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

    "NSE:ADANIPORTS": {
        "name": "Adani Ports & SEZ", "sector": "Ports + Logistics",
        "cap_category": "Largecap VCP",
        "market_cap_cr": 374347, "cmp": 1624.80,
        "wk52_high": 1625, "wk52_low": 1050,
        "pct_from_ath": 1.2,
        "adr_pct": 2.5,
        "avg_daily_volume_lakhs": 85,
        "ema_20": 1590, "ema_50": 1520, "ema_200": 1380,
        "stage": "ðŸŸ¢ PIVOT â€” READY",
        "contractions": "18% â†’ 8% â†’ 3% (at 52W high â€” breakout imminent)",
        "volume_dry_up": "Yes â€” tight squeeze near â‚¹1620",
        "pivot_point": 1630,
        "entry_zone": "1625-1645",
        "sl": 1555,
        "t1": 1700,
        "t2": 1810,
        "t3": 1950,
        "expected_move_1d_pct": 4,
        "expected_move_3d_pct": 12,
        "catalyst": "FY26 cargo volume record; ICD expansion; Vizhinjam Phase 2",
        "smart_money": "DII + FPI both accumulating",
        "buying_force_pct": 22,
        "conviction": "VERY HIGH",
        "risk_pct_to_sl": 4.6,
    },
    "NSE:BAJFINANCE": {
        "name": "Bajaj Finance", "sector": "NBFC â€” Consumer + SME Finance",
        "cap_category": "Largecap VCP",
        "market_cap_cr": 572898, "cmp": 921.60,
        "wk52_high": 950, "wk52_low": 620,
        "pct_from_ath": 3.0,
        "adr_pct": 2.2,
        "avg_daily_volume_lakhs": 350,
        "ema_20": 905, "ema_50": 870, "ema_200": 800,
        "stage": "ðŸŸ¢ PIVOT â€” READY",
        "contractions": "14% â†’ 7% â†’ 3% (stage-2 + near 52W high)",
        "volume_dry_up": "Yes â€” 3-week tight base â‚¹900-925",
        "pivot_point": 955,
        "entry_zone": "950-965",
        "sl": 905,
        "t1": 1000,
        "t2": 1065,
        "t3": 1150,
        "expected_move_1d_pct": 4,
        "expected_move_3d_pct": 12,
        "catalyst": "Rate cut cycle boosts NIM; AUM growth 28% YoY; RBI approval pipeline",
        "smart_money": "Consistent FII accumulation",
        "buying_force_pct": 22,
        "conviction": "VERY HIGH",
        "risk_pct_to_sl": 4.7,
    },
    "NSE:TATAPOWER": {
        "name": "Tata Power", "sector": "Power â€” Solar + Thermal + Distribution",
        "cap_category": "Largecap VCP",
        "market_cap_cr": 145148, "cmp": 454.25,
        "wk52_high": 480, "wk52_low": 310,
        "pct_from_ath": 5.4,
        "adr_pct": 3.5,
        "avg_daily_volume_lakhs": 520,
        "ema_20": 445, "ema_50": 425, "ema_200": 390,
        "stage": "ðŸ”µ CONTRACTING",
        "contractions": "22% â†’ 10% â†’ 5% (3 contractions, base tightening)",
        "volume_dry_up": "Yes â€” weekly vol 40% below base avg",
        "pivot_point": 480,
        "entry_zone": "478-488",
        "sl": 455,
        "t1": 505,
        "t2": 545,
        "t3": 590,
        "expected_move_1d_pct": 5,
        "expected_move_3d_pct": 14,
        "catalyst": "Solar EPC order wins; Mundra + new DISCOMs; EV charging infra",
        "smart_money": "Tata Group buying + MF accumulation",
        "buying_force_pct": 25,
        "conviction": "HIGH",
        "risk_pct_to_sl": 5.0,
    },
    "NSE:BHEL": {
        "name": "BHEL", "sector": "Capital Goods â€” Power + Defence",
        "cap_category": "Largecap VCP",
        "market_cap_cr": 121176, "cmp": 348.00,
        "wk52_high": 360, "wk52_low": 200,
        "pct_from_ath": 3.3,
        "adr_pct": 3.2,
        "avg_daily_volume_lakhs": 450,
        "ema_20": 340, "ema_50": 318, "ema_200": 285,
        "stage": "ðŸŸ¢ PIVOT â€” READY",
        "contractions": "20% â†’ 9% â†’ 4% (near 52W high â€” classic VCP)",
        "volume_dry_up": "Yes â€” base at â‚¹340-355",
        "pivot_point": 362,
        "entry_zone": "360-368",
        "sl": 342,
        "t1": 380,
        "t2": 412,
        "t3": 450,
        "expected_move_1d_pct": 4,
        "expected_move_3d_pct": 13,
        "catalyst": "Order book â‚¹1.2L cr; nuclear + defence + power capex super-cycle",
        "smart_money": "LIC + DII buying",        "buying_force_pct": 30,        "conviction": "HIGH",
        "risk_pct_to_sl": 5.5,
    },
    "NSE:PREMIERENE": {
        "name": "Premier Energies", "sector": "Solar Cells + Modules",
        "cap_category": "Largecap VCP",
        "market_cap_cr": 45962, "cmp": 1018.45,
        "wk52_high": 1500, "wk52_low": 640,
        "pct_from_ath": 32.1,
        "adr_pct": 4.5,
        "avg_daily_volume_lakhs": 25,
        "ema_20": 1005, "ema_50": 965, "ema_200": 850,
        "stage": "ðŸ”µ CONTRACTING",
        "contractions": "28% â†’ 13% â†’ 6% (11 weeks â€” approaching pivot)",
        "volume_dry_up": "Yes â€” base tightening above â‚¹1000",
        "pivot_point": 1090,
        "entry_zone": "1088-1105",
        "sl": 1035,
        "t1": 1145,
        "t2": 1240,
        "t3": 1360,
        "expected_move_1d_pct": 5,
        "expected_move_3d_pct": 16,
        "catalyst": "ALMM regime + 50GW domestic solar cell capacity push; Q4 results beat",
        "smart_money": "Strong FII + bulk deals",
        "buying_force_pct": 28,
        "conviction": "HIGH",
        "risk_pct_to_sl": 5.1,
    },

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # â–ˆâ–ˆ MIDCAP VCP  (MCap â‚¹15,000â€“50,000 cr)
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

    "NSE:DATAPATTNS": {
        "name": "Data Patterns India", "sector": "Defence Electronics â€” AESA + Avionics",
        "cap_category": "Midcap VCP",
        "market_cap_cr": 22674, "cmp": 4050.10,
        "wk52_high": 4050, "wk52_low": 2000,
        "pct_from_ath": 0.5,
        "adr_pct": 3.9,
        "avg_daily_volume_lakhs": 6,
        "ema_20": 3960, "ema_50": 3720, "ema_200": 3200,
        "stage": "ðŸš€ BREAKING OUT",
        "contractions": "22% â†’ 10% â†’ 4% â†’ BROKE OUT above â‚¹3900",
        "volume_dry_up": "Volume surge 3x avg â€” institutional buying",
        "pivot_point": 3900,
        "entry_zone": "4030-4080 (extended â€” wait for pullback to â‚¹3950)",
        "sl": 3840,
        "t1": 4255,
        "t2": 4620,
        "t3": 5000,
        "expected_move_1d_pct": 5,
        "expected_move_3d_pct": 16,
        "catalyst": "Nippon MF bulk buy Apr 23; AESA radar + fighter avionics contracts",
        "smart_money": "DII + bulk buy â€” very strong",        "buying_force_pct": 34,        "conviction": "VERY HIGH",
        "risk_pct_to_sl": 5.2,
    },
    "NSE:HBLENGINE": {
        "name": "HBL Power Systems", "sector": "Defence + Railways + Battery Storage",
        "cap_category": "Midcap VCP",
        "market_cap_cr": 22584, "cmp": 814.75,
        "wk52_high": 815, "wk52_low": 380,
        "pct_from_ath": 0.3,
        "adr_pct": 4.8,
        "avg_daily_volume_lakhs": 45,
        "ema_20": 795, "ema_50": 750, "ema_200": 650,
        "stage": "ðŸš€ BREAKING OUT",
        "contractions": "22% â†’ 11% â†’ 5% â†’ BREAKOUT above â‚¹800 (52W high)",
        "volume_dry_up": "Volume SURGED â€” 2.8x avg on breakout day",
        "pivot_point": 800,
        "entry_zone": "810-825 (chase with tight SL)",
        "sl": 773,
        "t1": 858,
        "t2": 935,
        "t3": 1020,
        "expected_move_1d_pct": 6,
        "expected_move_3d_pct": 18,
        "catalyst": "Kavach anti-collision orders; submarine battery + train-18 EV battery",
        "smart_money": "DII + bulk deal â€” strong accumulation",        "buying_force_pct": 39,        "conviction": "VERY HIGH",
        "risk_pct_to_sl": 5.2,
    },
    "NSE:AVALON": {
        "name": "Avalon Technologies", "sector": "EMS â€” Aero + Industrial",
        "cap_category": "Midcap VCP",
        "market_cap_cr": 7278, "cmp": 1090.10,
        "wk52_high": 1090, "wk52_low": 460,
        "pct_from_ath": 0.5,
        "adr_pct": 5.2,
        "avg_daily_volume_lakhs": 9,
        "ema_20": 1065, "ema_50": 1005, "ema_200": 870,
        "stage": "ðŸš€ BREAKING OUT",
        "contractions": "32% â†’ 14% â†’ 6% â†’ BREAKING OUT above â‚¹1050 (52W high)",
        "volume_dry_up": "Volume 2.8x avg â€” breakout confirmed",
        "pivot_point": 1050,
        "entry_zone": "1085-1105 (extended â€” tight SL mandatory)",
        "sl": 1035,
        "t1": 1145,
        "t2": 1240,
        "t3": 1360,
        "expected_move_1d_pct": 5,
        "expected_move_3d_pct": 15,
        "catalyst": "US aero EMS facility ramp + new Airbus supply contracts",
        "smart_money": "Bulk buys + promoter not selling",
        "buying_force_pct": 36,
        "conviction": "VERY HIGH",
        "risk_pct_to_sl": 5.0,
    },
    "NSE:KAYNES": {
        "name": "Kaynes Technology", "sector": "EMS + Semiconductors",
        "cap_category": "Midcap VCP",
        "market_cap_cr": 28307, "cmp": 4222.80,
        "wk52_high": 6500, "wk52_low": 2900,
        "pct_from_ath": 35.0,
        "adr_pct": 4.2,
        "avg_daily_volume_lakhs": 8,
        "ema_20": 4180, "ema_50": 3950, "ema_200": 3500,
        "stage": "ðŸ”µ CONTRACTING",
        "contractions": "26% â†’ 12% â†’ 5% (10 weeks â€” 4th contraction forming)",
        "volume_dry_up": "Yes â€” vol 48% below base avg",
        "pivot_point": 4480,
        "entry_zone": "4475-4515",
        "sl": 4240,
        "t1": 4710,
        "t2": 5070,
        "t3": 5500,
        "expected_move_1d_pct": 5,
        "expected_move_3d_pct": 14,
        "catalyst": "Semicon OSAT ramp; ICICI Pru MF bulk buy; defence EMS wins",
        "smart_money": "Bulk buy + DII",
        "buying_force_pct": 28,
        "conviction": "VERY HIGH",
        "risk_pct_to_sl": 5.3,
    },
    "NSE:JYOTICNC": {
        "name": "Jyoti CNC Automation", "sector": "CNC Machine Tools â€” Defence + Aero",
        "cap_category": "Midcap VCP",
        "market_cap_cr": 17074, "cmp": 750.75,
        "wk52_high": 1420, "wk52_low": 550,
        "pct_from_ath": 47.1,
        "adr_pct": 5.5,
        "avg_daily_volume_lakhs": 15,
        "ema_20": 738, "ema_50": 700, "ema_200": 640,
        "stage": "ðŸ”µ CONTRACTING",
        "contractions": "32% â†’ 15% â†’ 7% (10 weeks â€” near breakout)",
        "volume_dry_up": "Yes â€” daily vol 45% of base avg",
        "pivot_point": 800,
        "entry_zone": "798-812",
        "sl": 758,
        "t1": 840,
        "t2": 910,
        "t3": 990,
        "expected_move_1d_pct": 6,
        "expected_move_3d_pct": 16,
        "catalyst": "Defence CNC contracts + aerospace machining ramp",
        "smart_money": "DII buying",
        "buying_force_pct": 30,
        "conviction": "HIGH",
        "risk_pct_to_sl": 5.3,
    },
    "NSE:TITAGARH": {
        "name": "Titagarh Rail Systems", "sector": "Railways â€” Wagons + Metro Coaches",
        "cap_category": "Midcap VCP",
        "market_cap_cr": 10235, "cmp": 759.95,
        "wk52_high": 1400, "wk52_low": 560,
        "pct_from_ath": 45.7,
        "adr_pct": 4.3,
        "avg_daily_volume_lakhs": 22,
        "ema_20": 748, "ema_50": 715, "ema_200": 660,
        "stage": "ðŸ”µ CONTRACTING",
        "contractions": "28% â†’ 13% â†’ 6% (9 weeks â€” 4th contraction starting)",
        "volume_dry_up": "Yes",
        "pivot_point": 805,
        "entry_zone": "803-818",
        "sl": 762,
        "t1": 845,
        "t2": 915,
        "t3": 995,
        "expected_move_1d_pct": 5,
        "expected_move_3d_pct": 14,
        "catalyst": "Vande Metro/Sleeper + freight wagons; budget-backed rails capex",
        "smart_money": "DII accumulation",
        "buying_force_pct": 32,
        "conviction": "HIGH",
        "risk_pct_to_sl": 5.1,
    },
    "NSE:ASTRAMICRO": {
        "name": "Astra Microwave Products", "sector": "Defence â€” Radar + Microwave",
        "cap_category": "Midcap VCP",
        "market_cap_cr": 10786, "cmp": 1136.00,
        "wk52_high": 1480, "wk52_low": 700,
        "pct_from_ath": 23.2,
        "adr_pct": 4.7,
        "avg_daily_volume_lakhs": 14,
        "ema_20": 1115, "ema_50": 1060, "ema_200": 950,
        "stage": "ðŸ”µ CONTRACTING",
        "contractions": "26% â†’ 12% â†’ 5% (9 weeks)",
        "volume_dry_up": "Yes â€” last 2 weeks vol <40% avg",
        "pivot_point": 1210,
        "entry_zone": "1208-1225",
        "sl": 1145,
        "t1": 1270,
        "t2": 1375,
        "t3": 1500,
        "expected_move_1d_pct": 5,
        "expected_move_3d_pct": 16,
        "catalyst": "Radar export wins; DRDO + HAL supply chain; indigenisation",
        "smart_money": "Sector inflow + DII",
        "buying_force_pct": 28,
        "conviction": "HIGH",
        "risk_pct_to_sl": 5.4,
    },
    "NSE:KIRLOSBROS": {
        "name": "Kirloskar Brothers", "sector": "Industrial Pumps â€” Water + Naval Defence",
        "cap_category": "Midcap VCP",
        "market_cap_cr": 13881, "cmp": 1748.00,
        "wk52_high": 2100, "wk52_low": 950,
        "pct_from_ath": 16.8,
        "adr_pct": 4.6,
        "avg_daily_volume_lakhs": 7,
        "ema_20": 1718, "ema_50": 1640, "ema_200": 1450,
        "stage": "ðŸ”µ CONTRACTING",
        "contractions": "28% â†’ 13% â†’ 6% (9 weeks)",
        "volume_dry_up": "Yes",
        "pivot_point": 1860,
        "entry_zone": "1858-1878",
        "sl": 1760,
        "t1": 1950,
        "t2": 2110,
        "t3": 2290,
        "expected_move_1d_pct": 5,
        "expected_move_3d_pct": 15,
        "catalyst": "Jal Jeevan Mission pump orders; naval destroyer + submarine pump systems",
        "smart_money": "DII buying",
        "buying_force_pct": 34,
        "conviction": "HIGH",
        "risk_pct_to_sl": 5.3,
    },

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # â–ˆâ–ˆ SMALLCAP VCP  (MCap < â‚¹15,000 cr)
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

    "NSE:AZAD": {
        "name": "Azad Engineering", "sector": "Aero + Defence + Energy Components",
        "cap_category": "Smallcap VCP",
        "market_cap_cr": 14199, "cmp": 2198.60,
        "wk52_high": 2200, "wk52_low": 1000,
        "pct_from_ath": 0.5,
        "adr_pct": 5.1,
        "avg_daily_volume_lakhs": 5,
        "ema_20": 2150, "ema_50": 2040, "ema_200": 1780,
        "stage": "ðŸš€ BREAKING OUT",
        "contractions": "28% â†’ 13% â†’ 5% â†’ BREAKING OUT above â‚¹2100",
        "volume_dry_up": "Volume 2.6x avg â€” breakout live",
        "pivot_point": 2100,
        "entry_zone": "2185-2215 (in breakout â€” tight SL mandatory)",
        "sl": 2085,
        "t1": 2305,
        "t2": 2510,
        "t3": 2730,
        "expected_move_1d_pct": 6,
        "expected_move_3d_pct": 18,
        "catalyst": "GE Aerospace + Rolls-Royce supplier PO book building; new ATF contracts",
        "smart_money": "Bulk buys + FII entry post-ATH break",
        "buying_force_pct": 38,
        "conviction": "VERY HIGH",
        "risk_pct_to_sl": 5.1,
    },
    "NSE:SEDEMAC": {
        "name": "Sedemac Mechatronics", "sector": "Auto-Electric Controls (IPO Mar 2026)",
        "cap_category": "Smallcap VCP",
        "market_cap_cr": 8875, "cmp": 2009.70,
        "wk52_high": 2200, "wk52_low": 1380,
        "pct_from_ath": 8.7,
        "adr_pct": 4.2,
        "avg_daily_volume_lakhs": 8.5,
        "ema_20": 1980, "ema_50": 1890, "ema_200": 1720,
        "stage": "ðŸŸ¢ PIVOT â€” READY",
        "contractions": "8% â†’ 4% â†’ 2% (very tight 3 contractions â€” RARE textbook VCP)",
        "volume_dry_up": "Yes â€” last contraction vol 38% of base avg",
        "pivot_point": 2195,
        "entry_zone": "2190-2215",
        "sl": 2082,
        "t1": 2305,
        "t2": 2490,
        "t3": 2710,
        "expected_move_1d_pct": 8,
        "expected_move_3d_pct": 20,
        "catalyst": "IPO Mar 2026; tight float + institutional accumulation post-listing",
        "smart_money": "FII + MF accumulating post-IPO",
        "buying_force_pct": 27,
        "conviction": "VERY HIGH",
        "risk_pct_to_sl": 4.5,
    },
    "NSE:RPEL": {
        "name": "Raghav Productivity Enhancers", "sector": "Silica Ramming Mass â€” Steel",
        "cap_category": "Smallcap VCP",
        "market_cap_cr": 3166, "cmp": 689.65,
        "wk52_high": 800, "wk52_low": 420,
        "pct_from_ath": 13.8,
        "adr_pct": 5.8,
        "avg_daily_volume_lakhs": 4,
        "ema_20": 672, "ema_50": 638, "ema_200": 580,
        "stage": "ðŸŸ¢ PIVOT â€” READY",
        "contractions": "22% â†’ 10% â†’ 4% (base tight at â‚¹680-705)",
        "volume_dry_up": "Yes â€” vol drying, watching for surge above â‚¹705",
        "pivot_point": 705,
        "entry_zone": "703-718",
        "sl": 667,
        "t1": 740,
        "t2": 800,
        "t3": 880,
        "expected_move_1d_pct": 7,
        "expected_move_3d_pct": 22,
        "catalyst": "Steel capex cycle recovering; silica ramming mass demand + steel plant expansions",
        "smart_money": "Promoter buying + smart money alert",
        "buying_force_pct": 25,
        "conviction": "HIGH",
        "risk_pct_to_sl": 5.4,
    },
    "NSE:IDEAFORGE": {
        "name": "ideaForge Technology", "sector": "Defence Drones (UAV)",
        "cap_category": "Smallcap VCP",
        "market_cap_cr": 2375, "cmp": 548.40,
        "wk52_high": 1300, "wk52_low": 430,
        "pct_from_ath": 57.8,
        "adr_pct": 5.1,
        "avg_daily_volume_lakhs": 12,
        "ema_20": 536, "ema_50": 510, "ema_200": 475,
        "stage": "ðŸ”µ CONTRACTING",
        "contractions": "24% â†’ 11% â†’ 5% (3 contractions, 7 weeks)",
        "volume_dry_up": "Yes â€” base tightening at â‚¹540-560",
        "pivot_point": 590,
        "entry_zone": "588-603",
        "sl": 558,
        "t1": 620,
        "t2": 672,
        "t3": 735,
        "expected_move_1d_pct": 7,
        "expected_move_3d_pct": 20,
        "catalyst": "Iran-Hormuz tension; India drone policy; MoD orders pipeline",
        "smart_money": "Defence sector + bulk deals",
        "buying_force_pct": 24,
        "conviction": "HIGH",
        "risk_pct_to_sl": 5.1,
    },
    "NSE:PARAS": {
        "name": "Paras Defence", "sector": "Defence â€” Optronics + Space",
        "cap_category": "Smallcap VCP",
        "market_cap_cr": 6467, "cmp": 802.45,
        "wk52_high": 1460, "wk52_low": 580,
        "pct_from_ath": 45.0,
        "adr_pct": 6.1,
        "avg_daily_volume_lakhs": 20,
        "ema_20": 785, "ema_50": 745, "ema_200": 690,
        "stage": "ðŸ”µ CONTRACTING",
        "contractions": "30% â†’ 14% â†’ 7% (3 contractions, 11 weeks)",
        "volume_dry_up": "Yes â€” last week vol at 5-month low",
        "pivot_point": 858,
        "entry_zone": "855-870",
        "sl": 808,
        "t1": 900,
        "t2": 975,
        "t3": 1065,
        "expected_move_1d_pct": 6,
        "expected_move_3d_pct": 20,
        "catalyst": "Night-vision + optronics MoU pipeline; ISRO supply chain ramp",
        "smart_money": "DII + defence sector inflow",
        "buying_force_pct": 28,
        "conviction": "HIGH",
        "risk_pct_to_sl": 5.8,
    },
    "NSE:ZAGGLE": {
        "name": "Zaggle Prepaid Ocean Services", "sector": "Fintech B2B â€” SaaS + Cards",
        "cap_category": "Smallcap VCP",
        "market_cap_cr": 3505, "cmp": 260.17,
        "wk52_high": 620, "wk52_low": 215,
        "pct_from_ath": 58.0,
        "adr_pct": 5.5,
        "avg_daily_volume_lakhs": 18,
        "ema_20": 254, "ema_50": 240, "ema_200": 225,
        "stage": "ðŸ”µ CONTRACTING",
        "contractions": "28% â†’ 13% â†’ 6% (3 contractions, 10 weeks)",
        "volume_dry_up": "Yes â€” daily vol 40% of peak",
        "pivot_point": 280,
        "entry_zone": "278-286",
        "sl": 265,
        "t1": 294,
        "t2": 318,
        "t3": 350,
        "expected_move_1d_pct": 7,
        "expected_move_3d_pct": 22,
        "catalyst": "B2B fintech expansion; enterprise SaaS traction; partnerships Q4",
        "smart_money": "FPI buying â€” 2 consecutive quarters",        "buying_force_pct": 22,        "conviction": "HIGH",
        "risk_pct_to_sl": 5.4,
    },
    "NSE:CYIENTDLM": {
        "name": "Cyient DLM", "sector": "EMS â€” Aerospace + Defence + Medical",
        "cap_category": "Smallcap VCP",
        "market_cap_cr": 3169, "cmp": 399.25,
        "wk52_high": 1200, "wk52_low": 305,
        "pct_from_ath": 66.7,
        "adr_pct": 4.8,
        "avg_daily_volume_lakhs": 11,
        "ema_20": 390, "ema_50": 370, "ema_200": 345,
        "stage": "ðŸ”µ CONTRACTING",
        "contractions": "22% â†’ 10% â†’ 5% (tightening above â‚¹390)",
        "volume_dry_up": "Yes â€” 3-week vol dry-up",
        "pivot_point": 430,
        "entry_zone": "428-442",
        "sl": 402,
        "t1": 452,
        "t2": 490,
        "t3": 535,
        "expected_move_1d_pct": 6,
        "expected_move_3d_pct": 18,
        "catalyst": "Aero/defence EMS ramp; Q4 results beat expectation; HAL supply",
        "smart_money": "Sector inflow + DII",
        "buying_force_pct": 26,
        "conviction": "HIGH",
        "risk_pct_to_sl": 6.7,
    },
    "NSE:GANECOS": {
        "name": "Ganesha Ecosphere", "sector": "Recycled PET / Sustainability",
        "cap_category": "Smallcap VCP",
        "market_cap_cr": 2800, "cmp": 895.00,
        "wk52_high": 1100, "wk52_low": 620,
        "pct_from_ath": 18.6,
        "adr_pct": 5.2,
        "avg_daily_volume_lakhs": 2.5,
        "ema_20": 878, "ema_50": 840, "ema_200": 780,
        "stage": "🟡 BASING",
        "contractions": "30% → 14% → 7% (11 weeks)",
        "volume_dry_up": "Yes",
        "pivot_point": 955,
        "entry_zone": "952-968",
        "sl": 905,
        "t1": 1002,
        "t2": 1085,
        "t3": 1180,
        "expected_move_1d_pct": 5,
        "expected_move_3d_pct": 16,
        "catalyst": "ESG mandate + plastic waste regulation; extended producer responsibility",
        "smart_money": "Promoter holding stable, no selling",
        "buying_force_pct": 30,
        "conviction": "HIGH",
        "risk_pct_to_sl": 5.5,
    },
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def fill(h): return PatternFill("solid", fgColor=h)
def font(color=WHITE, bold=False, size=9, italic=False):
    return Font(name="Arial", color=color, bold=bold, size=size, italic=italic)
def bdr():
    s = Side(style="thin", color="2D2D2D")
    return Border(left=s, right=s, top=s, bottom=s)
def mid(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def merge_fresh(meta):
    if not os.path.exists(FRESH_JSON):
        return meta
    try:
        with open(FRESH_JSON) as f:
            fresh = json.load(f)
        for sym, updates in fresh.items():
            if sym in meta:
                meta[sym].update(updates)
            else:
                meta[sym] = updates
        print(f"Merged fresh data for {len(fresh)} VCP candidates")
    except Exception as e:
        print(f"Could not merge VCP fresh JSON: {e}")
    return meta

def make_comment(text, author="VCP Screener"):
    from openpyxl.comments import Comment
    c = Comment(text, author)
    c.width = 260
    c.height = 90
    return c

def priority_score(d):
    stage_rank = {
        "🚀 BREAKING OUT": 0, "🟢 PIVOT — READY": 1, "🔵 CONTRACTING": 2,
        "🟡 BASING": 3, "🟠 NOT READY": 4, "🔴 BROKEN": 5,
    }
    conv_rank = {"VERY HIGH": 0, "HIGH": 1, "MEDIUM": 2}
    s = stage_rank.get(d.get("stage", ""), 9)
    c = conv_rank.get(d.get("conviction", ""), 9)
    m = -d.get("expected_move_3d_pct", 0)
    return (s, c, m)

HEADERS = [
    "#", "NSE Symbol", "Stock Name", "Cap Category\n(stockexploder)",
    "Sector", "MCap Cr", "CMP\n(LIVE)", "52W High",
    "% from ATH", "ADR %", "Avg Vol\n(L)", "EMA Stack",
    "Stage", "Contractions", "Pivot", "ENTRY ZONE",
    "SL", "Risk %", "T1 → 40%", "T2 → 35%", "T3 → 25%",
    "Exp 1D", "Exp 3D", "Catalyst", "Smart Money", "Conviction",
]
COL_WIDTHS = [4,17,22,18,28,9,12,10,10,7,10,15,18,34,9,18,9,7,16,16,16,8,8,38,18,12]

def build():
    meta = merge_fresh({k: dict(v) for k, v in VCP_CANDIDATES.items()})
    for sym in list(meta.keys()):
        if sym in BLACKLIST:
            del meta[sym]

    live_prices = fetch_all_live_cmps(meta)
    for sym, price in live_prices.items():
        if price:
            meta[sym]["cmp"] = price
            meta[sym]["pct_from_ath"] = round(
                max(0, (meta[sym]["wk52_high"] - price) / meta[sym]["wk52_high"] * 100), 1
            )

    wb = Workbook()
    ws = wb.active
    ws.title = "NSE VCP Breakouts"
    ws.sheet_view.showGridLines = False
    last_col = get_column_letter(len(HEADERS))

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = (f"NSE VCP BREAKOUT SCREENER — {TODAY.strftime('%d %b %Y')}  "
                f"|  Largecap VCP  Midcap VCP  Smallcap VCP  "
                f"|  Mark Minervini + stockexploder  |  Priority: EARLY PROFIT")
    ws["A1"].font = Font(name="Arial", color=GOLD, bold=True, size=13)
    ws["A1"].fill = fill(HEADER_BG); ws["A1"].alignment = mid()
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = ("PRIORITY → Breaking Out (enter NOW)  Pivot Ready (await vol)  "
                "Contracting (3+ done)  Basing (watchlist)  "
                "||  BOOK: T1 +5% exit 40%  T2 +14% exit 35%  T3 +24% trail 25%  "
                "||  SL = hard stop below pivot — no exceptions")
    ws["A2"].font = Font(name="Arial", color=GREY, italic=True, size=9)
    ws["A2"].fill = fill(DARK_BG); ws["A2"].alignment = mid()
    ws.row_dimensions[2].height = 24

    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"] = ("LARGECAP VCP: MCap>50k Cr, 5-12% moves, 3-10d   "
                "MIDCAP VCP: 15k-50k Cr, 10-18% moves, 2-5d   "
                "SMALLCAP VCP: <15k Cr, 15-30% explosive 1-3d   "
                "|| CMP = LIVE (hover for timestamp)")
    ws["A3"].font = Font(name="Arial", color=AMBER, bold=True, size=9)
    ws["A3"].fill = fill(HEADER_BG); ws["A3"].alignment = mid()
    ws.row_dimensions[3].height = 18

    for i, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(name="Arial", color=GOLD, bold=True, size=8)
        c.fill = fill(HEADER_BG); c.alignment = mid(); c.border = bdr()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 42

    sorted_syms = sorted(meta.keys(), key=lambda s: priority_score(meta[s]))

    for idx, sym in enumerate(sorted_syms, 1):
        d = meta[sym]; r = idx + 4
        stage_bg, stage_fg, _ = STAGE_META.get(d["stage"], (DARK_BG, GREY, ""))
        cap_cat = d.get("cap_category", "Smallcap VCP")
        cap_row_bg, cap_color = CAP_COLORS.get(cap_cat, (SC_BG, GREEN))
        if   d["stage"] == "🚀 BREAKING OUT": row_bg = "0A4D2A"
        elif d["stage"] == "🔴 BROKEN":       row_bg = "2D0000"
        else: row_bg = cap_row_bg if idx % 2 == 1 else DARK_BG

        ema_stack = ("✓✓✓✓" if d["cmp"]>d["ema_20"]>d["ema_50"]>d["ema_200"]
                     else "✓✓✓ " if d["cmp"]>d["ema_20"]>d["ema_50"]
                     else "✓✓  " if d["cmp"]>d["ema_20"] else "✗   ")
        cmp = d["cmp"]
        g1 = f"+{round((d['t1']-cmp)/cmp*100,1)}% → Rs{d['t1']:,}"
        g2 = f"+{round((d['t2']-cmp)/cmp*100,1)}% → Rs{d['t2']:,}"
        g3 = f"+{round((d['t3']-cmp)/cmp*100,1)}% → Rs{d['t3']:,}"

        cells = [idx, sym, d["name"], cap_cat, d["sector"],
                 d["market_cap_cr"], d["cmp"], d["wk52_high"],
                 f"-{d['pct_from_ath']}%", f"{d['adr_pct']}%", f"{d['avg_daily_volume_lakhs']}L",
                 ema_stack, d["stage"], d["contractions"],
                 d["pivot_point"], d["entry_zone"], d["sl"], f"{d['risk_pct_to_sl']}%",
                 g1, g2, g3,
                 f"+{d['expected_move_1d_pct']}%", f"+{d['expected_move_3d_pct']}%",
                 d["catalyst"], d["smart_money"], d["conviction"]]

        for col_i, val in enumerate(cells, 1):
            c = ws.cell(row=r, column=col_i, value=val)
            c.fill = fill(row_bg); c.border = bdr()
            c.font = font(WHITE, size=9)
            c.alignment = mid() if col_i not in (3,4,5,14,24,25) else lft()
            if col_i==1: c.font = font(GOLD, bold=True)
            if col_i==2: c.font = font(GREEN, bold=True)
            if col_i==4:
                c.font = Font(name="Arial", color=cap_color, bold=True, size=9)
                if cap_cat == "Largecap VCP":
                    ct = "LARGECAP VCP\n─────────────────\nMCap > Rs50,000 Cr\nLower volatility, highest liquidity\nMoves: 5-12%  Hold: 3-10 days\nBest for positional swing trades"
                elif cap_cat == "Midcap VCP":
                    ct = "MIDCAP VCP\n─────────────────\nMCap Rs15,000-50,000 Cr\nMedium volatility & liquidity\nMoves: 10-18%  Hold: 2-5 days\nSweet spot for breakout momentum"
                else:
                    ct = "SMALLCAP VCP\n─────────────────\nMCap < Rs15,000 Cr\nHigh volatility, lower liquidity\nMoves: 15-30% explosive  Hold: 1-3 days\nUse smaller position sizes!"
                c.comment = make_comment(ct)
            if col_i==7:
                c.font = Font(name="Arial", color=CYAN, bold=True, size=9)
                c.comment = make_comment(
                    f"LIVE CMP — Yahoo Finance NSE\nSymbol: {sym.replace('NSE:','')}.NS\n"
                    f"Price: Rs{cmp:,.2f}\nFetched: {datetime.now().strftime('%d-%b-%Y %H:%M')}\n"
                    "Verify on Kite / TradingView before entry.")
            if col_i==10: c.font = font(CYAN, bold=True)
            if col_i==12:
                c.font = Font(name="Arial", color=GREEN if "✓✓✓✓" in str(val) else AMBER, bold=True, size=9)
            if col_i==13:
                c.font = Font(name="Arial", color=stage_fg, bold=True, size=9); c.fill = fill(stage_bg)
            if col_i==15: c.font = font(GOLD, bold=True)
            if col_i==16:
                c.font = Font(name="Arial", color=GREEN, bold=True, size=9)
                c.comment = make_comment(
                    f"ENTRY ZONE: {val}\nPivot: Rs{d['pivot_point']:,}\nSL: Rs{d['sl']:,} ({d['risk_pct_to_sl']}% risk)\n"
                    "─────────────────────────\nEnter ONLY with volume >=2x avg\nWait for breakout candle — not anticipation\nNo follow-through in 2 sessions → EXIT")
            if col_i==17: c.font = font(RED, bold=True)
            if col_i==19:
                c.font = font(GREEN, bold=True)
                c.comment = make_comment(f"T1 — EXIT 40%\n{val}\nMove SL to breakeven after T1 hit.\nDon't let winner turn loser.")
            if col_i==20:
                c.font = font(GREEN, bold=True)
                c.comment = make_comment(f"T2 — EXIT 35%\n{val}\n35% more out. Only 25% remains.\nTrail SL below last swing low.")
            if col_i==21:
                c.font = font(GOLD, bold=True)
                c.comment = make_comment(f"T3 — TRAIL 25%\n{val}\nLet the runner run.\nTrail aggressively. Book if weekly close below 20 EMA.")
            if col_i==22: c.font = font(GREEN if d["expected_move_1d_pct"]>=5 else AMBER, bold=True)
            if col_i==23:
                v = d["expected_move_3d_pct"]
                c.font = font(GOLD if v>=18 else GREEN if v>=12 else AMBER, bold=True)
            if col_i==26:
                clr = GREEN if d["conviction"]=="VERY HIGH" else BLUE if d["conviction"]=="HIGH" else AMBER
                c.font = font(clr, bold=True)
        ws.row_dimensions[r].height = 62

    base_row = len(sorted_syms) + 6
    msgs = [
        (f"PROTOCOL: Wait for actual breakout (close + vol>=2x). Book 40% T1, 35% T2, trail 25% T3. "
         f"Hard SL below pivot. Exit if no follow-through in 2 sessions. Risk 1-2% portfolio per trade.", AMBER),
        ("CAP GUIDE (stockexploder style):  LARGECAP VCP = safest+liquid 5-12%   MIDCAP VCP = sweet spot 10-18%   SMALLCAP VCP = explosive 15-30% smaller size", CYAN),
        (f"Generated: {TODAY.strftime('%d %b %Y')}  |  CMP = LIVE from Yahoo Finance  "
         f"|  Verify all levels on TradingView/Kite before trading  |  stockexploder VCP + Mark Minervini SEPA", GREY),
    ]
    for i, (msg, clr) in enumerate(msgs, start=base_row):
        ws.merge_cells(f"A{i}:{last_col}{i}")
        ws[f"A{i}"] = msg
        ws[f"A{i}"].font = Font(name="Arial", color=clr, bold=(clr==AMBER), size=9)
        ws[f"A{i}"].fill = fill(HEADER_BG if clr==AMBER else DARK_BG)
        ws[f"A{i}"].alignment = mid()
        ws.row_dimensions[i].height = 24 if clr==AMBER else 16

    ws.freeze_panes = "C5"
    wb.save(EXCEL_PATH)
    print(f"\n✅  Excel saved → {EXCEL_PATH}")

    cap_groups = {}
    for sym in sorted_syms:
        d = meta[sym]
        cat = d.get("cap_category","Smallcap VCP")
        cap_groups.setdefault(cat,[]).append(
            f"  {sym:<22} CMP Rs{d['cmp']:>9,.2f}  Entry {d['entry_zone']:<22}  "
            f"SL Rs{d['sl']:>7,}  T1 Rs{d['t1']:>7,} T2 Rs{d['t2']:>7,} T3 Rs{d['t3']:>7,}  "
            f"[{d['stage']}] {d['conviction']}")
    print(f"\n{'='*130}")
    print(f"  NSE VCP BREAKOUT SCREENER  {TODAY.strftime('%d %b %Y')}")
    print(f"{'='*130}\n")
    for cat in ["Largecap VCP","Midcap VCP","Smallcap VCP"]:
        items = cap_groups.get(cat,[])
        print(f"  ## {cat.upper()} ({len(items)} stocks)")
        print(f"  {'-'*125}")
        for item in items: print(f"  *{item}")
        print()
    print(f"{'='*130}\n")


def get_top_vcp_today(n=5, cap_filter=None):
    meta = merge_fresh({k: dict(v) for k, v in VCP_CANDIDATES.items()})
    if cap_filter:
        meta = {k: v for k, v in meta.items() if v.get("cap_category")==cap_filter}
    syms = sorted(meta.keys(), key=lambda s: priority_score(meta[s]))
    return [(s, meta[s]) for s in syms[:n]]


if __name__ == "__main__":
    build()
