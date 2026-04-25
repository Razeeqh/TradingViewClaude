"""
NSE / BSE IPO Listing-Gain Hunter
─────────────────────────────────────────────────────────────────────────────
Identifies IPOs with HIGH probability of strong listing gains based on:

  1. Subscription ratio — total ≥ 30x, QIB ≥ 20x, HNI ≥ 50x, Retail ≥ 5x
  2. GMP — sustained ≥ 30% premium over issue price 2-3 days before listing
  3. Anchor quality — top MFs (SBI, HDFC, ICICI Pru, Nippon, Mirae) + FPIs
  4. Issue size sweet spot — ₹500-3,000 cr (smaller often = better pop)
  5. PE vs sector peers — at-discount or fairly priced
  6. Analyst SUBSCRIBE rating — Motilal Oswal, ICICI Sec, Anand Rathi, etc.

The bi-weekly Opus 4.7 task refreshes via WebSearch (chittorgarh.com,
investorgain.com, NSE/BSE filings) and writes ipo_pipeline_fresh.json.

Sections:
  • OPEN_NOW       — apply this week
  • UPCOMING_2W    — get ready, watch anchor book
  • RECENT_30D     — listed last 30 days, post-listing analysis
  • BIG_PIPELINE   — major 2026 IPOs to watch (Jio, OYO, NSE, etc.)
─────────────────────────────────────────────────────────────────────────────
"""
import json, os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH = r"C:\Users\razee\OneDrive\Desktop\TradingClaude\NSE_IPO_Pipeline.xlsx"
FRESH_JSON = r"C:\Users\razee\OneDrive\Desktop\TradingClaude\ipo_pipeline_fresh.json"
TODAY      = date.today()

DARK_BG="0D0D0D"; HEADER_BG="1A1A2E"; ROW_ALT="141414"
GOLD="FFD700"; GREEN="00C896"; BLUE="4FC3F7"; WHITE="FFFFFF"
GREY="9E9E9E"; RED="FF5252";  AMBER="FFB300"; ORANGE="FF6B35"; CYAN="00BCD4"; PURPLE="9C27B0"

RATING_META = {
    "STRONG SUBSCRIBE": ("1B4332", GREEN,  "🔥 Apply Day 1 max retail; expected 30%+ listing pop"),
    "SUBSCRIBE":        ("003566", BLUE,   "✅ Apply for listing gains"),
    "NEUTRAL":          ("3D2C00", AMBER,  "👁 Apply only with surplus funds"),
    "AVOID":            ("2D0000", RED,    "❌ Skip — overvalued / weak demand / governance"),
}

# ── IPO Pipeline (template — refresh weekly via Opus 4.7 task) ───────────────
# All numbers MUST be verified before applying. The bi-weekly scheduled task
# writes ipo_pipeline_fresh.json with live chittorgarh.com / investorgain.com
# data which fully replaces these template entries.
IPO_PIPELINE = {
    "OPEN_NOW": {
        # Example schema — actual entries written by the scheduled task
        "[Template — Verify Before Apply]": {
            "exchange": "NSE Mainboard",
            "issue_dates": "TBD",
            "listing_date": "TBD",
            "issue_size_cr": 0,
            "price_band": "—",
            "lot_size": 0,
            "min_investment": 0,
            "qib_subscription": "—",
            "hni_subscription": "—",
            "retail_subscription": "—",
            "total_subscription": "—",
            "gmp_rs": 0,
            "gmp_pct": 0,
            "gmp_trend": "—",
            "anchor_quality": "—",
            "anchor_amount_cr": 0,
            "fy26_revenue_cr": 0,
            "fy26_pat_cr": 0,
            "pe_ratio": 0,
            "sector_pe": 0,
            "sector": "—",
            "promoter_post_ipo_pct": 0,
            "use_of_proceeds": "—",
            "analyst_recommendation": "Run weekly task to refresh",
            "expected_listing_gain_pct": 0,
            "rating": "NEUTRAL",
            "rationale": "Template — actual entries refreshed by bi-weekly Opus 4.7 task via chittorgarh.com + investorgain.com",
        },
    },
    "UPCOMING_2W": {
        # Filled by scheduled task with SEBI-approved DRHPs in next 2 weeks
    },
    "RECENT_30D": {
        # Filled with last-30-days listings + listing gain % + post-listing performance
    },
    "BIG_PIPELINE_2026": {
        "Reliance Jio Infocomm": {
            "expected_quarter": "Q2-Q3 FY27",
            "expected_size_cr": 80000,
            "valuation_cr": 1200000,
            "anchor_demand": "Massive — likely fully covered in anchor",
            "watch_for": "DRHP filing date (SEBI portal)",
            "trade_idea": "Buy parent RIL pre-announcement; Jio IPO typically gives 30-50% pop for parent stock",
            "status": "Awaiting DRHP",
        },
        "OYO": {
            "expected_quarter": "Q3-Q4 FY27",
            "expected_size_cr": 8000,
            "valuation_cr": 60000,
            "anchor_demand": "Mixed — depends on FY26 PAT trajectory",
            "watch_for": "FY26 audited results + new DRHP filing",
            "trade_idea": "Wait for fresh DRHP + GMP signal post-anchor day",
            "status": "Re-filing pending",
        },
        "PhonePe": {
            "expected_quarter": "FY27",
            "expected_size_cr": 50000,
            "valuation_cr": 1000000,
            "anchor_demand": "Strong (Walmart-backed)",
            "watch_for": "Conversion to Indian-domiciled entity completion",
            "trade_idea": "Mainboard-only application; expect tight allocation",
            "status": "Domicile shift complete; DRHP pending",
        },
        "NSE (National Stock Exchange)": {
            "expected_quarter": "FY27",
            "expected_size_cr": 30000,
            "valuation_cr": 400000,
            "anchor_demand": "Very High",
            "watch_for": "SEBI clearance for self-listing",
            "trade_idea": "Apply for full retail allocation; monopoly franchise",
            "status": "Awaiting SEBI nod",
        },
        "Tata Capital": {
            "expected_quarter": "Q1-Q2 FY27",
            "expected_size_cr": 15000,
            "valuation_cr": 110000,
            "anchor_demand": "Strong (Tata pedigree)",
            "watch_for": "DRHP + RBI shadow-NBFC compliance",
            "trade_idea": "Apply at moderate band; long-term hold candidate",
            "status": "DRHP filed",
        },
        "HDB Financial Services": {
            "expected_quarter": "Q2 FY27",
            "expected_size_cr": 12500,
            "valuation_cr": 80000,
            "anchor_demand": "Strong (HDFC parent)",
            "watch_for": "Listing gains likely modest (large issue)",
            "trade_idea": "Mainboard apply; expect listing premium 10-15%",
            "status": "DRHP under SEBI review",
        },
        "LG Electronics India": {
            "expected_quarter": "FY27",
            "expected_size_cr": 15000,
            "valuation_cr": 90000,
            "anchor_demand": "Very Strong",
            "watch_for": "Subscription multiples on Day 2-3",
            "trade_idea": "Apply max retail; brand strength = listing pop",
            "status": "DRHP filed Q4 FY26",
        },
        "Swiggy (already listed Nov 2024)": {
            "expected_quarter": "LISTED",
            "trade_idea": "Trade post-lock-in expiry; track quick-commerce burn rate",
            "status": "LISTED — track for swing trade only",
        },
    },
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def fill(h): return PatternFill("solid", fgColor=h)
def font(color=WHITE, bold=False, size=9, italic=False):
    return Font(name="Arial", color=color, bold=bold, size=size, italic=italic)
def bdr():
    s = Side(style="thin", color="2D2D2D")
    return Border(left=s, right=s, top=s, bottom=s)
def mid(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def merge_fresh(pipeline):
    """Merge or fully replace pipeline sections from fresh JSON."""
    if not os.path.exists(FRESH_JSON):
        return pipeline
    try:
        with open(FRESH_JSON) as f:
            fresh = json.load(f)
        for section, entries in fresh.items():
            if section in pipeline:
                # OPEN_NOW + UPCOMING_2W + RECENT_30D get fully replaced
                if section in ("OPEN_NOW", "UPCOMING_2W", "RECENT_30D"):
                    pipeline[section] = entries
                else:
                    pipeline[section].update(entries)
            else:
                pipeline[section] = entries
        print(f"Merged fresh IPO pipeline data — {sum(len(v) for v in fresh.values())} entries")
    except Exception as e:
        print(f"Could not merge IPO fresh JSON: {e}")
    return pipeline

# ── Excel build ───────────────────────────────────────────────────────────────
OPEN_HEADERS = [
    "#", "Company", "Exchange", "Issue Dates", "List Date",
    "Size ₹cr", "Price Band ₹", "Lot", "Min Inv ₹",
    "QIB Sub", "HNI Sub", "Retail Sub", "Total Sub",
    "GMP ₹", "GMP %", "GMP Trend",
    "Anchor Quality", "Anchor ₹cr",
    "Rev FY26 cr", "PAT FY26 cr", "PE", "Sector PE",
    "Sector", "Promoter Post-IPO %",
    "Use of Proceeds", "Analyst Reco",
    "Exp List Gain %", "Rating", "Rationale",
]
OPEN_WIDTHS = [4,24,14,16,12,8,12,6,11,9,9,9,9,8,8,11,28,10,11,11,7,9,18,10,24,28,10,18,42]

UPCOMING_HEADERS = [
    "#", "Company", "Exchange", "Issue Dates", "Size ₹cr",
    "Price Band", "Anchor Date", "Early GMP", "DRHP Status",
    "Sector", "Analyst Preview", "Watch Signal",
]
UPCOMING_WIDTHS = [4,24,14,18,9,14,12,11,16,18,32,42]

RECENT_HEADERS = [
    "#", "Company", "Listed", "Issue ₹", "Listing ₹", "List Gain %",
    "Current ₹", "Post-List %", "Lessons / Pattern",
]
RECENT_WIDTHS = [4,24,12,9,10,11,9,11,48]

BIG_HEADERS = [
    "#", "Company", "Expected Q", "Size ₹cr", "Valuation ₹cr",
    "Anchor Demand", "Watch For", "Trade Idea", "Status",
]
BIG_WIDTHS = [4,28,14,10,12,18,28,38,18]

def add_section_block(ws, start_row, title, headers, widths):
    """Adds a section title row + header row, returns the row index for first data row."""
    cols = len(headers)
    last_col = get_column_letter(cols)
    ws.merge_cells(f"A{start_row}:{last_col}{start_row}")
    ws[f"A{start_row}"] = title
    ws[f"A{start_row}"].font = Font(name="Arial", color=GOLD, bold=True, size=11)
    ws[f"A{start_row}"].fill = fill(HEADER_BG)
    ws[f"A{start_row}"].alignment = mid()
    ws.row_dimensions[start_row].height = 22

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=start_row + 1, column=i, value=h)
        c.font = Font(name="Arial", color=GOLD, bold=True, size=8)
        c.fill = fill(HEADER_BG)
        c.alignment = mid()
        c.border = bdr()
        if ws.column_dimensions[get_column_letter(i)].width is None or \
           ws.column_dimensions[get_column_letter(i)].width < w:
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[start_row + 1].height = 30
    return start_row + 2

def build():
    pipeline = merge_fresh({k: dict(v) for k, v in IPO_PIPELINE.items()})

    wb = Workbook()
    ws = wb.active
    ws.title = "NSE IPO Pipeline"
    ws.sheet_view.showGridLines = False

    # ── Master title ──
    ws.merge_cells("A1:Z1")
    ws["A1"] = (f"NSE / BSE IPO LISTING-GAIN HUNTER  |  Updated {TODAY.strftime('%d %b %Y')}  |  "
                "Subscription · GMP · Anchor Quality · Analyst Rating-Driven Filter")
    ws["A1"].font = Font(name="Arial", color=GOLD, bold=True, size=13)
    ws["A1"].fill = fill(HEADER_BG)
    ws["A1"].alignment = mid()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:Z2")
    ws["A2"] = ("Strong setup = QIB ≥ 20x · HNI ≥ 50x · Retail ≥ 5x · GMP sustained ≥ 30% · "
                "Top MF anchors · Reasonable PE vs sector  |  Issue size sweet spot ₹500-3,000 cr")
    ws["A2"].font = Font(name="Arial", color=GREY, italic=True, size=9)
    ws["A2"].fill = fill(DARK_BG)
    ws["A2"].alignment = mid()
    ws.row_dimensions[2].height = 18

    cur = 4

    # ── OPEN_NOW section ──
    cur = add_section_block(ws, cur, "🟢 OPEN NOW — APPLY THIS WEEK",
                             OPEN_HEADERS, OPEN_WIDTHS)
    open_entries = pipeline.get("OPEN_NOW", {})
    if not open_entries:
        ws.merge_cells(f"A{cur}:{get_column_letter(len(OPEN_HEADERS))}{cur}")
        ws[f"A{cur}"] = "  (No open IPOs this week — bi-weekly Opus task will populate this section)"
        ws[f"A{cur}"].font = font(GREY, italic=True)
        ws[f"A{cur}"].fill = fill(DARK_BG)
        ws[f"A{cur}"].alignment = lft()
        ws.row_dimensions[cur].height = 18
        cur += 1
    else:
        for idx, (company, d) in enumerate(open_entries.items(), 1):
            r = cur
            rating = d.get("rating", "NEUTRAL")
            r_bg, r_fg, _ = RATING_META.get(rating, (DARK_BG, GREY, ""))
            row_bg = ROW_ALT if idx % 2 else DARK_BG
            cells = [
                idx, company, d.get("exchange","—"), d.get("issue_dates","—"), d.get("listing_date","—"),
                d.get("issue_size_cr",0), d.get("price_band","—"), d.get("lot_size",0), d.get("min_investment",0),
                d.get("qib_subscription","—"), d.get("hni_subscription","—"),
                d.get("retail_subscription","—"), d.get("total_subscription","—"),
                d.get("gmp_rs",0), f"{d.get('gmp_pct',0)}%", d.get("gmp_trend","—"),
                d.get("anchor_quality","—"), d.get("anchor_amount_cr",0),
                d.get("fy26_revenue_cr",0), d.get("fy26_pat_cr",0),
                d.get("pe_ratio",0), d.get("sector_pe",0),
                d.get("sector","—"), f"{d.get('promoter_post_ipo_pct',0)}%",
                d.get("use_of_proceeds","—"), d.get("analyst_recommendation","—"),
                f"{d.get('expected_listing_gain_pct',0)}%",
                rating, d.get("rationale","—"),
            ]
            for col_i, val in enumerate(cells, 1):
                c = ws.cell(row=r, column=col_i, value=val)
                c.fill = fill(row_bg)
                c.border = bdr()
                c.font = font(WHITE, size=9)
                c.alignment = mid() if col_i not in (2,17,23,25,26,29) else lft()
                if col_i == 2: c.font = font(GREEN, bold=True)
                if col_i == 14 or col_i == 15: c.font = font(GOLD, bold=True)
                if col_i == 28:
                    c.font = Font(name="Arial", color=r_fg, bold=True, size=9)
                    c.fill = fill(r_bg)
            ws.row_dimensions[r].height = 50
            cur += 1

    cur += 2  # spacer

    # ── UPCOMING_2W section ──
    cur = add_section_block(ws, cur, "🟡 UPCOMING (Next 2 Weeks) — Watch Anchor Book",
                             UPCOMING_HEADERS, UPCOMING_WIDTHS)
    upcoming = pipeline.get("UPCOMING_2W", {})
    if not upcoming:
        ws.merge_cells(f"A{cur}:{get_column_letter(len(UPCOMING_HEADERS))}{cur}")
        ws[f"A{cur}"] = "  (No upcoming IPOs identified — Opus task scans SEBI DRHP approvals)"
        ws[f"A{cur}"].font = font(GREY, italic=True)
        ws[f"A{cur}"].fill = fill(DARK_BG)
        ws[f"A{cur}"].alignment = lft()
        ws.row_dimensions[cur].height = 18
        cur += 1
    else:
        for idx, (company, d) in enumerate(upcoming.items(), 1):
            r = cur
            row_bg = ROW_ALT if idx % 2 else DARK_BG
            cells = [
                idx, company, d.get("exchange","—"), d.get("issue_dates","—"),
                d.get("issue_size_cr",0), d.get("price_band","—"),
                d.get("anchor_book_date","—"), d.get("early_gmp","—"),
                d.get("drhp_status","—"), d.get("sector","—"),
                d.get("analyst_preview","—"), d.get("watch","—"),
            ]
            for col_i, val in enumerate(cells, 1):
                c = ws.cell(row=r, column=col_i, value=val)
                c.fill = fill(row_bg)
                c.border = bdr()
                c.font = font(WHITE, size=9)
                c.alignment = mid() if col_i in (1,3,5,7,8) else lft()
                if col_i == 2: c.font = font(BLUE, bold=True)
            ws.row_dimensions[r].height = 40
            cur += 1

    cur += 2  # spacer

    # ── RECENT_30D section ──
    cur = add_section_block(ws, cur, "📈 RECENT LISTINGS (Last 30 Days) — Pattern Library",
                             RECENT_HEADERS, RECENT_WIDTHS)
    recent = pipeline.get("RECENT_30D", {})
    if not recent:
        ws.merge_cells(f"A{cur}:{get_column_letter(len(RECENT_HEADERS))}{cur}")
        ws[f"A{cur}"] = "  (No recent listings logged — Opus task pulls last-30-day mainboard listings)"
        ws[f"A{cur}"].font = font(GREY, italic=True)
        ws[f"A{cur}"].fill = fill(DARK_BG)
        ws[f"A{cur}"].alignment = lft()
        ws.row_dimensions[cur].height = 18
        cur += 1
    else:
        for idx, (company, d) in enumerate(recent.items(), 1):
            r = cur
            row_bg = ROW_ALT if idx % 2 else DARK_BG
            list_gain = d.get("listing_gain_pct", 0)
            post = d.get("post_listing_pct", 0)
            cells = [
                idx, company, d.get("listed_date","—"),
                d.get("issue_price",0), d.get("listing_price",0), f"{list_gain}%",
                d.get("current_price",0), f"{post}%", d.get("lessons","—"),
            ]
            for col_i, val in enumerate(cells, 1):
                c = ws.cell(row=r, column=col_i, value=val)
                c.fill = fill(row_bg)
                c.border = bdr()
                c.font = font(WHITE, size=9)
                c.alignment = mid() if col_i != 2 and col_i != 9 else lft()
                if col_i == 2: c.font = font(CYAN, bold=True)
                if col_i == 6:
                    c.font = font(GREEN if list_gain >= 0 else RED, bold=True)
                if col_i == 8:
                    c.font = font(GREEN if post >= 0 else RED, bold=True)
            ws.row_dimensions[r].height = 36
            cur += 1

    cur += 2  # spacer

    # ── BIG_PIPELINE section ──
    cur = add_section_block(ws, cur, "🚀 BIG PIPELINE 2026-2027 — Major IPOs to Track",
                             BIG_HEADERS, BIG_WIDTHS)
    big = pipeline.get("BIG_PIPELINE_2026", {})
    for idx, (company, d) in enumerate(big.items(), 1):
        r = cur
        row_bg = ROW_ALT if idx % 2 else DARK_BG
        cells = [
            idx, company, d.get("expected_quarter","—"),
            d.get("expected_size_cr",0), d.get("valuation_cr",0),
            d.get("anchor_demand","—"), d.get("watch_for","—"),
            d.get("trade_idea","—"), d.get("status","—"),
        ]
        for col_i, val in enumerate(cells, 1):
            c = ws.cell(row=r, column=col_i, value=val)
            c.fill = fill(row_bg)
            c.border = bdr()
            c.font = font(WHITE, size=9)
            c.alignment = mid() if col_i in (1,3,4,5) else lft()
            if col_i == 2: c.font = font(PURPLE, bold=True)
            if col_i == 9:
                clr = GREEN if "filed" in str(val).lower() or "listed" in str(val).lower() else AMBER
                c.font = font(clr, bold=True)
        ws.row_dimensions[r].height = 36
        cur += 1

    # ── Footer ──
    cur += 2
    ws.merge_cells(f"A{cur}:Z{cur}")
    ws[f"A{cur}"] = ("⚠️  IPO PROTOCOL: 1) ONLY apply if rating = STRONG SUBSCRIBE or SUBSCRIBE  "
                    "2) Apply Day 1 to maximize allocation chance  "
                    "3) GMP cut-off threshold ≥ ₹50 OR ≥ 30% before close  "
                    "4) Subscription cut-off: total ≥ 30x by Day 3 final hour  "
                    "5) Sell Day 1 listing if listing gain ≥ 25%, else hold for Day 5 momentum  "
                    "6) Never apply with borrowed funds  "
                    "7) ASBA only — never use UPI block + cash mix")
    ws[f"A{cur}"].font      = Font(name="Arial", color=AMBER, bold=True, size=9)
    ws[f"A{cur}"].fill      = fill(HEADER_BG)
    ws[f"A{cur}"].alignment = mid()
    ws.row_dimensions[cur].height = 50

    cur += 1
    ws.merge_cells(f"A{cur}:Z{cur}")
    ws[f"A{cur}"] = (f"Auto-generated by Claude Opus 4.7  |  Run date: {TODAY.strftime('%d %b %Y')}  |  "
                    "Sources: chittorgarh.com · investorgain.com · NSE/BSE filings · SEBI DRHP portal · "
                    "Motilal Oswal · ICICI Sec · Anand Rathi · Choice IPO notes  |  Verify all data.  NOT financial advice.")
    ws[f"A{cur}"].font      = Font(name="Arial", color=GREY, italic=True, size=8)
    ws[f"A{cur}"].fill      = fill(DARK_BG)
    ws[f"A{cur}"].alignment = mid()
    ws.row_dimensions[cur].height = 18

    ws.freeze_panes = "B4"

    wb.save(EXCEL_PATH)
    print(f"✅ IPO Pipeline Excel saved: {EXCEL_PATH}\n")

    # Summary
    print(f"📊 IPO Pipeline Summary:")
    print(f"  🟢 OPEN NOW            : {len(pipeline.get('OPEN_NOW', {}))} IPOs")
    print(f"  🟡 UPCOMING (2W)       : {len(pipeline.get('UPCOMING_2W', {}))} IPOs")
    print(f"  📈 RECENT (30 days)    : {len(pipeline.get('RECENT_30D', {}))} listings")
    print(f"  🚀 BIG PIPELINE 2026-27: {len(pipeline.get('BIG_PIPELINE_2026', {}))} mega-IPOs tracked")

if __name__ == "__main__":
    build()
